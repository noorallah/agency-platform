"""Transactional service for profile-driven product master operations."""

# ruff: noqa: D102, D107

from collections.abc import Iterable
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from sqlalchemy import String, case, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.sql import Select

from app.business.models import (
    BusinessProfile,
    CategoryAttributeRule,
    FirmBusinessProfile,
    ProfileFeature,
)
from app.business.services import AttributeInput, AttributeService
from app.common.audit.services import record_audit
from app.core.exceptions import ConflictError, ResourceNotFoundError, ValidationError
from app.core.utils.dates import utc_now
from app.firms.models import Firm
from app.products.models import (
    Product,
    ProductAttributeValue,
    ProductCategory,
    ProductMedia,
)
from app.products.schemas import (
    ProductAttributeInput,
    ProductAttributeResponse,
    ProductCategoryCreate,
    ProductCategoryFilter,
    ProductCategoryUpdate,
    ProductCreate,
    ProductListFilters,
    ProductMediaInput,
    ProductMetadataResponse,
    ProductSummary,
    ProductTaxProfileOption,
    ProductUpdate,
)
from app.products.schemas.product import ProductCategoryResponse
from app.tax.models import TaxProfile
from app.uom.models import Uom


class ProductService:
    """Coordinate dynamic product validation, persistence, and retrieval."""

    def __init__(self, session: Session) -> None:
        self._session = session
        # Scoped to this request; the service is constructed per call.
        self._attribute_match_cache: dict[tuple[UUID, str], frozenset[UUID]] = {}

    def list_products(
        self,
        *,
        firm_scope: UUID,
        filters: ProductListFilters,
        page: int,
        page_size: int,
        search: str | None,
        sort_by: str,
        descending: bool,
    ) -> tuple[list[Product], int]:
        columns = {
            "code": Product.code,
            "name": Product.name,
            "status": Product.status,
            "selling_price": Product.selling_price,
            "created_at": Product.created_at,
        }
        statement = (
            select(Product)
            .where(Product.firm_id == firm_scope)
            .options(selectinload(Product.media))
        )
        count = (
            select(func.count())
            .select_from(Product)
            .where(Product.firm_id == firm_scope)
        )
        statement, count = self._apply_filters(statement, count, filters=filters)
        ordering = columns[sort_by].desc() if descending else columns[sort_by].asc()
        rows = list(self._session.scalars(statement.order_by(ordering)).all())
        if search or filters.attribute_query:
            term = f"%{search.strip()}%" if search else None
            attr_term = (
                f"%{filters.attribute_query.strip()}%"
                if filters.attribute_query
                else None
            )
            filtered = [
                row
                for row in rows
                if self._matches_search(
                    row=row, search_term=term, attribute_term=attr_term
                )
            ]
            total = len(filtered)
            start = (page - 1) * page_size
            return filtered[start : start + page_size], total
        start = (page - 1) * page_size
        return rows[start : start + page_size], int(self._session.scalar(count) or 0)

    def summary(
        self, *, firm_scope: UUID, filters: ProductListFilters
    ) -> ProductSummary:
        base = select(Product).where(Product.firm_id == firm_scope)
        base, _ = self._apply_filters(base, base, filters=filters)
        subquery = base.subquery()
        total, active, inactive, draft, archived, deleted = self._session.execute(
            select(
                func.count(subquery.c.id),
                func.sum(case((subquery.c.status == "ACTIVE", 1), else_=0)),
                func.sum(case((subquery.c.status == "INACTIVE", 1), else_=0)),
                func.sum(case((subquery.c.status == "DRAFT", 1), else_=0)),
                func.sum(case((subquery.c.status == "ARCHIVED", 1), else_=0)),
                func.sum(case((subquery.c.is_deleted.is_(True), 1), else_=0)),
            )
        ).one()
        return ProductSummary(
            total=int(total or 0),
            active=int(active or 0),
            inactive=int(inactive or 0),
            draft=int(draft or 0),
            archived=int(archived or 0),
            deleted=int(deleted or 0),
        )

    def create_product(
        self, data: ProductCreate, *, firm_id: UUID, actor_id: UUID
    ) -> Product:
        self._assert_unique_code(firm_id, data.code)
        self._assert_unique_barcode(firm_id, data.barcode)
        feature_codes = self._active_feature_codes(firm_id)
        category = self._validate_category_reference(firm_id, data.category_id)
        self._validate_sub_category_reference(
            firm_id=firm_id,
            category_id=data.category_id,
            sub_category_id=data.sub_category_id,
        )
        self._validate_tax_profile_group_code(firm_id, data.tax_profile_group_code)
        self._validate_uom_references(data)
        self._validate_feature_gated_fields(data, feature_codes)
        product = Product(
            **self._product_values(data),
            firm_id=firm_id,
            created_by=actor_id,
            updated_by=actor_id,
        )
        product.media = [
            self._build_media(firm_id, item, actor_id) for item in data.media
        ]
        self._session.add(product)
        self._session.flush()
        self._store_attributes(
            product, data.attributes, category=category, actor_id=actor_id
        )
        record_audit(
            self._session,
            action="product.created",
            entity_type="product",
            entity_id=product.id,
            actor_id=actor_id,
            firm_id=firm_id,
            after_data={"code": product.code},
        )
        self._commit()
        self._session.refresh(product)
        return product

    def get_product(
        self,
        product_id: UUID,
        *,
        firm_scope: UUID,
        include_deleted: bool = False,
    ) -> Product:
        statement = (
            select(Product)
            .where(Product.id == product_id, Product.firm_id == firm_scope)
            .options(selectinload(Product.media))
        )
        if not include_deleted:
            statement = statement.where(Product.is_deleted.is_(False))
        row = self._session.scalar(statement)
        if row is None:
            raise ResourceNotFoundError("Product not found.")
        return row

    def update_product(
        self,
        product_id: UUID,
        data: ProductUpdate,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> Product:
        product = self.get_product(
            product_id, firm_scope=firm_scope, include_deleted=True
        )
        self._assert_unique_code(firm_scope, data.code, current_id=product.id)
        self._assert_unique_barcode(firm_scope, data.barcode, current_id=product.id)
        category = self._validate_category_reference(firm_scope, data.category_id)
        self._validate_sub_category_reference(
            firm_id=firm_scope,
            category_id=data.category_id,
            sub_category_id=data.sub_category_id,
        )
        self._validate_tax_profile_group_code(firm_scope, data.tax_profile_group_code)
        self._validate_uom_references(data)
        feature_codes = self._active_feature_codes(firm_scope)
        self._validate_feature_gated_fields(data, feature_codes)
        before: dict[str, object] = {
            "code": product.code,
            "category_id": str(product.category_id),
        }
        for field, value in self._product_values(data).items():
            setattr(product, field, value)
        product.updated_by = actor_id
        self._store_attributes(
            product, data.attributes, category=category, actor_id=actor_id
        )
        self._reconcile_media(product, data.media, actor_id)
        record_audit(
            self._session,
            action="product.updated",
            entity_type="product",
            entity_id=product.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data=before,
            after_data={"code": product.code, "status": product.status},
        )
        self._commit()
        self._session.refresh(product)
        return product

    def delete_product(
        self, product_id: UUID, *, firm_scope: UUID, actor_id: UUID
    ) -> None:
        product = self.get_product(product_id, firm_scope=firm_scope)
        product.is_deleted = True
        product.deleted_at = utc_now()
        product.deleted_by = actor_id
        product.updated_by = actor_id
        record_audit(
            self._session,
            action="product.deleted",
            entity_type="product",
            entity_id=product.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data={"code": product.code},
        )
        self._commit()

    def restore_product(
        self, product_id: UUID, *, firm_scope: UUID, actor_id: UUID
    ) -> Product:
        product = self.get_product(
            product_id, firm_scope=firm_scope, include_deleted=True
        )
        if not product.is_deleted:
            return product
        product.is_deleted = False
        product.deleted_at = None
        product.deleted_by = None
        product.updated_by = actor_id
        record_audit(
            self._session,
            action="product.restored",
            entity_type="product",
            entity_id=product.id,
            actor_id=actor_id,
            firm_id=firm_scope,
        )
        self._commit()
        return product

    def duplicate_product(
        self, product_id: UUID, *, firm_scope: UUID, actor_id: UUID
    ) -> Product:
        source = self.get_product(product_id, firm_scope=firm_scope)
        duplicated = ProductCreate.model_validate(
            {
                **self._product_values_from_model(source),
                "code": self._next_duplicate_code(firm_scope, source.code),
                "attributes": self._attribute_inputs_for(source),
                "media": [
                    self._media_input_from_model(media)
                    for media in source.media
                    if not media.is_deleted
                ],
            }
        )
        product = self.create_product(duplicated, firm_id=firm_scope, actor_id=actor_id)
        record_audit(
            self._session,
            action="product.duplicated",
            entity_type="product",
            entity_id=product.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data={"source_product_id": str(source.id)},
        )
        self._commit()
        return product

    def bulk_delete(
        self, ids: Iterable[UUID], *, firm_scope: UUID, actor_id: UUID
    ) -> int:
        count = 0
        for product_id in ids:
            product = self.get_product(product_id, firm_scope=firm_scope)
            if product.is_deleted:
                continue
            product.is_deleted = True
            product.deleted_at = utc_now()
            product.deleted_by = actor_id
            product.updated_by = actor_id
            count += 1
        if count > 0:
            self._commit()
        return count

    def bulk_restore(
        self, ids: Iterable[UUID], *, firm_scope: UUID, actor_id: UUID
    ) -> int:
        count = 0
        for product_id in ids:
            product = self.get_product(
                product_id, firm_scope=firm_scope, include_deleted=True
            )
            if not product.is_deleted:
                continue
            product.is_deleted = False
            product.deleted_at = None
            product.deleted_by = None
            product.updated_by = actor_id
            count += 1
        if count > 0:
            self._commit()
        return count

    def metadata(
        self, *, firm_scope: UUID, category_id: UUID | None = None
    ) -> ProductMetadataResponse:
        profile = self._resolved_profile(firm_scope)
        feature_codes = self._active_feature_codes(firm_scope)
        categories = self._session.scalars(
            select(ProductCategory)
            .where(
                ProductCategory.firm_id == firm_scope,
                ProductCategory.is_deleted.is_(False),
                ProductCategory.is_active.is_(True),
            )
            .order_by(ProductCategory.path.asc())
        ).all()
        required_ids, optional_ids = self._category_attribute_ids(
            profile.id, category_id
        )
        return ProductMetadataResponse(
            profile_code=profile.code,
            features=[
                {"code": code, "enabled": True} for code in sorted(feature_codes)
            ],
            categories=[
                ProductCategoryResponse.model_validate(item) for item in categories
            ],
            tax_profiles=[
                ProductTaxProfileOption(
                    id=item.id,
                    code=item.code,
                    group_code=item.group_code,
                    label=item.label,
                    tax_system_id=item.tax_system_id,
                )
                for item in self._session.scalars(
                    select(TaxProfile)
                    .where(
                        TaxProfile.firm_id == firm_scope,
                        TaxProfile.is_deleted.is_(False),
                        TaxProfile.status == "ACTIVE",
                    )
                    .order_by(TaxProfile.display_order.asc(), TaxProfile.code.asc())
                ).all()
            ],
            required_attribute_definition_ids=required_ids,
            optional_attribute_definition_ids=optional_ids,
        )

    def create_category(
        self, data: ProductCategoryCreate, *, firm_id: UUID, actor_id: UUID
    ) -> ProductCategory:
        parent = self._validate_category_reference(firm_id, data.parent_id)
        level = 0 if parent is None else parent.level + 1
        path = data.code if parent is None else f"{parent.path}/{data.code}"
        row = ProductCategory(
            firm_id=firm_id,
            code=data.code,
            name=data.name,
            parent_id=data.parent_id,
            level=level,
            path=path,
            is_active=data.is_active,
            created_by=actor_id,
            updated_by=actor_id,
        )
        self._session.add(row)
        self._session.flush()
        self._commit()
        return row

    def list_categories(
        self, *, firm_scope: UUID, filters: ProductCategoryFilter
    ) -> list[ProductCategory]:
        statement = select(ProductCategory).where(
            ProductCategory.firm_id == firm_scope,
            ProductCategory.is_deleted.is_(False),
        )
        if filters.parent_id is not None:
            statement = statement.where(ProductCategory.parent_id == filters.parent_id)
        if not filters.include_inactive:
            statement = statement.where(ProductCategory.is_active.is_(True))
        return list(
            self._session.scalars(statement.order_by(ProductCategory.path.asc())).all()
        )

    def get_category(self, category_id: UUID, *, firm_scope: UUID) -> ProductCategory:
        row = self._validate_category_reference(firm_scope, category_id)
        if row is None:
            raise ResourceNotFoundError("Product category not found.")
        return row

    def update_category(
        self,
        category_id: UUID,
        data: ProductCategoryUpdate,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> ProductCategory:
        row = self.get_category(category_id, firm_scope=firm_scope)
        parent = self._validate_category_reference(firm_scope, data.parent_id)
        row.code = data.code
        row.name = data.name
        row.parent_id = data.parent_id
        row.level = 0 if parent is None else parent.level + 1
        row.path = data.code if parent is None else f"{parent.path}/{data.code}"
        row.is_active = data.is_active
        row.updated_by = actor_id
        self._commit()
        return row

    def delete_category(
        self, category_id: UUID, *, firm_scope: UUID, actor_id: UUID
    ) -> None:
        row = self.get_category(category_id, firm_scope=firm_scope)
        has_products = self._session.scalar(
            select(Product.id).where(
                Product.firm_id == firm_scope,
                Product.category_id == row.id,
                Product.is_deleted.is_(False),
            )
        )
        if has_products is not None:
            raise ValidationError("Categories used by products cannot be deleted.")
        row.is_deleted = True
        row.deleted_at = utc_now()
        row.deleted_by = actor_id
        row.updated_by = actor_id
        self._commit()

    def export_products_csv(self, *, firm_scope: UUID, search: str | None) -> str:
        rows, _ = self.list_products(
            firm_scope=firm_scope,
            filters=ProductListFilters(include_deleted=False),
            page=1,
            page_size=5000,
            search=search,
            sort_by="code",
            descending=False,
        )
        output = ["Code,Name,Type,Brand,HSN,SellingPrice,Status"]
        for item in rows:
            output.append(
                ",".join(
                    [
                        item.code,
                        item.name,
                        item.product_type,
                        item.brand or "",
                        item.hsn_sac or "",
                        str(item.selling_price or ""),
                        item.status,
                    ]
                )
            )
        return "\n".join(output)

    def export_products_xlsx(self, *, firm_scope: UUID, search: str | None) -> bytes:
        try:
            from openpyxl import Workbook  # type: ignore[import-untyped]
        except ImportError as error:
            raise ValidationError(
                "XLSX export dependency is unavailable. Install openpyxl."
            ) from error
        rows, _ = self.list_products(
            firm_scope=firm_scope,
            filters=ProductListFilters(include_deleted=False),
            page=1,
            page_size=5000,
            search=search,
            sort_by="code",
            descending=False,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        sheet.append(["Code", "Name", "Type", "Brand", "HSN", "SellingPrice", "Status"])
        for item in rows:
            sheet.append(
                [
                    item.code,
                    item.name,
                    item.product_type,
                    item.brand or "",
                    item.hsn_sac or "",
                    float(item.selling_price or 0),
                    item.status,
                ]
            )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def import_products_json(
        self, records: list[ProductCreate], *, firm_scope: UUID, actor_id: UUID
    ) -> list[Product]:
        result: list[Product] = []
        for item in records:
            result.append(
                self.create_product(item, firm_id=firm_scope, actor_id=actor_id)
            )
        return result

    def import_products_csv(
        self, csv_content: str, *, firm_scope: UUID, actor_id: UUID
    ) -> list[Product]:
        import csv
        import io

        reader = csv.DictReader(io.StringIO(csv_content))
        records: list[ProductCreate] = []
        for row in reader:
            code = (row.get("Code") or "").strip().upper()
            if not code:
                continue
            records.append(
                ProductCreate(
                    code=code,
                    barcode=None,
                    qr_code=None,
                    name=(row.get("Name") or "").strip(),
                    short_name=None,
                    description=None,
                    product_type=(row.get("Type") or "STOCK_ITEM").strip().upper(),
                    category_id=None,
                    sub_category_id=None,
                    unit=None,
                    brand=(row.get("Brand") or "").strip() or None,
                    model=None,
                    hsn_sac=(row.get("HSN") or "").strip().upper() or None,
                    tax_profile_group_code=None,
                    purchase_price=None,
                    selling_price=(
                        Decimal(row["SellingPrice"])
                        if row.get("SellingPrice")
                        else None
                    ),
                    mrp=None,
                    status=(row.get("Status") or "ACTIVE").strip().upper(),
                    remarks=None,
                    attributes=[],
                    media=[],
                )
            )
        return self.import_products_json(
            records, firm_scope=firm_scope, actor_id=actor_id
        )

    def import_products_xlsx(
        self, workbook_bytes: bytes, *, firm_scope: UUID, actor_id: UUID
    ) -> list[Product]:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ValidationError(
                "XLSX import dependency is unavailable. Install openpyxl."
            ) from error
        workbook = load_workbook(filename=BytesIO(workbook_bytes), read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(value or "").strip() for value in rows[0]]
        index = {name: position for position, name in enumerate(header)}
        records: list[ProductCreate] = []
        for values in rows[1:]:
            code = str(values[index.get("Code", -1)] or "").strip().upper()
            if not code:
                continue
            records.append(
                ProductCreate(
                    code=code,
                    barcode=None,
                    qr_code=None,
                    name=str(values[index.get("Name", -1)] or "").strip(),
                    short_name=None,
                    description=None,
                    product_type=str(values[index.get("Type", -1)] or "STOCK_ITEM")
                    .strip()
                    .upper(),
                    category_id=None,
                    sub_category_id=None,
                    unit=None,
                    brand=str(values[index.get("Brand", -1)] or "").strip() or None,
                    model=None,
                    hsn_sac=str(values[index.get("HSN", -1)] or "").strip().upper()
                    or None,
                    tax_profile_group_code=None,
                    purchase_price=None,
                    selling_price=(
                        Decimal(str(values[index.get("SellingPrice", -1)]).strip())
                        if index.get("SellingPrice", -1) >= 0
                        and values[index["SellingPrice"]] is not None
                        else None
                    ),
                    mrp=None,
                    status=str(values[index.get("Status", -1)] or "ACTIVE")
                    .strip()
                    .upper(),
                    remarks=None,
                    attributes=[],
                    media=[],
                )
            )
        return self.import_products_json(
            records, firm_scope=firm_scope, actor_id=actor_id
        )

    def _apply_filters(
        self,
        statement: Select[Any],
        count: Select[Any],
        *,
        filters: ProductListFilters,
    ) -> tuple[Select[Any], Select[Any]]:
        if not filters.include_deleted:
            statement = statement.where(Product.is_deleted.is_(False))
            count = count.where(Product.is_deleted.is_(False))
        if filters.status is not None:
            statement = statement.where(Product.status == filters.status.value)
            count = count.where(Product.status == filters.status.value)
        if filters.product_type is not None:
            statement = statement.where(
                Product.product_type == filters.product_type.value
            )
            count = count.where(Product.product_type == filters.product_type.value)
        if filters.category_id is not None:
            statement = statement.where(Product.category_id == filters.category_id)
            count = count.where(Product.category_id == filters.category_id)
        if filters.sub_category_id is not None:
            statement = statement.where(
                Product.sub_category_id == filters.sub_category_id
            )
            count = count.where(Product.sub_category_id == filters.sub_category_id)
        if filters.tax_profile_group_code is not None:
            statement = statement.where(
                Product.tax_profile_group_code == filters.tax_profile_group_code
            )
            count = count.where(
                Product.tax_profile_group_code == filters.tax_profile_group_code
            )
        if filters.brand:
            statement = statement.where(
                Product.brand.ilike(f"%{filters.brand.strip()}%")
            )
            count = count.where(Product.brand.ilike(f"%{filters.brand.strip()}%"))
        if filters.hsn_sac:
            statement = statement.where(
                Product.hsn_sac == filters.hsn_sac.strip().upper()
            )
            count = count.where(Product.hsn_sac == filters.hsn_sac.strip().upper())
        return statement, count

    def _active_feature_codes(self, firm_id: UUID) -> set[str]:
        profile = self._resolved_profile(firm_id)
        feature_ids = {
            row.feature_id
            for row in self._session.scalars(
                select(ProfileFeature).where(
                    ProfileFeature.business_profile_id == profile.id,
                    ProfileFeature.is_deleted.is_(False),
                    ProfileFeature.is_enabled.is_(True),
                )
            )
        }
        if not feature_ids:
            return set()
        # Import lazily to avoid cyclic imports.
        from app.business.models import BusinessFeature

        return {
            row.code
            for row in self._session.scalars(
                select(BusinessFeature).where(BusinessFeature.id.in_(feature_ids))
            )
        }

    def _resolved_profile(self, firm_id: UUID) -> BusinessProfile:
        assignment = self._session.scalar(
            select(FirmBusinessProfile).where(
                FirmBusinessProfile.firm_id == firm_id,
                FirmBusinessProfile.is_deleted.is_(False),
                FirmBusinessProfile.is_active.is_(True),
            )
        )
        if assignment is not None:
            profile = self._session.scalar(
                select(BusinessProfile).where(
                    BusinessProfile.id == assignment.business_profile_id,
                    BusinessProfile.is_deleted.is_(False),
                    BusinessProfile.status == "ACTIVE",
                )
            )
            if profile is not None:
                return profile
        default_profile = self._session.scalar(
            select(BusinessProfile).where(
                BusinessProfile.is_deleted.is_(False),
                BusinessProfile.status == "ACTIVE",
                BusinessProfile.is_default.is_(True),
            )
        )
        if default_profile is None:
            raise ValidationError("No active business profile is configured.")
        return default_profile

    def _category_attribute_ids(
        self, profile_id: UUID, category_id: UUID | None
    ) -> tuple[list[UUID], list[UUID]]:
        if category_id is None:
            return [], []
        category = self._session.scalar(
            select(ProductCategory).where(ProductCategory.id == category_id)
        )
        if category is None:
            return [], []
        rules = self._session.scalars(
            select(CategoryAttributeRule).where(
                CategoryAttributeRule.is_deleted.is_(False),
                CategoryAttributeRule.category_code.in_(
                    [category.code, category.name.upper()]
                ),
                or_(
                    CategoryAttributeRule.business_profile_id == profile_id,
                    CategoryAttributeRule.business_profile_id.is_(None),
                ),
            )
        ).all()
        required = [rule.attribute_definition_id for rule in rules if rule.is_mandatory]
        optional = [
            rule.attribute_definition_id for rule in rules if not rule.is_mandatory
        ]
        return required, optional

    def _validate_feature_gated_fields(
        self, data: ProductCreate | ProductUpdate, feature_codes: set[str]
    ) -> None:
        if data.barcode and "BARCODE" not in feature_codes:
            raise ValidationError("Barcode is disabled by feature configuration.")
        if data.qr_code and "QR_CODE" not in feature_codes:
            raise ValidationError("QR code is disabled by feature configuration.")

    def _attribute_inputs_for(self, product: Product) -> list[dict[str, object]]:
        """Return a product's attributes shaped for ProductCreate validation."""
        return [
            {
                "attribute_definition_id": resolved.definition.id,
                "value": resolved.value,
            }
            for resolved in AttributeService(self._session).values_for(
                ProductAttributeValue, product.id
            )
            if resolved.value is not None
        ]

    def _store_attributes(
        self,
        product: Product,
        attributes: list[ProductAttributeInput],
        *,
        category: ProductCategory | None,
        actor_id: UUID,
    ) -> None:
        """Validate and persist a product's configurable attributes."""
        AttributeService(self._session).replace_values(
            ProductAttributeValue,
            product.id,
            [
                AttributeInput(
                    attribute_definition_id=item.attribute_definition_id,
                    value=item.value,
                )
                for item in attributes
            ],
            firm_id=product.firm_id,
            actor_id=actor_id,
            category_code=category.code if category is not None else None,
        )

    def attribute_responses(self, product: Product) -> list[ProductAttributeResponse]:
        """Return one product's stored attributes in response shape."""
        rows = self._session.scalars(
            select(ProductAttributeValue)
            .where(
                ProductAttributeValue.product_id == product.id,
                ProductAttributeValue.is_deleted.is_(False),
            )
            .order_by(ProductAttributeValue.created_at.asc())
        ).all()
        return [
            ProductAttributeResponse(
                id=row.id,
                attribute_definition_id=row.attribute_definition_id,
                value_text=row.value_text,
                value_number=row.value_number,
                value_date=row.value_date,
                value_boolean=row.value_boolean,
                created_at=row.created_at,
                updated_at=row.updated_at,
            )
            for row in rows
        ]

    def _assert_unique_code(
        self, firm_id: UUID, code: str, current_id: UUID | None = None
    ) -> None:
        statement = select(Product.id).where(
            Product.firm_id == firm_id,
            Product.code == code,
            Product.is_deleted.is_(False),
        )
        if current_id is not None:
            statement = statement.where(Product.id != current_id)
        if self._session.scalar(statement) is not None:
            raise ConflictError("Product code already exists in this firm.")

    def _assert_unique_barcode(
        self, firm_id: UUID, barcode: str | None, current_id: UUID | None = None
    ) -> None:
        normalized = (barcode or "").strip()
        if not normalized:
            return
        statement = select(Product.id).where(
            Product.firm_id == firm_id,
            Product.barcode == normalized,
            Product.is_deleted.is_(False),
        )
        if current_id is not None:
            statement = statement.where(Product.id != current_id)
        if self._session.scalar(statement) is not None:
            raise ConflictError("Barcode already exists in this firm.")

    def _validate_category_reference(
        self, firm_id: UUID, category_id: UUID | None
    ) -> ProductCategory | None:
        if category_id is None:
            return None
        row = self._session.scalar(
            select(ProductCategory).where(
                ProductCategory.id == category_id,
                ProductCategory.firm_id == firm_id,
                ProductCategory.is_deleted.is_(False),
            )
        )
        if row is None:
            raise ValidationError("Selected product category is unavailable.")
        return row

    def _validate_sub_category_reference(
        self,
        *,
        firm_id: UUID,
        category_id: UUID | None,
        sub_category_id: UUID | None,
    ) -> None:
        if sub_category_id is None:
            return
        if category_id is None:
            raise ValidationError("Category is required when sub category is selected.")
        sub_category = self._session.scalar(
            select(ProductCategory).where(
                ProductCategory.id == sub_category_id,
                ProductCategory.firm_id == firm_id,
                ProductCategory.is_deleted.is_(False),
            )
        )
        if sub_category is None:
            raise ValidationError("Selected sub category is unavailable.")
        if sub_category.parent_id != category_id:
            raise ValidationError(
                "Selected sub category does not belong to the selected category."
            )

    def _validate_tax_profile_group_code(
        self, firm_id: UUID, tax_profile_group_code: str | None
    ) -> None:
        if tax_profile_group_code is None:
            return
        row = self._session.scalar(
            select(TaxProfile.id).where(
                TaxProfile.firm_id == firm_id,
                TaxProfile.group_code == tax_profile_group_code,
                TaxProfile.is_deleted.is_(False),
                TaxProfile.status == "ACTIVE",
            )
        )
        if row is None:
            raise ValidationError(
                "No active tax profile found for the given group code."
            )

    def _validate_uom_references(self, data: ProductCreate | ProductUpdate) -> None:
        references = {
            data.base_uom_id,
            data.inventory_uom_id,
            data.purchase_uom_id,
            data.sales_uom_id,
            data.default_receiving_uom_id,
            data.default_dispatch_uom_id,
            data.minimum_sales_uom_id,
        }
        ids = {item for item in references if item is not None}
        if not ids:
            return
        available = set(
            self._session.scalars(
                select(Uom.id).where(Uom.id.in_(ids), Uom.is_deleted.is_(False))
            ).all()
        )
        missing = [item for item in ids if item not in available]
        if missing:
            raise ValidationError(
                "One or more selected UOM references are unavailable.",
                details={"unknown_uom_ids": [str(item) for item in missing]},
            )

    @staticmethod
    def _product_values(data: ProductCreate | ProductUpdate) -> dict[str, object]:
        payload = data.model_dump(exclude={"attributes", "media"}, mode="python")
        payload["product_type"] = data.product_type.value
        payload["status"] = data.status.value
        return payload

    @staticmethod
    def _build_media(
        firm_id: UUID, data: ProductMediaInput, actor_id: UUID
    ) -> ProductMedia:
        return ProductMedia(
            firm_id=firm_id,
            media_kind=data.media_kind,
            file_name=data.file_name,
            mime_type=data.mime_type,
            storage_path=data.storage_path,
            is_primary=data.is_primary,
            file_size_bytes=data.file_size_bytes,
            created_by=actor_id,
            updated_by=actor_id,
        )

    def _matches_search(
        self, *, row: Product, search_term: str | None, attribute_term: str | None
    ) -> bool:
        """Return whether a product matches the free-text or attribute search."""
        if search_term:
            search_text = search_term.strip("%").lower()
            haystacks = [
                row.code,
                row.barcode,
                row.qr_code,
                row.name,
                row.short_name,
                row.brand,
                row.hsn_sac,
            ]
            if any(
                value is not None and value.lower().find(search_text) >= 0
                for value in haystacks
                if value is not None
            ):
                return True
        if attribute_term:
            attr_text = attribute_term.strip("%").lower()
            if row.id in self._products_matching_attribute(row.firm_id, attr_text):
                return True
        return False

    def _products_matching_attribute(
        self, firm_id: UUID, needle: str
    ) -> frozenset[UUID]:
        """Return every product in a firm whose attributes contain the text.

        Resolved once per search rather than per candidate row: the previous
        per-row lookup issued one query for every product being filtered.
        """
        cached = self._attribute_match_cache.get((firm_id, needle))
        if cached is not None:
            return cached
        pattern = f"%{needle}%"
        rows = self._session.scalars(
            select(ProductAttributeValue.product_id).where(
                ProductAttributeValue.firm_id == firm_id,
                ProductAttributeValue.is_deleted.is_(False),
                or_(
                    func.lower(ProductAttributeValue.value_text).like(pattern),
                    func.lower(
                        func.cast(ProductAttributeValue.value_number, String)
                    ).like(pattern),
                    func.lower(
                        func.cast(ProductAttributeValue.value_date, String)
                    ).like(pattern),
                ),
            )
        ).all()
        matched = frozenset(rows)
        self._attribute_match_cache[(firm_id, needle)] = matched
        return matched

    def _reconcile_media(
        self, product: Product, inputs: list[ProductMediaInput], actor_id: UUID
    ) -> None:
        existing = {
            (row.media_kind, row.file_name, row.storage_path): row
            for row in product.media
        }
        requested_keys = {
            (item.media_kind, item.file_name, item.storage_path) for item in inputs
        }
        now = utc_now()
        for item in inputs:
            key = (item.media_kind, item.file_name, item.storage_path)
            current = existing.get(key)
            if current is None:
                product.media.append(self._build_media(product.firm_id, item, actor_id))
                continue
            current.mime_type = item.mime_type
            current.is_primary = item.is_primary
            current.file_size_bytes = item.file_size_bytes
            current.updated_by = actor_id
            current.is_deleted = False
            current.deleted_at = None
            current.deleted_by = None
        for key, row in existing.items():
            if key not in requested_keys:
                row.is_deleted = True
                row.deleted_at = now
                row.deleted_by = actor_id
                row.updated_by = actor_id

    def _product_values_from_model(self, product: Product) -> dict[str, object]:
        return {
            "code": product.code,
            "barcode": product.barcode,
            "qr_code": product.qr_code,
            "name": product.name,
            "short_name": product.short_name,
            "description": product.description,
            "product_type": product.product_type,
            "category_id": product.category_id,
            "sub_category_id": product.sub_category_id,
            "unit": product.unit,
            "brand": product.brand,
            "model": product.model,
            "hsn_sac": product.hsn_sac,
            "tax_profile_group_code": product.tax_profile_group_code,
            "base_uom_id": product.base_uom_id,
            "inventory_uom_id": product.inventory_uom_id,
            "purchase_uom_id": product.purchase_uom_id,
            "sales_uom_id": product.sales_uom_id,
            "default_receiving_uom_id": product.default_receiving_uom_id,
            "default_dispatch_uom_id": product.default_dispatch_uom_id,
            "minimum_sales_uom_id": product.minimum_sales_uom_id,
            "weight": product.weight,
            "volume": product.volume,
            "length": product.length,
            "width": product.width,
            "height": product.height,
            "allow_fraction": product.allow_fraction,
            "allow_decimal": product.allow_decimal,
            "purchase_price": product.purchase_price,
            "selling_price": product.selling_price,
            "mrp": product.mrp,
            "status": product.status,
            "remarks": product.remarks,
            "track_batch": product.track_batch,
            "track_lot": product.track_lot,
            "track_serial": product.track_serial,
            "track_expiry": product.track_expiry,
            "track_manufacturing_date": product.track_manufacturing_date,
            "track_warranty": product.track_warranty,
            "allow_negative_stock": product.allow_negative_stock,
            "require_batch_on_receipt": product.require_batch_on_receipt,
            "require_batch_on_issue": product.require_batch_on_issue,
            "require_serial_on_receipt": product.require_serial_on_receipt,
            "require_serial_on_issue": product.require_serial_on_issue,
        }

    @staticmethod
    def _media_input_from_model(row: ProductMedia) -> dict[str, object]:
        return {
            "media_kind": row.media_kind,
            "file_name": row.file_name,
            "mime_type": row.mime_type,
            "storage_path": row.storage_path,
            "is_primary": row.is_primary,
            "file_size_bytes": row.file_size_bytes,
        }

    def _next_duplicate_code(self, firm_id: UUID, code: str) -> str:
        base = f"{code}-COPY"
        candidate = base
        index = 1
        while (
            self._session.scalar(
                select(Product.id).where(
                    Product.firm_id == firm_id,
                    Product.code == candidate,
                    Product.is_deleted.is_(False),
                )
            )
            is not None
        ):
            candidate = f"{base}-{index}"
            index += 1
        return candidate

    def _commit(self) -> None:
        try:
            self._session.commit()
        except IntegrityError as error:
            self._session.rollback()
            raise ConflictError(
                "Product operation conflicts with existing data."
            ) from error

    def ensure_firm(self, firm_id: UUID) -> None:
        row = self._session.scalar(
            select(Firm.id).where(Firm.id == firm_id, Firm.is_deleted.is_(False))
        )
        if row is None:
            raise ResourceNotFoundError("Firm not found.")
