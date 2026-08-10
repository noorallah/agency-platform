"""Transactional service for enterprise purchase management."""

from __future__ import annotations

import csv
import io
import json
from datetime import date
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from sqlalchemy import false, func, or_, select
from sqlalchemy.orm import Session

from app.batch_serial.models import BatchRecord
from app.branches.models import Branch, Warehouse, WarehouseStorageNode
from app.business.gating import assert_feature_fields
from app.common.audit.services import record_audit
from app.core.exceptions import ConflictError, ResourceNotFoundError, ValidationError
from app.core.utils.dates import utc_now
from app.document_framework.models import (
    DocumentTypeDefinition,
)
from app.document_framework.schemas import (
    DocumentLifecycleEventCreate,
)
from app.document_framework.services.transactional_document_service import (
    DocumentStateSpec,
    DocumentTypeSpec,
    TransactionalDocumentService,
)
from app.goods_receipt.models import GoodsReceipt
from app.products.models import Product
from app.purchase.models import (
    PurchaseAttachment,
    PurchaseDeliverySchedule,
    PurchaseNote,
    PurchaseOrder,
    PurchaseOrderHistory,
    PurchaseOrderLine,
)
from app.purchase.schemas import (
    PurchaseAttachmentResponse,
    PurchaseDeliveryScheduleResponse,
    PurchaseNoteResponse,
    PurchaseOrderCreate,
    PurchaseOrderImportRequest,
    PurchaseOrderLineResponse,
    PurchaseOrderListFilters,
    PurchaseOrderResponse,
    PurchaseOrderStatus,
    PurchaseOrderUpdate,
    PurchaseSummary,
)
from app.tax.models import TaxProfile
from app.tax.schemas import TaxRuleSimulationRequest
from app.tax.services.tax_framework_service import TaxFrameworkService
from app.tax.services.tax_rule_service import TaxRuleService
from app.uom.schemas import ConversionRequest
from app.uom.services import UomService
from app.vendors.models import Vendor


def _batch_is_expired(batch: BatchRecord) -> bool:
    """Return whether a batch has expired as of today.

    Expiry is decided by the date. Nothing sets ``status = 'EXPIRED'``, so the
    original status check never fired and expired stock could be purchased.
    """
    if batch.status == "DESTROYED":
        return False
    if batch.status == "EXPIRED":
        return True
    return batch.expiry_date is not None and batch.expiry_date <= utc_now().date()


class PurchaseService(TransactionalDocumentService):
    """Coordinate purchase order lifecycle, calculations, and integrations."""

    DOCUMENT = DocumentTypeSpec(
        code="PURCHASE_ORDER",
        name="Purchase Order",
        description="Reusable purchase document type.",
        category="PURCHASE",
        module="purchase",
        prefix="PO",
        include_branch_code=True,
        include_company_code=True,
        rule_code="PURCHASE_ORDER_DEFAULT",
        rule_name="Purchase Order Default Numbering",
        states=(
            DocumentStateSpec("DRAFT", "Draft", 10, allows_edit=True),
            DocumentStateSpec("APPROVED", "Approved", 20, allows_edit=True),
            DocumentStateSpec("CANCELLED", "Cancelled", 90, is_terminal=True),
            DocumentStateSpec("CLOSED", "Closed", 100, is_terminal=True),
        ),
    )

    def __init__(self, session: Session) -> None:
        """Bind the lifecycle base plus this module's collaborators."""
        super().__init__(session)
        self._uom = UomService(session)
        self._tax = TaxRuleService(session)

    def list_orders(
        self,
        *,
        firm_scope: UUID,
        filters: PurchaseOrderListFilters,
        page: int,
        page_size: int,
        search: str | None,
        sort_by: str,
        descending: bool,
    ) -> tuple[list[PurchaseOrder], int]:
        """List orders."""
        statement = select(PurchaseOrder).where(PurchaseOrder.firm_id == firm_scope)
        count = (
            select(func.count())
            .select_from(PurchaseOrder)
            .where(PurchaseOrder.firm_id == firm_scope)
        )
        if not filters.include_deleted:
            statement = statement.where(PurchaseOrder.is_deleted.is_(False))
            count = count.where(PurchaseOrder.is_deleted.is_(False))
        if filters.vendor_id is not None:
            statement = statement.where(PurchaseOrder.vendor_id == filters.vendor_id)
            count = count.where(PurchaseOrder.vendor_id == filters.vendor_id)
        if filters.status is not None:
            statement = statement.where(PurchaseOrder.status == filters.status.value)
            count = count.where(PurchaseOrder.status == filters.status.value)
        if filters.branch_id is not None:
            statement = statement.where(PurchaseOrder.branch_id == filters.branch_id)
            count = count.where(PurchaseOrder.branch_id == filters.branch_id)
        if filters.warehouse_id is not None:
            statement = statement.where(
                PurchaseOrder.warehouse_id == filters.warehouse_id
            )
            count = count.where(PurchaseOrder.warehouse_id == filters.warehouse_id)
        if filters.buyer_id is not None:
            statement = statement.where(PurchaseOrder.buyer_id == filters.buyer_id)
            count = count.where(PurchaseOrder.buyer_id == filters.buyer_id)
        if filters.purchase_type is not None:
            statement = statement.where(
                PurchaseOrder.purchase_type == filters.purchase_type.value
            )
            count = count.where(
                PurchaseOrder.purchase_type == filters.purchase_type.value
            )
        if filters.created_from is not None:
            statement = statement.where(
                PurchaseOrder.purchase_date >= filters.created_from
            )
            count = count.where(PurchaseOrder.purchase_date >= filters.created_from)
        if filters.created_to is not None:
            statement = statement.where(
                PurchaseOrder.purchase_date <= filters.created_to
            )
            count = count.where(PurchaseOrder.purchase_date <= filters.created_to)
        if search:
            term = f"%{search.strip()}%"
            condition = or_(
                PurchaseOrder.po_number.ilike(term),
                PurchaseOrder.reference_number.ilike(term),
                PurchaseOrder.external_reference.ilike(term),
                PurchaseOrder.remarks.ilike(term),
            )
            statement = statement.where(condition)
            count = count.where(condition)
        sort_column = getattr(PurchaseOrder, sort_by, PurchaseOrder.created_at)
        order_by = sort_column.desc() if descending else sort_column.asc()
        rows = list(
            self._session.scalars(
                statement.order_by(order_by)
                .offset((page - 1) * page_size)
                .limit(page_size)
            ).all()
        )
        return rows, int(self._session.scalar(count) or 0)

    def summary(self, *, firm_scope: UUID) -> PurchaseSummary:
        """Summarize ."""
        rows = list(
            self._session.scalars(
                select(PurchaseOrder).where(
                    PurchaseOrder.firm_id == firm_scope,
                    PurchaseOrder.is_deleted.is_(False),
                )
            ).all()
        )
        overdue = sum(
            1
            for row in rows
            if row.expected_delivery_date is not None
            and row.expected_delivery_date < utc_now().date()
            and row.status
            not in {
                PurchaseOrderStatus.CANCELLED.value,
                PurchaseOrderStatus.CLOSED.value,
                PurchaseOrderStatus.RECEIVED.value,
            }
        )
        total_value = sum((row.grand_total for row in rows), Decimal("0"))
        return PurchaseSummary(
            total=len(rows),
            draft=sum(
                1 for row in rows if row.status == PurchaseOrderStatus.DRAFT.value
            ),
            open=sum(
                1
                for row in rows
                if row.status
                in {
                    PurchaseOrderStatus.SUBMITTED.value,
                    PurchaseOrderStatus.APPROVED.value,
                    PurchaseOrderStatus.ORDERED.value,
                    PurchaseOrderStatus.PARTIALLY_ORDERED.value,
                    PurchaseOrderStatus.PARTIALLY_RECEIVED.value,
                }
            ),
            cancelled=sum(
                1 for row in rows if row.status == PurchaseOrderStatus.CANCELLED.value
            ),
            closed=sum(
                1 for row in rows if row.status == PurchaseOrderStatus.CLOSED.value
            ),
            total_value=self._q(total_value),
            overdue_delivery=overdue,
        )

    def create_order(
        self, data: PurchaseOrderCreate, *, firm_id: UUID, actor_id: UUID
    ) -> PurchaseOrder:
        """Create order."""
        assert_feature_fields(
            self._session,
            firm_id,
            feature="ATTACHMENTS",
            values={"attachments": data.attachments},
        )
        document_type, numbering_rule = self._ensure_document_setup(
            firm_id=firm_id, actor_id=actor_id
        )
        self._validate_scope_references(
            firm_id=firm_id,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            vendor_id=data.vendor_id,
        )
        branch_code, company_code = self._scope_codes(
            firm_id=firm_id, branch_id=data.branch_id
        )
        po_number = (
            data.po_number.strip()
            if data.po_number
            else self._documents.reserve_number(
                numbering_rule.id,
                firm_id=firm_id,
                financial_year_label=self._financial_year_label(
                    data.purchase_date, firm_id
                ),
                branch_code=branch_code,
                company_code=company_code,
                document_date=data.purchase_date,
                actor_id=actor_id,
            )
        )
        row = PurchaseOrder(
            firm_id=firm_id,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            vendor_id=data.vendor_id,
            buyer_id=data.buyer_id,
            tax_profile_id=data.tax_profile_id,
            po_number=po_number,
            vendor_contact=data.vendor_contact,
            vendor_address=data.vendor_address,
            department=data.department,
            purchase_type=data.purchase_type.value,
            purchase_category=data.purchase_category,
            purchase_date=data.purchase_date,
            expected_delivery_date=data.expected_delivery_date,
            payment_terms=data.payment_terms,
            delivery_terms=data.delivery_terms,
            currency_code=data.currency_code,
            exchange_rate=data.exchange_rate,
            reference_number=data.reference_number,
            external_reference=data.external_reference,
            priority=data.priority,
            remarks=data.remarks,
            status=data.status.value,
            header_discount_amount=data.header_discount_amount,
            additional_charges=data.additional_charges,
            round_off=data.round_off,
            created_by=actor_id,
            updated_by=actor_id,
        )
        self._session.add(row)
        self._flush_or_conflict("Purchase order number already exists in this firm.")
        totals = self._replace_lines(row, data=data, actor_id=actor_id)
        row.subtotal = totals["subtotal"]
        row.line_discount_total = totals["line_discount_total"]
        row.tax_total = totals["tax_total"]
        row.grand_total = totals["grand_total"]
        self._replace_schedules(row, data=data, actor_id=actor_id)
        self._replace_attachments(row, data=data, actor_id=actor_id)
        self._replace_notes(row, data=data, actor_id=actor_id)
        self._history(
            order=row,
            action="purchase.created",
            from_status=None,
            to_status=row.status,
            actor_id=actor_id,
            details={"po_number": row.po_number},
        )
        self._record_document_event(
            firm_id=firm_id,
            document_type=document_type,
            order=row,
            action="CREATED",
            from_state=None,
            to_state=row.status,
            actor_id=actor_id,
            details={"po_number": row.po_number},
        )
        record_audit(
            self._session,
            action="purchase.created",
            entity_type="purchase_order",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_id,
            after_data={"po_number": row.po_number, "status": row.status},
        )
        self._flush_or_conflict("Purchase order number already exists in this firm.")
        self._session.commit()
        return row

    def update_order(
        self,
        order_id: UUID,
        data: PurchaseOrderUpdate,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> PurchaseOrder:
        """Change order."""
        assert_feature_fields(
            self._session,
            firm_scope,
            feature="ATTACHMENTS",
            values={"attachments": data.attachments},
        )
        row = self.get_order(order_id, firm_scope=firm_scope)
        if row.status in {
            PurchaseOrderStatus.CANCELLED.value,
            PurchaseOrderStatus.CLOSED.value,
        }:
            raise ValidationError("Cancelled/closed purchase orders cannot be updated.")
        self._validate_scope_references(
            firm_id=firm_scope,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            vendor_id=data.vendor_id,
        )
        before_status = row.status
        row.branch_id = data.branch_id
        row.warehouse_id = data.warehouse_id
        row.vendor_id = data.vendor_id
        row.buyer_id = data.buyer_id
        row.tax_profile_id = data.tax_profile_id
        row.vendor_contact = data.vendor_contact
        row.vendor_address = data.vendor_address
        row.department = data.department
        row.purchase_type = data.purchase_type.value
        row.purchase_category = data.purchase_category
        row.purchase_date = data.purchase_date
        row.expected_delivery_date = data.expected_delivery_date
        row.payment_terms = data.payment_terms
        row.delivery_terms = data.delivery_terms
        row.currency_code = data.currency_code
        row.exchange_rate = data.exchange_rate
        row.reference_number = data.reference_number
        row.external_reference = data.external_reference
        row.priority = data.priority
        row.remarks = data.remarks
        row.status = data.status.value
        row.header_discount_amount = data.header_discount_amount
        row.additional_charges = data.additional_charges
        row.round_off = data.round_off
        row.updated_by = actor_id
        totals = self._replace_lines(row, data=data, actor_id=actor_id)
        row.subtotal = totals["subtotal"]
        row.line_discount_total = totals["line_discount_total"]
        row.tax_total = totals["tax_total"]
        row.grand_total = totals["grand_total"]
        self._replace_schedules(row, data=data, actor_id=actor_id)
        self._replace_attachments(row, data=data, actor_id=actor_id)
        self._replace_notes(row, data=data, actor_id=actor_id)
        self._history(
            order=row,
            action="purchase.updated",
            from_status=before_status,
            to_status=row.status,
            actor_id=actor_id,
            details={"po_number": row.po_number},
        )
        self._record_document_event(
            firm_id=firm_scope,
            document_type=self._ensure_document_setup(
                firm_id=firm_scope, actor_id=actor_id
            )[0],
            order=row,
            action="EDITED",
            from_state=before_status,
            to_state=row.status,
            actor_id=actor_id,
            details={"po_number": row.po_number},
        )
        record_audit(
            self._session,
            action="purchase.updated",
            entity_type="purchase_order",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data={"status": before_status},
            after_data={"status": row.status, "grand_total": str(row.grand_total)},
        )
        self._flush_or_conflict("Purchase order update conflicts with existing data.")
        self._session.commit()
        return row

    def get_order(
        self, order_id: UUID, *, firm_scope: UUID, include_deleted: bool = False
    ) -> PurchaseOrder:
        """Return order."""
        statement = select(PurchaseOrder).where(
            PurchaseOrder.id == order_id,
            PurchaseOrder.firm_id == firm_scope,
        )
        if not include_deleted:
            statement = statement.where(PurchaseOrder.is_deleted.is_(False))
        row = self._session.scalar(statement)
        if row is None:
            raise ResourceNotFoundError("Purchase order not found.")
        return row

    def _assert_order_removable(self, order: PurchaseOrder) -> None:
        """Refuse to delete an order goods have already been received against.

        Cancelling a received order was refused; deleting one was not, and
        delete is the more destructive of the two. A receipt records its
        purchase_order_id, so removing the order leaves the receipt pointing at
        a document no listing shows.
        """
        if order.status == PurchaseOrderStatus.RECEIVED.value:
            raise ValidationError("Received purchase orders cannot be deleted.")
        received = self._session.scalar(
            select(GoodsReceipt.id)
            .where(
                GoodsReceipt.purchase_order_id == order.id,
                GoodsReceipt.is_deleted.is_(False),
            )
            .limit(1)
        )
        if received is not None:
            raise ValidationError(
                "Goods have been received against this order; cancel it instead."
            )

    def delete_order(self, order_id: UUID, *, firm_scope: UUID, actor_id: UUID) -> None:
        """Soft delete a purchase order nothing has been received against."""
        row = self.get_order(order_id, firm_scope=firm_scope)
        self._assert_order_removable(row)
        row.is_deleted = True
        row.deleted_at = utc_now()
        row.deleted_by = actor_id
        row.updated_by = actor_id
        document_type = self._ensure_document_setup(
            firm_id=firm_scope, actor_id=actor_id
        )[0]
        self._history(
            order=row,
            action="purchase.deleted",
            from_status=row.status,
            to_status=row.status,
            actor_id=actor_id,
        )
        self._record_document_event(
            firm_id=firm_scope,
            document_type=document_type,
            order=row,
            action="EDITED",
            from_state=row.status,
            to_state=row.status,
            actor_id=actor_id,
        )
        record_audit(
            self._session,
            action="purchase.deleted",
            entity_type="purchase_order",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data={"status": row.status},
        )
        self._session.commit()

    def restore_order(
        self, order_id: UUID, *, firm_scope: UUID, actor_id: UUID
    ) -> PurchaseOrder:
        """Restore order."""
        row = self.get_order(order_id, firm_scope=firm_scope, include_deleted=True)
        row.is_deleted = False
        row.deleted_at = None
        row.deleted_by = None
        row.updated_by = actor_id
        document_type = self._ensure_document_setup(
            firm_id=firm_scope, actor_id=actor_id
        )[0]
        self._history(
            order=row,
            action="purchase.restored",
            from_status=row.status,
            to_status=row.status,
            actor_id=actor_id,
        )
        self._record_document_event(
            firm_id=firm_scope,
            document_type=document_type,
            order=row,
            action="EDITED",
            from_state=row.status,
            to_state=row.status,
            actor_id=actor_id,
        )
        record_audit(
            self._session,
            action="purchase.restored",
            entity_type="purchase_order",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            after_data={"status": row.status},
        )
        self._session.commit()
        return row

    def cancel_order(
        self, order_id: UUID, *, firm_scope: UUID, actor_id: UUID, reason: str | None
    ) -> PurchaseOrder:
        """Cancel order."""
        row = self.get_order(order_id, firm_scope=firm_scope)
        if row.status in {
            PurchaseOrderStatus.CANCELLED.value,
            PurchaseOrderStatus.CLOSED.value,
        }:
            return row
        if row.status == PurchaseOrderStatus.RECEIVED.value:
            raise ValidationError("Received purchase orders cannot be cancelled.")
        before = row.status
        row.status = PurchaseOrderStatus.CANCELLED.value
        row.cancel_reason = reason
        row.updated_by = actor_id
        document_type = self._ensure_document_setup(
            firm_id=firm_scope, actor_id=actor_id
        )[0]
        self._history(
            order=row,
            action="purchase.cancelled",
            from_status=before,
            to_status=row.status,
            actor_id=actor_id,
            details={"reason": reason or ""},
        )
        self._record_document_event(
            firm_id=firm_scope,
            document_type=document_type,
            order=row,
            action="CANCELLED",
            from_state=before,
            to_state=row.status,
            actor_id=actor_id,
            remarks=reason,
            details={"reason": reason or ""},
        )
        record_audit(
            self._session,
            action="purchase.cancelled",
            entity_type="purchase_order",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data={"status": before},
            after_data={"status": row.status, "reason": reason or ""},
        )
        self._session.commit()
        return row

    def close_order(
        self, order_id: UUID, *, firm_scope: UUID, actor_id: UUID, reason: str | None
    ) -> PurchaseOrder:
        """Close order."""
        row = self.get_order(order_id, firm_scope=firm_scope)
        if row.status == PurchaseOrderStatus.CLOSED.value:
            return row
        if row.status == PurchaseOrderStatus.CANCELLED.value:
            raise ValidationError("Cancelled purchase orders cannot be closed.")
        before = row.status
        row.status = PurchaseOrderStatus.CLOSED.value
        row.close_reason = reason
        row.updated_by = actor_id
        document_type = self._ensure_document_setup(
            firm_id=firm_scope, actor_id=actor_id
        )[0]
        self._history(
            order=row,
            action="purchase.closed",
            from_status=before,
            to_status=row.status,
            actor_id=actor_id,
            details={"reason": reason or ""},
        )
        self._record_document_event(
            firm_id=firm_scope,
            document_type=document_type,
            order=row,
            action="CLOSED",
            from_state=before,
            to_state=row.status,
            actor_id=actor_id,
            remarks=reason,
            details={"reason": reason or ""},
        )
        record_audit(
            self._session,
            action="purchase.closed",
            entity_type="purchase_order",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data={"status": before},
            after_data={"status": row.status, "reason": reason or ""},
        )
        self._session.commit()
        return row

    def import_orders(
        self,
        data: PurchaseOrderImportRequest,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> list[PurchaseOrder]:
        """Import orders."""
        self._validate_import_records(data.records, firm_scope=firm_scope)
        return [
            self.create_order(record, firm_id=firm_scope, actor_id=actor_id)
            for record in data.records
        ]

    def import_orders_csv(
        self, csv_content: str, *, firm_scope: UUID, actor_id: UUID
    ) -> list[PurchaseOrder]:
        """Import orders csv."""
        reader = csv.DictReader(io.StringIO(csv_content))
        records: list[PurchaseOrderCreate] = []
        for row in reader:
            branch_id = (row.get("BranchId") or "").strip()
            warehouse_id = (row.get("WarehouseId") or "").strip()
            vendor_id = (row.get("VendorId") or "").strip()
            product_id = (row.get("ProductId") or "").strip()
            purchase_date = (row.get("PurchaseDate") or "").strip()
            ordered_qty = (row.get("OrderedQty") or "").strip()
            unit_price = (row.get("UnitPrice") or "").strip()
            if not (
                branch_id
                and warehouse_id
                and vendor_id
                and product_id
                and purchase_date
                and ordered_qty
                and unit_price
            ):
                continue
            records.append(
                self._import_record(
                    branch_id=branch_id,
                    warehouse_id=warehouse_id,
                    vendor_id=vendor_id,
                    product_id=product_id,
                    purchase_date=purchase_date,
                    ordered_quantity=ordered_qty,
                    unit_price=unit_price,
                    po_number=(row.get("PoNumber") or "").strip() or None,
                    remarks=(row.get("Remarks") or "").strip() or None,
                    purchase_uom_id=(row.get("PurchaseUomId") or "").strip() or None,
                    inventory_uom_id=(row.get("InventoryUomId") or "").strip() or None,
                    tax_profile_id=(row.get("TaxProfileId") or "").strip() or None,
                    expected_delivery_date=(
                        (row.get("ExpectedDeliveryDate") or "").strip() or None
                    ),
                    status=(row.get("Status") or "").strip() or None,
                )
            )
        return self.import_orders(
            PurchaseOrderImportRequest(records=records),
            firm_scope=firm_scope,
            actor_id=actor_id,
        )

    def import_orders_xlsx(
        self, workbook_bytes: bytes, *, firm_scope: UUID, actor_id: UUID
    ) -> list[PurchaseOrder]:
        """Import orders xlsx."""
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError as error:
            raise ValidationError(
                "XLSX import dependency is unavailable. Install openpyxl."
            ) from error
        workbook = load_workbook(filename=BytesIO(workbook_bytes), read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(value or "").strip() for value in rows[0]]
        index = {name: position for position, name in enumerate(header)}
        records: list[PurchaseOrderCreate] = []
        for values in rows[1:]:

            def _cell(name: str, row: tuple[object, ...] = tuple(values)) -> str:
                """Read one named cell from this row.

                ``row`` is bound at definition time on purpose: without it the
                closure would read whichever row the loop had reached by the
                time it was called.
                """
                position = index.get(name, -1)
                if position < 0 or position >= len(row):
                    return ""
                return str(row[position] or "").strip()

            branch_id = _cell("BranchId")
            warehouse_id = _cell("WarehouseId")
            vendor_id = _cell("VendorId")
            product_id = _cell("ProductId")
            purchase_date = _cell("PurchaseDate")
            ordered_qty = _cell("OrderedQty")
            unit_price = _cell("UnitPrice")
            if not (
                branch_id
                and warehouse_id
                and vendor_id
                and product_id
                and purchase_date
                and ordered_qty
                and unit_price
            ):
                continue
            records.append(
                self._import_record(
                    branch_id=branch_id,
                    warehouse_id=warehouse_id,
                    vendor_id=vendor_id,
                    product_id=product_id,
                    purchase_date=purchase_date,
                    ordered_quantity=ordered_qty,
                    unit_price=unit_price,
                    po_number=_cell("PoNumber") or None,
                    remarks=_cell("Remarks") or None,
                    purchase_uom_id=_cell("PurchaseUomId") or None,
                    inventory_uom_id=_cell("InventoryUomId") or None,
                    tax_profile_id=_cell("TaxProfileId") or None,
                    expected_delivery_date=_cell("ExpectedDeliveryDate") or None,
                    status=_cell("Status") or None,
                )
            )
        return self.import_orders(
            PurchaseOrderImportRequest(records=records),
            firm_scope=firm_scope,
            actor_id=actor_id,
        )

    def export_orders_csv(self, *, firm_scope: UUID, search: str | None) -> str:
        """Export orders csv."""
        rows, _ = self.list_orders(
            firm_scope=firm_scope,
            filters=PurchaseOrderListFilters(include_deleted=False),
            page=1,
            page_size=5000,
            search=search,
            sort_by="purchase_date",
            descending=True,
        )
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "PO Number",
                "Date",
                "Vendor ID",
                "Branch ID",
                "Warehouse ID",
                "Status",
                "Subtotal",
                "Tax Total",
                "Grand Total",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.po_number,
                    row.purchase_date.isoformat(),
                    str(row.vendor_id),
                    str(row.branch_id),
                    str(row.warehouse_id),
                    row.status,
                    str(row.subtotal),
                    str(row.tax_total),
                    str(row.grand_total),
                ]
            )
        return output.getvalue()

    def export_orders_xlsx(self, *, firm_scope: UUID, search: str | None) -> bytes:
        """Export orders xlsx."""
        try:
            from openpyxl import Workbook
        except ImportError as error:
            raise ValidationError(
                "XLSX export dependency is unavailable. Install openpyxl."
            ) from error
        rows, _ = self.list_orders(
            firm_scope=firm_scope,
            filters=PurchaseOrderListFilters(include_deleted=False),
            page=1,
            page_size=5000,
            search=search,
            sort_by="purchase_date",
            descending=True,
        )
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise ValidationError("Unable to generate purchase export workbook.")
        sheet.title = "PurchaseOrders"
        sheet.append(
            [
                "PO Number",
                "Date",
                "Vendor ID",
                "Branch ID",
                "Warehouse ID",
                "Status",
                "Subtotal",
                "Tax Total",
                "Grand Total",
            ]
        )
        for row in rows:
            sheet.append(
                [
                    row.po_number,
                    row.purchase_date.isoformat(),
                    str(row.vendor_id),
                    str(row.branch_id),
                    str(row.warehouse_id),
                    row.status,
                    # openpyxl writes Decimal natively, so there is no reason to
                    # round-trip money through binary floating point here.
                    self._q(row.subtotal),
                    self._q(row.tax_total),
                    self._q(row.grand_total),
                ]
            )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def order_response(self, row: PurchaseOrder) -> PurchaseOrderResponse:
        """Order response."""
        lines = list(
            self._session.scalars(
                select(PurchaseOrderLine)
                .where(
                    PurchaseOrderLine.purchase_order_id == row.id,
                    PurchaseOrderLine.is_deleted.is_(False),
                )
                .order_by(PurchaseOrderLine.line_number.asc())
            ).all()
        )
        schedules = list(
            self._session.scalars(
                select(PurchaseDeliverySchedule)
                .where(
                    PurchaseDeliverySchedule.firm_id == row.firm_id,
                    PurchaseDeliverySchedule.is_deleted.is_(False),
                    (
                        PurchaseDeliverySchedule.purchase_order_line_id.in_(
                            [item.id for item in lines]
                        )
                        if lines
                        else false()
                    ),
                )
                .order_by(PurchaseDeliverySchedule.delivery_date.asc())
            ).all()
        )
        attachments = list(
            self._session.scalars(
                select(PurchaseAttachment).where(
                    PurchaseAttachment.purchase_order_id == row.id,
                    PurchaseAttachment.is_deleted.is_(False),
                )
            ).all()
        )
        notes = list(
            self._session.scalars(
                select(PurchaseNote).where(
                    PurchaseNote.purchase_order_id == row.id,
                    PurchaseNote.is_deleted.is_(False),
                )
            ).all()
        )
        payload = PurchaseOrderResponse.model_validate(row).model_dump(mode="python")
        payload["lines"] = [
            PurchaseOrderLineResponse.model_validate(item).model_dump(mode="python")
            for item in lines
        ]
        payload["delivery_schedules"] = [
            PurchaseDeliveryScheduleResponse.model_validate(
                {
                    "id": entry.id,
                    "purchase_order_line_id": entry.purchase_order_line_id,
                    "line_number": self._line_number(
                        lines, entry.purchase_order_line_id
                    ),
                    "delivery_date": entry.delivery_date,
                    "quantity": entry.quantity,
                    "status": entry.status,
                    "remarks": entry.remarks,
                    "created_at": entry.created_at,
                    "updated_at": entry.updated_at,
                }
            ).model_dump(mode="python")
            for entry in schedules
        ]
        payload["attachments"] = [
            PurchaseAttachmentResponse.model_validate(item).model_dump(mode="python")
            for item in attachments
        ]
        payload["notes"] = [
            PurchaseNoteResponse.model_validate(item).model_dump(mode="python")
            for item in notes
        ]
        return PurchaseOrderResponse.model_validate(payload)

    def order_history(
        self, *, order_id: UUID, firm_scope: UUID
    ) -> list[PurchaseOrderHistory]:
        """Order history."""
        self.get_order(order_id, firm_scope=firm_scope, include_deleted=True)
        return list(
            self._session.scalars(
                select(PurchaseOrderHistory)
                .where(
                    PurchaseOrderHistory.purchase_order_id == order_id,
                    PurchaseOrderHistory.firm_id == firm_scope,
                    PurchaseOrderHistory.is_deleted.is_(False),
                )
                .order_by(PurchaseOrderHistory.created_at.asc())
            ).all()
        )

    def _line_number(self, lines: list[PurchaseOrderLine], line_id: UUID) -> int:
        """Line number."""
        for line in lines:
            if line.id == line_id:
                return line.line_number
        return 0

    def _replace_lines(
        self,
        order: PurchaseOrder,
        *,
        data: PurchaseOrderCreate | PurchaseOrderUpdate,
        actor_id: UUID,
    ) -> dict[str, Decimal]:
        """Replace lines."""
        # Lines are matched on their line number and updated in place;
        # re-inserting them minted a new UUID per line on every save, and
        # downstream documents reference those ids with no foreign key.
        existing = {
            existing_line.line_number: existing_line
            for existing_line in self._session.scalars(
                select(PurchaseOrderLine).where(
                    PurchaseOrderLine.purchase_order_id == order.id
                )
            ).all()
        }
        seen: set[int] = set()
        gross_total = Decimal("0")
        line_discount_total = Decimal("0")
        tax_total = Decimal("0")
        for idx, line in enumerate(data.lines, start=1):
            product = self._active_product(order.firm_id, line.product_id)
            conversion = self._conversion(
                quantity=line.ordered_quantity + line.free_quantity,
                purchase_uom_id=line.purchase_uom_id,
                inventory_uom_id=line.inventory_uom_id,
                product_id=product.id,
                purchase_date=order.purchase_date,
                firm_id=order.firm_id,
            )
            gross_amount = self._q(line.ordered_quantity * line.unit_price)
            discount_amount = (
                self._q(gross_amount * line.discount_percent / Decimal("100"))
                if line.discount_amount <= 0
                else self._q(line.discount_amount)
            )
            taxable = self._q(gross_amount - discount_amount)
            # A product names its tax group, not a version, so the rate is
            # resolved from the document date. product.tax_profile_id has not
            # existed since the group_code refactor and raised AttributeError
            # whenever neither the line nor the order named a profile.
            tax_profile_id = line.tax_profile_id or order.tax_profile_id
            if tax_profile_id is None:
                resolved = TaxFrameworkService(
                    self._session
                ).resolve_profile_for_product(
                    product, order.purchase_date, firm_scope=order.firm_id
                )
                tax_profile_id = resolved.id if resolved else None
            tax_amount = self._line_tax_amount(
                firm_id=order.firm_id,
                actor_id=actor_id,
                tax_profile_id=tax_profile_id,
                vendor_id=order.vendor_id,
                product_id=product.id,
                purchase_date=order.purchase_date,
                taxable=taxable,
            )
            net_amount = self._q(taxable + tax_amount)
            row = PurchaseOrderLine(
                purchase_order_id=order.id,
                firm_id=order.firm_id,
                line_number=idx,
                product_id=product.id,
                description=line.description or product.name,
                vendor_product_code=line.vendor_product_code,
                purchase_uom_id=line.purchase_uom_id or product.purchase_uom_id,
                inventory_uom_id=line.inventory_uom_id or product.inventory_uom_id,
                conversion_factor=conversion["factor"],
                conversion_version=conversion["version"],
                ordered_quantity=self._q(line.ordered_quantity),
                free_quantity=self._q(line.free_quantity),
                base_quantity=conversion["converted"],
                unit_price=self._q(line.unit_price),
                discount_percent=self._q(line.discount_percent),
                discount_amount=discount_amount,
                gross_amount=gross_amount,
                tax_profile_id=tax_profile_id,
                tax_amount=tax_amount,
                net_amount=net_amount,
                batch_required=line.batch_required,
                expiry_required=line.expiry_required,
                serial_required=line.serial_required,
                manufacturing_date=line.manufacturing_date,
                expiry_date=line.expiry_date,
                warehouse_id=line.warehouse_id or order.warehouse_id,
                storage_node_id=line.storage_node_id,
                remarks=line.remarks,
                status=PurchaseOrderStatus.ORDERED.value,
                created_by=actor_id,
                updated_by=actor_id,
            )
            self._validate_line_dates(row)
            self._validate_storage_scope(
                order.firm_id, row.warehouse_id, row.storage_node_id
            )
            persisted = existing.get(idx)
            if persisted is None:
                self._session.add(row)
            else:
                self._apply_line_values(
                    persisted,
                    row,
                    actor_id=actor_id,
                    preserve=("received_quantity", "invoiced_quantity"),
                )
            seen.add(idx)
            gross_total += gross_amount
            line_discount_total += discount_amount
            tax_total += tax_amount
        for line_number, obsolete in existing.items():
            if line_number not in seen:
                self._session.delete(obsolete)
        self._session.flush()
        # subtotal is the taxable base — gross less line discount, before tax —
        # which is what every other transactional document reports. This module
        # used to report gross before discount under the same name.
        subtotal = self._q(gross_total - line_discount_total)
        grand_total = self._q(
            subtotal
            - data.header_discount_amount
            + tax_total
            + data.additional_charges
            + data.round_off
        )
        return {
            "subtotal": subtotal,
            "line_discount_total": self._q(line_discount_total),
            "tax_total": self._q(tax_total),
            "grand_total": grand_total,
        }

    def _replace_schedules(
        self,
        order: PurchaseOrder,
        *,
        data: PurchaseOrderCreate | PurchaseOrderUpdate,
        actor_id: UUID,
    ) -> None:
        """Replace schedules."""
        line_map = {
            item.line_number: item
            for item in self._session.scalars(
                select(PurchaseOrderLine).where(
                    PurchaseOrderLine.purchase_order_id == order.id
                )
            ).all()
        }
        self._session.query(PurchaseDeliverySchedule).where(
            PurchaseDeliverySchedule.firm_id == order.firm_id,
            (
                PurchaseDeliverySchedule.purchase_order_line_id.in_(
                    [item.id for item in line_map.values()]
                )
                if line_map
                else false()
            ),
        ).delete(synchronize_session=False)
        for schedule in data.delivery_schedules:
            line = line_map.get(schedule.line_number)
            if line is None:
                raise ValidationError(
                    f"Delivery schedule references unknown line {schedule.line_number}."
                )
            self._session.add(
                PurchaseDeliverySchedule(
                    purchase_order_line_id=line.id,
                    firm_id=order.firm_id,
                    delivery_date=schedule.delivery_date,
                    quantity=self._q(schedule.quantity),
                    status="PENDING",
                    remarks=schedule.remarks,
                    created_by=actor_id,
                    updated_by=actor_id,
                )
            )

    def _replace_attachments(
        self,
        order: PurchaseOrder,
        *,
        data: PurchaseOrderCreate | PurchaseOrderUpdate,
        actor_id: UUID,
    ) -> None:
        """Replace attachments."""
        self._session.query(PurchaseAttachment).filter(
            PurchaseAttachment.purchase_order_id == order.id
        ).delete(synchronize_session=False)
        for item in data.attachments:
            self._session.add(
                PurchaseAttachment(
                    purchase_order_id=order.id,
                    firm_id=order.firm_id,
                    file_name=item.file_name,
                    mime_type=item.mime_type,
                    file_path=item.file_path,
                    attachment_kind=item.attachment_kind.strip().upper(),
                    created_by=actor_id,
                    updated_by=actor_id,
                )
            )

    def _replace_notes(
        self,
        order: PurchaseOrder,
        *,
        data: PurchaseOrderCreate | PurchaseOrderUpdate,
        actor_id: UUID,
    ) -> None:
        """Replace notes."""
        self._session.query(PurchaseNote).filter(
            PurchaseNote.purchase_order_id == order.id
        ).delete(synchronize_session=False)
        for item in data.notes:
            self._session.add(
                PurchaseNote(
                    purchase_order_id=order.id,
                    firm_id=order.firm_id,
                    note_type=item.note_type.value,
                    note=item.note,
                    created_by=actor_id,
                    updated_by=actor_id,
                )
            )

    def _history(
        self,
        *,
        order: PurchaseOrder,
        action: str,
        from_status: str | None,
        to_status: str | None,
        actor_id: UUID,
        details: dict[str, object] | None = None,
        remarks: str | None = None,
    ) -> None:
        """History ."""
        self._session.add(
            PurchaseOrderHistory(
                purchase_order_id=order.id,
                firm_id=order.firm_id,
                action=action,
                from_status=from_status,
                to_status=to_status,
                remarks=remarks,
                details_json=json.dumps(details or {}),
                created_by=actor_id,
                updated_by=actor_id,
            )
        )

    def _record_document_event(
        self,
        *,
        firm_id: UUID,
        document_type: DocumentTypeDefinition,
        order: PurchaseOrder,
        action: str,
        from_state: str | None,
        to_state: str | None,
        actor_id: UUID,
        details: dict[str, object] | None = None,
        remarks: str | None = None,
    ) -> None:
        """Record document event."""
        self._documents.record_event(
            firm_id,
            DocumentLifecycleEventCreate(
                document_type_id=document_type.id,
                source_document_id=order.id,
                source_module_code="PURCHASE",
                document_number=order.po_number,
                action=action,
                from_state=from_state,
                to_state=to_state,
                remarks=remarks,
                details_json=details,
                actor_id=actor_id,
            ),
            actor_id,
        )

    def _validate_scope_references(
        self, *, firm_id: UUID, branch_id: UUID, warehouse_id: UUID, vendor_id: UUID
    ) -> None:
        """Validate scope references."""
        branch = self._session.scalar(
            select(Branch).where(
                Branch.id == branch_id,
                Branch.firm_id == firm_id,
                Branch.is_deleted.is_(False),
            )
        )
        if branch is None:
            raise ValidationError("Selected branch is not available in this firm.")
        if branch.status != "ACTIVE":
            raise ValidationError("Inactive branches cannot be used in purchases.")
        warehouse = self._session.scalar(
            select(Warehouse).where(
                Warehouse.id == warehouse_id,
                Warehouse.firm_id == firm_id,
                Warehouse.is_deleted.is_(False),
            )
        )
        if warehouse is None:
            raise ValidationError("Selected warehouse is not available in this firm.")
        if warehouse.status != "ACTIVE":
            raise ValidationError("Inactive warehouses cannot be used in purchases.")
        if warehouse.branch_id != branch.id:
            raise ValidationError("Warehouse does not belong to selected branch.")
        vendor = self._session.scalar(
            select(Vendor).where(
                Vendor.id == vendor_id,
                Vendor.firm_id == firm_id,
                Vendor.is_deleted.is_(False),
            )
        )
        if vendor is None:
            raise ValidationError("Selected vendor is not available in this firm.")
        if vendor.status != "ACTIVE":
            raise ValidationError(
                "Inactive or blocked vendors cannot be used in purchases."
            )

    def _validate_storage_scope(
        self, firm_id: UUID, warehouse_id: UUID | None, storage_node_id: UUID | None
    ) -> None:
        """Validate storage scope."""
        if warehouse_id is None or storage_node_id is None:
            return
        warehouse = self._session.scalar(
            select(Warehouse).where(
                Warehouse.id == warehouse_id,
                Warehouse.firm_id == firm_id,
                Warehouse.is_deleted.is_(False),
            )
        )
        if warehouse is None:
            raise ValidationError("Line warehouse is unavailable for this firm.")
        storage = self._session.scalar(
            select(WarehouseStorageNode).where(
                WarehouseStorageNode.id == storage_node_id,
                WarehouseStorageNode.warehouse_id == warehouse_id,
                WarehouseStorageNode.is_deleted.is_(False),
            )
        )
        if storage is None:
            raise ValidationError(
                "Line storage area is unavailable for selected warehouse."
            )
        if not storage.is_active:
            raise ValidationError("Inactive storage areas cannot be used in purchases.")

    def _active_product(self, firm_id: UUID, product_id: UUID) -> Product:
        """Active product."""
        row = self._session.scalar(
            select(Product).where(
                Product.id == product_id,
                Product.firm_id == firm_id,
                Product.is_deleted.is_(False),
            )
        )
        if row is None:
            raise ValidationError("Selected product is not available in this firm.")
        if row.status != "ACTIVE":
            raise ValidationError("Inactive/blocked products cannot be purchased.")
        return row

    def _line_tax_amount(
        self,
        *,
        firm_id: UUID,
        actor_id: UUID,
        tax_profile_id: UUID | None,
        vendor_id: UUID,
        product_id: UUID,
        purchase_date: date,
        taxable: Decimal,
    ) -> Decimal:
        """Line tax amount."""
        if tax_profile_id is None:
            return Decimal("0")
        self._assert_tax_profile_available(tax_profile_id, firm_id=firm_id)
        simulation = self._tax.simulate(
            TaxRuleSimulationRequest(
                transaction_type="PURCHASE",
                transaction_date=purchase_date,
                tax_profile_id=tax_profile_id,
                vendor_id=vendor_id,
                product_id=product_id,
                invoice_value=self._q(taxable),
            ),
            firm_scope=firm_id,
            actor_id=actor_id,
        )
        return self._q(simulation.total_tax_amount)

    def _conversion(
        self,
        *,
        quantity: Decimal,
        purchase_uom_id: UUID | None,
        inventory_uom_id: UUID | None,
        product_id: UUID,
        purchase_date: date,
        firm_id: UUID,
    ) -> dict[str, Decimal | int | None]:
        """Conversion ."""
        if (
            purchase_uom_id is None
            or inventory_uom_id is None
            or purchase_uom_id == inventory_uom_id
        ):
            return {
                "factor": Decimal("1"),
                "converted": self._q(quantity),
                "version": None,
            }
        response = self._uom.convert_quantity(
            ConversionRequest(
                quantity=quantity,
                from_uom_id=purchase_uom_id,
                to_uom_id=inventory_uom_id,
                product_id=product_id,
                conversion_date=purchase_date,
            ),
            firm_scope=firm_id,
        )
        return {
            "factor": response.conversion_factor,
            "converted": self._q(response.converted_quantity),
            "version": response.version,
        }

    def _validate_line_dates(self, line: PurchaseOrderLine) -> None:
        """Validate line dates."""
        if line.expiry_required and line.expiry_date is None:
            raise ValidationError(
                "Expiry date is required for lines marked as expiry-required."
            )
        if (
            line.manufacturing_date is not None
            and line.expiry_date is not None
            and line.expiry_date < line.manufacturing_date
        ):
            raise ValidationError("Expiry date cannot be before manufacturing date.")
        if line.expiry_date is not None:
            existing_batch = self._session.scalar(
                select(BatchRecord).where(
                    BatchRecord.firm_id == line.firm_id,
                    BatchRecord.product_id == line.product_id,
                    BatchRecord.expiry_date == line.expiry_date,
                    BatchRecord.is_deleted.is_(False),
                )
            )
            if existing_batch is not None and _batch_is_expired(existing_batch):
                raise ValidationError("Expired products cannot be purchased.")

    def _assert_tax_profile_available(
        self, profile_id: UUID, *, firm_id: UUID, document_date: date | None = None
    ) -> None:
        """Reject a profile that is unavailable, or not in force on the document."""
        service = TaxFrameworkService(self._session)
        if document_date is None:
            profile = self._session.scalar(
                select(TaxProfile).where(
                    TaxProfile.id == profile_id,
                    TaxProfile.firm_id == firm_id,
                    TaxProfile.is_deleted.is_(False),
                )
            )
            if profile is None:
                raise ValidationError(
                    "Selected tax profile is not available in this firm."
                )
            if profile.status != "ACTIVE":
                raise ValidationError(
                    "Inactive tax profiles cannot be used in purchases."
                )
            return
        service.assert_profile_effective_on(
            profile_id, document_date, firm_scope=firm_id
        )

    def _validate_import_records(
        self, records: list[PurchaseOrderCreate], *, firm_scope: UUID
    ) -> None:
        """Validate import records."""
        explicit_numbers = [
            record.po_number.strip().upper()
            for record in records
            if record.po_number and record.po_number.strip()
        ]
        duplicate_numbers = {
            number for number in explicit_numbers if explicit_numbers.count(number) > 1
        }
        if duplicate_numbers:
            duplicate_list = ", ".join(sorted(duplicate_numbers))
            raise ConflictError(
                "Duplicate purchase order numbers found in import payload: "
                f"{duplicate_list}."
            )
        if not explicit_numbers:
            return
        existing = self._session.scalars(
            select(PurchaseOrder.po_number).where(
                PurchaseOrder.firm_id == firm_scope,
                func.upper(PurchaseOrder.po_number).in_(explicit_numbers),
            )
        ).all()
        if existing:
            duplicate_list = ", ".join(sorted({item for item in existing if item}))
            raise ConflictError(
                f"Purchase order numbers already exist in this firm: {duplicate_list}."
            )

    def _import_record(
        self,
        *,
        branch_id: str,
        warehouse_id: str,
        vendor_id: str,
        product_id: str,
        purchase_date: str,
        ordered_quantity: str,
        unit_price: str,
        po_number: str | None,
        remarks: str | None,
        purchase_uom_id: str | None,
        inventory_uom_id: str | None,
        tax_profile_id: str | None,
        expected_delivery_date: str | None,
        status: str | None,
    ) -> PurchaseOrderCreate:
        """Import record."""
        payload: dict[str, object] = {
            "po_number": po_number,
            "branch_id": branch_id,
            "warehouse_id": warehouse_id,
            "vendor_id": vendor_id,
            "purchase_date": purchase_date,
            "remarks": remarks,
            "lines": [
                {
                    "product_id": product_id,
                    "purchase_uom_id": purchase_uom_id,
                    "inventory_uom_id": inventory_uom_id,
                    "ordered_quantity": ordered_quantity,
                    "unit_price": unit_price,
                    "tax_profile_id": tax_profile_id,
                }
            ],
        }
        if expected_delivery_date:
            payload["expected_delivery_date"] = expected_delivery_date
        if status:
            payload["status"] = status
        return PurchaseOrderCreate.model_validate(payload)
