"""Transactional service for the enterprise inventory foundation."""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import UUID

from sqlalchemy import case, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.orm.attributes import InstrumentedAttribute
from sqlalchemy.sql import Select

from app.batch_serial.models import BatchRecord
from app.branches.models import Branch, Warehouse, WarehouseStorageNode
from app.business.models import BusinessProfile, FirmBusinessProfile
from app.common.audit.services import record_audit
from app.core.exceptions import ConflictError, ResourceNotFoundError, ValidationError
from app.core.utils.money import quantize_money
from app.finance.services.document_posting import DocumentPostingService
from app.inventory.models import (
    InventoryRecord,
    InventoryTransaction,
    OpeningStockBatch,
    OpeningStockLine,
    ProductValuation,
    StockLedgerEntry,
)
from app.inventory.schemas import (
    REVERSAL_SUFFIX,
    BatchStockTotals,
    InventoryAdjustmentCreate,
    InventoryCreate,
    InventoryListFilters,
    InventoryLocationSummary,
    InventoryResponse,
    InventorySummary,
    InventoryTransactionListFilters,
    InventoryTransactionResponse,
    InventoryTransactionType,
    InventoryUpdate,
    OpeningStockBatchCreate,
    OpeningStockBatchListFilters,
    OpeningStockBatchResponse,
    OpeningStockImportRequest,
    OpeningStockLineCreate,
    OpeningStockLineResponse,
    OpeningStockUpdate,
    StockLedgerListFilters,
    StockLedgerResponse,
)
from app.products.models import Product
from app.uom.models import ConversionRule

ZERO = Decimal("0")


@dataclass(slots=True)
class _Movement:
    transaction_type: str
    reference_number: str
    reference_type: str
    transaction_date: date
    quantity: Decimal
    current_delta: Decimal = ZERO
    reserved_delta: Decimal = ZERO
    blocked_delta: Decimal = ZERO
    damaged_delta: Decimal = ZERO
    quarantine_delta: Decimal = ZERO
    in_transit_delta: Decimal = ZERO
    entered_quantity: Decimal | None = None
    entered_uom_id: UUID | None = None
    conversion_version: int | None = None
    remarks: str | None = None
    unit_cost: Decimal | None = None
    #: The batch that moved, mirrored onto the transaction and the ledger so
    #: "where did batch B-2405 go" is answerable from the history rather than
    #: only from the current balance.
    batch_id: UUID | None = None


class InventoryService:
    """Coordinate inventory projections, immutable movements, and opening stock."""

    def __init__(self, session: Session) -> None:
        """Bind the service to the request unit of work."""
        self._session = session

    def list_inventory(
        self,
        *,
        firm_scope: UUID,
        filters: InventoryListFilters,
        page: int,
        page_size: int,
        search: str | None,
        sort_by: str,
        descending: bool,
    ) -> tuple[list[InventoryRecord], int]:
        """Return a page of stock projections and the total."""
        columns = {
            "created_at": InventoryRecord.created_at,
            "updated_at": InventoryRecord.updated_at,
            "current_quantity": InventoryRecord.current_quantity,
            "available_quantity": InventoryRecord.available_quantity,
            "status": InventoryRecord.status,
            "product_code": Product.code,
        }
        statement = (
            select(InventoryRecord)
            .join(Product, Product.id == InventoryRecord.product_id)
            .join(Branch, Branch.id == InventoryRecord.branch_id)
            .join(Warehouse, Warehouse.id == InventoryRecord.warehouse_id)
            .outerjoin(
                WarehouseStorageNode,
                WarehouseStorageNode.id == InventoryRecord.storage_node_id,
            )
            .where(InventoryRecord.firm_id == firm_scope)
            .options(
                selectinload(InventoryRecord.transactions),
            )
        )
        count = (
            select(func.count())
            .select_from(InventoryRecord)
            .join(Product, Product.id == InventoryRecord.product_id)
            .join(Branch, Branch.id == InventoryRecord.branch_id)
            .join(Warehouse, Warehouse.id == InventoryRecord.warehouse_id)
            .outerjoin(
                WarehouseStorageNode,
                WarehouseStorageNode.id == InventoryRecord.storage_node_id,
            )
            .where(InventoryRecord.firm_id == firm_scope)
        )
        statement, count = self._apply_inventory_filters(statement, count, filters)
        if search:
            term = f"%{search.strip()}%"
            condition = or_(
                Product.code.ilike(term),
                Product.name.ilike(term),
                Branch.code.ilike(term),
                Branch.name.ilike(term),
                Warehouse.code.ilike(term),
                Warehouse.name.ilike(term),
                WarehouseStorageNode.code.ilike(term),
                WarehouseStorageNode.name.ilike(term),
            )
            statement = statement.where(condition)
            count = count.where(condition)
        ordering = columns.get(sort_by, InventoryRecord.created_at)
        rows = self._session.scalars(
            statement.order_by(
                ordering.desc() if descending else ordering.asc(),
                # Tiebreaker: see list_ledger for why every paged query needs one.
                InventoryRecord.id.desc() if descending else InventoryRecord.id.asc(),
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
        ).all()
        return list(rows), int(self._session.scalar(count) or 0)

    def inventory_summary(
        self, *, firm_scope: UUID, filters: InventoryListFilters
    ) -> InventorySummary:
        """Return stock counts, value and exception totals."""
        statement = (
            select(
                func.count(InventoryRecord.id),
                func.coalesce(func.sum(InventoryRecord.current_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.reserved_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.available_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.blocked_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.damaged_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.quarantine_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.in_transit_quantity), 0),
                func.coalesce(
                    func.sum(
                        case(
                            (
                                InventoryRecord.current_quantity
                                <= func.coalesce(
                                    InventoryRecord.reorder_level,
                                    InventoryRecord.minimum_level,
                                    ZERO,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ),
                    0,
                ),
                func.coalesce(
                    func.sum(case((InventoryRecord.current_quantity <= 0, 1), else_=0)),
                    0,
                ),
                func.coalesce(
                    func.sum(
                        case((InventoryRecord.available_quantity < 0, 1), else_=0)
                    ),
                    0,
                ),
            )
            .select_from(InventoryRecord)
            .where(InventoryRecord.firm_id == firm_scope)
        )
        statement, _ = self._apply_inventory_filters(statement, statement, filters)
        (
            total,
            current_quantity,
            reserved_quantity,
            available_quantity,
            blocked_quantity,
            damaged_quantity,
            quarantine_quantity,
            in_transit_quantity,
            low_stock_count,
            out_of_stock_count,
            negative_stock_count,
        ) = self._session.execute(statement).one()
        return InventorySummary(
            total_records=int(total or 0),
            current_quantity=Decimal(current_quantity or 0),
            reserved_quantity=Decimal(reserved_quantity or 0),
            available_quantity=Decimal(available_quantity or 0),
            blocked_quantity=Decimal(blocked_quantity or 0),
            damaged_quantity=Decimal(damaged_quantity or 0),
            quarantine_quantity=Decimal(quarantine_quantity or 0),
            in_transit_quantity=Decimal(in_transit_quantity or 0),
            low_stock_count=int(low_stock_count or 0),
            out_of_stock_count=int(out_of_stock_count or 0),
            negative_stock_count=int(negative_stock_count or 0),
        )

    def stock_by_firm(self, *, firm_scope: UUID) -> list[InventoryLocationSummary]:
        """Return stock totals rolled up to the firm."""
        row = self._session.execute(
            select(
                func.count(InventoryRecord.id),
                func.coalesce(func.sum(InventoryRecord.current_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.reserved_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.available_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.blocked_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.damaged_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.quarantine_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.in_transit_quantity), 0),
            ).where(
                InventoryRecord.firm_id == firm_scope,
                InventoryRecord.is_deleted.is_(False),
            )
        ).first()
        if row is None or int(row[0] or 0) == 0:
            return []
        return [
            InventoryLocationSummary(
                scope_id=firm_scope,
                scope_code="FIRM",
                scope_name="Selected Firm",
                current_quantity=Decimal(row[1] or 0),
                reserved_quantity=Decimal(row[2] or 0),
                available_quantity=Decimal(row[3] or 0),
                blocked_quantity=Decimal(row[4] or 0),
                damaged_quantity=Decimal(row[5] or 0),
                quarantine_quantity=Decimal(row[6] or 0),
                in_transit_quantity=Decimal(row[7] or 0),
            )
        ]

    def stock_by_branch(self, *, firm_scope: UUID) -> list[InventoryLocationSummary]:
        """Return stock totals per branch."""
        rows = self._session.execute(
            select(
                Branch.id,
                Branch.code,
                Branch.name,
                func.coalesce(func.sum(InventoryRecord.current_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.reserved_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.available_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.blocked_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.damaged_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.quarantine_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.in_transit_quantity), 0),
            )
            .join(InventoryRecord, InventoryRecord.branch_id == Branch.id)
            .where(
                Branch.firm_id == firm_scope,
                Branch.is_deleted.is_(False),
                InventoryRecord.is_deleted.is_(False),
            )
            .group_by(Branch.id, Branch.code, Branch.name)
            .order_by(Branch.code.asc())
        ).all()
        return [
            InventoryLocationSummary(
                scope_id=row[0],
                scope_code=row[1],
                scope_name=row[2],
                current_quantity=Decimal(row[3] or 0),
                reserved_quantity=Decimal(row[4] or 0),
                available_quantity=Decimal(row[5] or 0),
                blocked_quantity=Decimal(row[6] or 0),
                damaged_quantity=Decimal(row[7] or 0),
                quarantine_quantity=Decimal(row[8] or 0),
                in_transit_quantity=Decimal(row[9] or 0),
            )
            for row in rows
        ]

    def stock_by_batch(
        self, *, firm_scope: UUID, batch_ids: Sequence[UUID]
    ) -> dict[UUID, BatchStockTotals]:
        """Return what each of these batches is holding, in one query.

        A batch is held per location, so its total is a sum across however many
        stock rows carry it. ``batches`` used to keep its own copy of these six
        numbers, written by the batch API and reconciled against the projection
        by nothing -- so a batch could claim ten on the shelf while no stock row
        anywhere held any of it.

        Callers rendering a page of batches pass every id on the page. Asking
        per row would be a query per batch, which is the mistake
        ``values_for_many`` exists to avoid elsewhere.

        Returns:
            The totals for each batch that holds stock. A batch with no stock
            rows is absent -- ``BatchStockTotals()`` is its answer, and the
            caller supplies it rather than this running a query per empty
            batch.

        """
        if not batch_ids:
            return {}
        rows = self._session.execute(
            select(
                InventoryRecord.batch_id,
                func.coalesce(func.sum(InventoryRecord.current_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.available_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.reserved_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.blocked_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.damaged_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.quarantine_quantity), 0),
            )
            .where(
                InventoryRecord.firm_id == firm_scope,
                InventoryRecord.batch_id.in_(batch_ids),
                InventoryRecord.is_deleted.is_(False),
            )
            .group_by(InventoryRecord.batch_id)
        ).all()
        return {
            row[0]: BatchStockTotals(
                current_quantity=Decimal(row[1] or 0),
                available_quantity=Decimal(row[2] or 0),
                reserved_quantity=Decimal(row[3] or 0),
                blocked_quantity=Decimal(row[4] or 0),
                damaged_quantity=Decimal(row[5] or 0),
                quarantine_quantity=Decimal(row[6] or 0),
            )
            for row in rows
            if row[0] is not None
        }

    def stock_by_product(self, *, firm_scope: UUID) -> list[InventoryLocationSummary]:
        """Return stock totals per product, across batches and locations.

        A batch-tracked product is several stock rows -- one per batch per
        place -- so "how much amoxicillin do I have" stopped being a row and
        became a sum. The list endpoint deliberately still returns the
        individual rows, because which batch stock is in is the reason the
        grain changed; this is where the total lives.
        """
        rows = self._session.execute(
            select(
                Product.id,
                Product.code,
                Product.name,
                func.coalesce(func.sum(InventoryRecord.current_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.reserved_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.available_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.blocked_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.damaged_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.quarantine_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.in_transit_quantity), 0),
            )
            .join(InventoryRecord, InventoryRecord.product_id == Product.id)
            .where(
                InventoryRecord.firm_id == firm_scope,
                Product.is_deleted.is_(False),
                InventoryRecord.is_deleted.is_(False),
            )
            .group_by(Product.id, Product.code, Product.name)
            .order_by(Product.code.asc())
        ).all()
        return [
            InventoryLocationSummary(
                scope_id=row[0],
                scope_code=row[1],
                scope_name=row[2],
                current_quantity=Decimal(row[3] or 0),
                reserved_quantity=Decimal(row[4] or 0),
                available_quantity=Decimal(row[5] or 0),
                blocked_quantity=Decimal(row[6] or 0),
                damaged_quantity=Decimal(row[7] or 0),
                quarantine_quantity=Decimal(row[8] or 0),
                in_transit_quantity=Decimal(row[9] or 0),
            )
            for row in rows
        ]

    def stock_by_warehouse(self, *, firm_scope: UUID) -> list[InventoryLocationSummary]:
        """Return stock totals per warehouse."""
        rows = self._session.execute(
            select(
                Warehouse.id,
                Warehouse.code,
                Warehouse.name,
                func.coalesce(func.sum(InventoryRecord.current_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.reserved_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.available_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.blocked_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.damaged_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.quarantine_quantity), 0),
                func.coalesce(func.sum(InventoryRecord.in_transit_quantity), 0),
            )
            .join(InventoryRecord, InventoryRecord.warehouse_id == Warehouse.id)
            .where(
                Warehouse.firm_id == firm_scope,
                Warehouse.is_deleted.is_(False),
                InventoryRecord.is_deleted.is_(False),
            )
            .group_by(Warehouse.id, Warehouse.code, Warehouse.name)
            .order_by(Warehouse.code.asc())
        ).all()
        return [
            InventoryLocationSummary(
                scope_id=row[0],
                scope_code=row[1],
                scope_name=row[2],
                current_quantity=Decimal(row[3] or 0),
                reserved_quantity=Decimal(row[4] or 0),
                available_quantity=Decimal(row[5] or 0),
                blocked_quantity=Decimal(row[6] or 0),
                damaged_quantity=Decimal(row[7] or 0),
                quarantine_quantity=Decimal(row[8] or 0),
                in_transit_quantity=Decimal(row[9] or 0),
            )
            for row in rows
        ]

    def create_inventory_record(
        self, data: InventoryCreate, *, firm_id: UUID, actor_id: UUID
    ) -> InventoryRecord:
        """Create a stock projection for a product location."""
        branch, warehouse, storage_node, product, profile = self._validate_references(
            firm_id=firm_id,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            storage_node_id=data.storage_node_id,
            product_id=data.product_id,
        )
        locator = self._storage_locator(storage_node.id if storage_node else None)
        if (
            self._find_inventory_row(
                firm_id=firm_id,
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                storage_locator=locator,
                product_id=product.id,
                # Creating a stock row by hand creates the untracked one; a
                # batch's row is created by the movement that receives it.
                batch_id=None,
            )
            is not None
        ):
            raise ConflictError("An inventory record already exists for this location.")
        row = InventoryRecord(
            firm_id=firm_id,
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            storage_node_id=storage_node.id if storage_node else None,
            storage_locator=locator,
            product_id=product.id,
            business_profile_id=profile.id if profile is not None else None,
            minimum_level=data.minimum_level,
            maximum_level=data.maximum_level,
            reorder_level=data.reorder_level,
            safety_stock=data.safety_stock,
            status=data.status.value,
            current_quantity=ZERO,
            reserved_quantity=ZERO,
            available_quantity=ZERO,
            blocked_quantity=ZERO,
            damaged_quantity=ZERO,
            quarantine_quantity=ZERO,
            in_transit_quantity=ZERO,
            created_by=actor_id,
            updated_by=actor_id,
        )
        self._session.add(row)
        self._session.flush()
        record_audit(
            self._session,
            action="inventory.created",
            entity_type="inventory",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_id,
            after_data={
                "product_id": str(product.id),
                "warehouse_id": str(warehouse.id),
            },
        )
        self._commit()
        self._session.refresh(row)
        return row

    def get_inventory_record(
        self, inventory_id: UUID, *, firm_scope: UUID, include_deleted: bool = False
    ) -> InventoryRecord:
        """Return one stock projection the firm owns."""
        statement = select(InventoryRecord).where(
            InventoryRecord.id == inventory_id, InventoryRecord.firm_id == firm_scope
        )
        if not include_deleted:
            statement = statement.where(InventoryRecord.is_deleted.is_(False))
        row = self._session.scalar(statement)
        if row is None:
            raise ResourceNotFoundError("Inventory record not found.")
        return row

    def update_inventory_record(
        self,
        inventory_id: UUID,
        data: InventoryUpdate,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> InventoryRecord:
        """Change a projection's thresholds and status."""
        row = self.get_inventory_record(
            inventory_id, firm_scope=firm_scope, include_deleted=True
        )
        self._validate_references(
            firm_id=firm_scope,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            storage_node_id=data.storage_node_id,
            product_id=data.product_id,
        )
        locator = self._storage_locator(data.storage_node_id)
        existing = self._find_inventory_row(
            firm_id=firm_scope,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            storage_locator=locator,
            product_id=data.product_id,
            # Editing a stock row cannot move it between batches, so the clash
            # to look for is on the row's own batch.
            batch_id=row.batch_id,
        )
        if existing is not None and existing.id != row.id:
            raise ConflictError("The target inventory location already exists.")
        before: dict[str, object] = {
            "minimum_level": str(row.minimum_level or ""),
            "maximum_level": str(row.maximum_level or ""),
            "reorder_level": str(row.reorder_level or ""),
            "safety_stock": str(row.safety_stock or ""),
            "status": row.status,
        }
        row.branch_id = data.branch_id
        row.warehouse_id = data.warehouse_id
        row.storage_node_id = data.storage_node_id
        row.storage_locator = locator
        row.product_id = data.product_id
        row.minimum_level = data.minimum_level
        row.maximum_level = data.maximum_level
        row.reorder_level = data.reorder_level
        row.safety_stock = data.safety_stock
        row.status = data.status.value
        row.updated_by = actor_id
        record_audit(
            self._session,
            action="inventory.updated",
            entity_type="inventory",
            entity_id=row.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data=before,
            after_data={"status": row.status},
        )
        self._commit()
        self._session.refresh(row)
        return row

    def list_transactions(
        self,
        *,
        firm_scope: UUID,
        filters: InventoryTransactionListFilters,
        page: int,
        page_size: int,
        search: str | None,
        sort_by: str,
        descending: bool,
    ) -> tuple[list[InventoryTransaction], int]:
        """Return a page of inventory movements."""
        columns = {
            "created_at": InventoryTransaction.created_at,
            "transaction_date": InventoryTransaction.transaction_date,
            "transaction_type": InventoryTransaction.transaction_type,
            "reference_number": InventoryTransaction.reference_number,
            "quantity": InventoryTransaction.quantity,
        }
        statement = (
            select(InventoryTransaction)
            .join(Product, Product.id == InventoryTransaction.product_id)
            .join(Branch, Branch.id == InventoryTransaction.branch_id)
            .join(Warehouse, Warehouse.id == InventoryTransaction.warehouse_id)
            .outerjoin(
                WarehouseStorageNode,
                WarehouseStorageNode.id == InventoryTransaction.storage_node_id,
            )
            .where(InventoryTransaction.firm_id == firm_scope)
        )
        count = (
            select(func.count())
            .select_from(InventoryTransaction)
            .join(Product, Product.id == InventoryTransaction.product_id)
            .join(Branch, Branch.id == InventoryTransaction.branch_id)
            .join(Warehouse, Warehouse.id == InventoryTransaction.warehouse_id)
            .outerjoin(
                WarehouseStorageNode,
                WarehouseStorageNode.id == InventoryTransaction.storage_node_id,
            )
            .where(InventoryTransaction.firm_id == firm_scope)
        )
        statement, count = self._apply_transaction_filters(statement, count, filters)
        if search:
            term = f"%{search.strip()}%"
            condition = or_(
                Product.code.ilike(term),
                Product.name.ilike(term),
                InventoryTransaction.reference_number.ilike(term),
                InventoryTransaction.reference_type.ilike(term),
                Warehouse.code.ilike(term),
                Warehouse.name.ilike(term),
            )
            statement = statement.where(condition)
            count = count.where(condition)
        ordering = columns.get(sort_by, InventoryTransaction.created_at)
        rows = self._session.scalars(
            statement.order_by(
                ordering.desc() if descending else ordering.asc(),
                # Tiebreaker: see list_ledger for why every paged query needs one.
                (
                    InventoryTransaction.id.desc()
                    if descending
                    else InventoryTransaction.id.asc()
                ),
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
        ).all()
        return list(rows), int(self._session.scalar(count) or 0)

    def list_ledger(
        self,
        *,
        firm_scope: UUID,
        filters: StockLedgerListFilters,
        page: int,
        page_size: int,
        search: str | None,
        sort_by: str,
        descending: bool,
    ) -> tuple[list[StockLedgerEntry], int]:
        """Return a page of immutable ledger rows."""
        columns = {
            "created_at": StockLedgerEntry.created_at,
            "transaction_date": StockLedgerEntry.transaction_date,
            "transaction_type": StockLedgerEntry.transaction_type,
            "reference_number": StockLedgerEntry.reference_number,
            "quantity": StockLedgerEntry.quantity,
        }
        statement = (
            select(StockLedgerEntry)
            .join(Product, Product.id == StockLedgerEntry.product_id)
            .join(Branch, Branch.id == StockLedgerEntry.branch_id)
            .join(Warehouse, Warehouse.id == StockLedgerEntry.warehouse_id)
            .outerjoin(
                WarehouseStorageNode,
                WarehouseStorageNode.id == StockLedgerEntry.storage_node_id,
            )
            .where(StockLedgerEntry.firm_id == firm_scope)
        )
        count = (
            select(func.count())
            .select_from(StockLedgerEntry)
            .join(Product, Product.id == StockLedgerEntry.product_id)
            .join(Branch, Branch.id == StockLedgerEntry.branch_id)
            .join(Warehouse, Warehouse.id == StockLedgerEntry.warehouse_id)
            .outerjoin(
                WarehouseStorageNode,
                WarehouseStorageNode.id == StockLedgerEntry.storage_node_id,
            )
            .where(StockLedgerEntry.firm_id == firm_scope)
        )
        statement, count = self._apply_ledger_filters(statement, count, filters)
        if search:
            term = f"%{search.strip()}%"
            condition = or_(
                Product.code.ilike(term),
                Product.name.ilike(term),
                StockLedgerEntry.reference_number.ilike(term),
                StockLedgerEntry.reference_type.ilike(term),
                Warehouse.code.ilike(term),
                Warehouse.name.ilike(term),
            )
            statement = statement.where(condition)
            count = count.where(condition)
        ordering = columns.get(sort_by, StockLedgerEntry.created_at)
        rows = self._session.scalars(
            statement.order_by(
                ordering.desc() if descending else ordering.asc(),
                # A sort column alone is not a total order. `created_at` is not
                # unique -- a dispatch writes its DISPATCH and UNRESERVE rows in
                # one flush, so they share a timestamp to the microsecond -- and
                # OFFSET/LIMIT over a tie is free to hand the same row to two
                # pages and never show another. Paging the seeded ledger showed
                # one row twice and hid one entirely until this tiebreaker.
                StockLedgerEntry.id.desc() if descending else StockLedgerEntry.id.asc(),
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
        ).all()
        return list(rows), int(self._session.scalar(count) or 0)

    def list_opening_stock_batches(
        self,
        *,
        firm_scope: UUID,
        filters: OpeningStockBatchListFilters,
        page: int,
        page_size: int,
        search: str | None,
        sort_by: str,
        descending: bool,
    ) -> tuple[list[OpeningStockBatch], int]:
        """Return a page of opening-stock batches."""
        columns = {
            "created_at": OpeningStockBatch.created_at,
            "posting_date": OpeningStockBatch.posting_date,
            "reference_number": OpeningStockBatch.reference_number,
            "status": OpeningStockBatch.status,
        }
        statement = (
            select(OpeningStockBatch)
            .join(Branch, Branch.id == OpeningStockBatch.branch_id)
            .join(Warehouse, Warehouse.id == OpeningStockBatch.warehouse_id)
            .where(OpeningStockBatch.firm_id == firm_scope)
            .options(selectinload(OpeningStockBatch.lines))
        )
        count = (
            select(func.count())
            .select_from(OpeningStockBatch)
            .join(Branch, Branch.id == OpeningStockBatch.branch_id)
            .join(Warehouse, Warehouse.id == OpeningStockBatch.warehouse_id)
            .where(OpeningStockBatch.firm_id == firm_scope)
        )
        statement, count = self._apply_opening_stock_filters(statement, count, filters)
        if search:
            term = f"%{search.strip()}%"
            condition = or_(
                OpeningStockBatch.reference_number.ilike(term),
                Branch.code.ilike(term),
                Branch.name.ilike(term),
                Warehouse.code.ilike(term),
                Warehouse.name.ilike(term),
            )
            statement = statement.where(condition)
            count = count.where(condition)
        ordering = columns.get(sort_by, OpeningStockBatch.created_at)
        rows = self._session.scalars(
            statement.order_by(
                ordering.desc() if descending else ordering.asc(),
                # Tiebreaker: see list_ledger for why every paged query needs one.
                (
                    OpeningStockBatch.id.desc()
                    if descending
                    else OpeningStockBatch.id.asc()
                ),
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
        ).all()
        return list(rows), int(self._session.scalar(count) or 0)

    def get_opening_stock_batch(
        self, batch_id: UUID, *, firm_scope: UUID, include_deleted: bool = False
    ) -> OpeningStockBatch:
        """Return one opening-stock batch."""
        statement = (
            select(OpeningStockBatch)
            .where(
                OpeningStockBatch.id == batch_id,
                OpeningStockBatch.firm_id == firm_scope,
            )
            .options(selectinload(OpeningStockBatch.lines))
        )
        if not include_deleted:
            statement = statement.where(OpeningStockBatch.is_deleted.is_(False))
        row = self._session.scalar(statement)
        if row is None:
            raise ResourceNotFoundError("Opening stock batch not found.")
        return row

    def create_opening_stock_batch(
        self,
        data: OpeningStockBatchCreate,
        *,
        firm_id: UUID,
        actor_id: UUID,
        source_format: str = "MANUAL",
    ) -> OpeningStockBatch:
        """Create a draft opening-stock batch."""
        self._validate_branch_warehouse_scope(
            firm_id=firm_id, branch_id=data.branch_id, warehouse_id=data.warehouse_id
        )
        self._assert_unique_opening_reference(firm_id, data.reference_number)
        batch = OpeningStockBatch(
            firm_id=firm_id,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            reference_number=data.reference_number.strip().upper(),
            posting_date=data.posting_date,
            remarks=data.remarks,
            source_format=source_format.upper(),
            status="DRAFT",
            created_by=actor_id,
            updated_by=actor_id,
        )
        batch.lines = self._build_opening_stock_lines(
            firm_id=firm_id,
            warehouse_id=data.warehouse_id,
            lines=data.lines,
            actor_id=actor_id,
        )
        self._session.add(batch)
        self._session.flush()
        record_audit(
            self._session,
            action="opening_stock.created",
            entity_type="opening_stock_batch",
            entity_id=batch.id,
            actor_id=actor_id,
            firm_id=firm_id,
            after_data={
                "reference_number": batch.reference_number,
                "line_count": len(batch.lines),
            },
        )
        self._commit()
        self._session.refresh(batch)
        return batch

    def update_opening_stock_batch(
        self,
        batch_id: UUID,
        data: OpeningStockUpdate,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> OpeningStockBatch:
        """Change a draft opening-stock batch."""
        batch = self.get_opening_stock_batch(
            batch_id, firm_scope=firm_scope, include_deleted=True
        )
        if batch.status != "DRAFT":
            raise ValidationError("Only draft opening stock batches can be edited.")
        self._validate_branch_warehouse_scope(
            firm_id=firm_scope, branch_id=data.branch_id, warehouse_id=data.warehouse_id
        )
        self._assert_unique_opening_reference(
            firm_scope, data.reference_number, excluding_id=batch.id
        )
        before: dict[str, object] = {
            "reference_number": batch.reference_number,
            "status": batch.status,
        }
        batch.branch_id = data.branch_id
        batch.warehouse_id = data.warehouse_id
        batch.reference_number = data.reference_number.strip().upper()
        batch.posting_date = data.posting_date
        batch.remarks = data.remarks
        batch.updated_by = actor_id
        for existing in list(batch.lines):
            self._session.delete(existing)
        batch.lines = self._build_opening_stock_lines(
            firm_id=firm_scope,
            warehouse_id=data.warehouse_id,
            lines=data.lines,
            actor_id=actor_id,
        )
        record_audit(
            self._session,
            action="opening_stock.updated",
            entity_type="opening_stock_batch",
            entity_id=batch.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            before_data=before,
            after_data={
                "reference_number": batch.reference_number,
                "line_count": len(batch.lines),
            },
        )
        self._commit()
        self._session.refresh(batch)
        return batch

    def _opening_stock_batch_id(
        self,
        line: OpeningStockLine,
        batch: OpeningStockBatch,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> UUID | None:
        """Resolve the batch one opening-stock line puts its goods in.

        Opening stock is stock arriving, so it behaves like a goods receipt:
        an unknown number registers the batch rather than being refused, and a
        product whose profile requires a batch on receipt is refused without
        one. Enforcing it here as well as on receipts is what makes the rule
        mean something -- otherwise a firm could take in untraceable stock on
        day one and never be able to say where it came from.

        Returns:
            The batch id, or None where the line names none and the product
            does not require one.

        """
        number = (line.batch_number or "").strip()
        if not number:
            product = self._session.get(Product, line.product_id)
            if product is not None and product.require_batch_on_receipt:
                raise ValidationError(
                    f"{product.code} must be taken in with a batch number, "
                    "including as opening stock."
                )
            return None
        # Imported here rather than at the top: `batch_serial` reads stock
        # totals from this service, so the two modules would import each other
        # at load. The dependency is real in both directions -- a batch is
        # identity and stock is quantity, and each needs the other's answer.
        from app.batch_serial.services import BatchSerialService

        resolved = BatchSerialService(self._session).resolve_for_receipt(
            firm_scope=firm_scope,
            actor_id=actor_id,
            product_id=line.product_id,
            batch_number=number,
            branch_id=batch.branch_id,
            warehouse_id=batch.warehouse_id,
            expiry_date=line.expiry_date,
        )
        line.batch_id = resolved.id
        return resolved.id

    def post_opening_stock_batch(
        self, batch_id: UUID, *, firm_scope: UUID, actor_id: UUID
    ) -> OpeningStockBatch:
        """Post an opening-stock batch into the ledger."""
        batch = self.get_opening_stock_batch(batch_id, firm_scope=firm_scope)
        if batch.status == "POSTED":
            raise ConflictError("Opening stock batch has already been posted.")
        if not batch.lines:
            raise ValidationError("Opening stock batch must contain at least one line.")
        movement_ids: list[UUID] = []
        for line in batch.lines:
            (
                base_quantity,
                entered_quantity,
                entered_uom_id,
                conversion_version,
            ) = self._resolve_base_quantity(
                firm_scope=firm_scope,
                product_id=line.product_id,
                quantity=line.entered_quantity or line.quantity,
                entered_uom_id=line.entered_uom_id,
                conversion_version=line.conversion_version,
                on_date=batch.posting_date,
            )
            # Day-one stock arrives in a batch like any other stock. It was
            # the last way stock could enter with no batch behind it, so a
            # pharmacy's opening shelf was one untraceable heap while every
            # later delivery was traced -- and a product that requires a batch
            # on issue could never ship what it started with.
            line_batch_id = self._opening_stock_batch_id(
                line, batch, firm_scope=firm_scope, actor_id=actor_id
            )
            inventory = self._ensure_inventory_projection(
                firm_id=firm_scope,
                branch_id=batch.branch_id,
                warehouse_id=batch.warehouse_id,
                storage_node_id=line.storage_node_id,
                product_id=line.product_id,
                actor_id=actor_id,
                batch_id=line_batch_id,
            )
            self._apply_thresholds(
                inventory,
                minimum_level=line.minimum_level,
                maximum_level=line.maximum_level,
                reorder_level=line.reorder_level,
                safety_stock=line.safety_stock,
                actor_id=actor_id,
            )
            transaction = self._stage_movement(
                inventory,
                actor_id=actor_id,
                movement=_Movement(
                    transaction_type=InventoryTransactionType.OPENING_STOCK.value,
                    batch_id=line_batch_id,
                    reference_number=batch.reference_number,
                    reference_type="OPENING_STOCK",
                    transaction_date=batch.posting_date,
                    quantity=base_quantity,
                    current_delta=base_quantity,
                    unit_cost=line.unit_cost,
                    entered_quantity=entered_quantity,
                    entered_uom_id=entered_uom_id,
                    conversion_version=conversion_version,
                    remarks=line.remarks or batch.remarks,
                ),
            )
            line.transaction_id = transaction.id
            line.updated_by = actor_id
            movement_ids.append(transaction.id)
        # Day-one stock arrived from nowhere the ledger can see, so it is
        # debited to inventory against opening balance equity. The flush is
        # required for the same reason it is on adjustments: request sessions
        # do not autoflush, so the rows staged above are invisible to a query
        # until they are written.
        self._session.flush()
        opening_value = self._session.scalar(
            select(func.coalesce(func.sum(StockLedgerEntry.total_cost), 0)).where(
                StockLedgerEntry.transaction_id.in_(movement_ids)
            )
        )
        DocumentPostingService(self._session).post_opening_stock(
            firm_id=firm_scope,
            batch_id=batch.id,
            reference_number=batch.reference_number,
            posting_date=batch.posting_date,
            stock_value=Decimal(str(opening_value or ZERO)),
            actor_id=actor_id,
        )
        batch.status = "POSTED"
        batch.posted_at = batch.posting_date
        batch.updated_by = actor_id
        record_audit(
            self._session,
            action="opening_stock.posted",
            entity_type="opening_stock_batch",
            entity_id=batch.id,
            actor_id=actor_id,
            firm_id=firm_scope,
            after_data={
                "reference_number": batch.reference_number,
                "line_count": len(batch.lines),
            },
        )
        self._commit()
        self._session.refresh(batch)
        return batch

    def import_opening_stock_json(
        self,
        payload: OpeningStockImportRequest,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> OpeningStockBatch:
        """Build opening-stock lines from a JSON payload."""
        batch = self.create_opening_stock_batch(
            OpeningStockBatchCreate(
                branch_id=payload.branch_id,
                warehouse_id=payload.warehouse_id,
                reference_number=payload.reference_number,
                posting_date=payload.posting_date,
                remarks=payload.remarks,
                lines=payload.lines,
            ),
            firm_id=firm_scope,
            actor_id=actor_id,
            source_format="JSON",
        )
        if payload.auto_post:
            return self.post_opening_stock_batch(
                batch.id, firm_scope=firm_scope, actor_id=actor_id
            )
        return batch

    def import_opening_stock_csv(
        self,
        csv_content: str,
        *,
        reference_number: str,
        posting_date: date,
        branch_id: UUID,
        warehouse_id: UUID,
        remarks: str | None,
        auto_post: bool,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> OpeningStockBatch:
        """Build opening-stock lines from a CSV upload."""
        import csv
        import io

        reader = csv.DictReader(io.StringIO(csv_content))
        lines: list[OpeningStockLineCreate] = []
        for row in reader:
            product_id = str(row.get("ProductId") or "").strip()
            quantity = str(row.get("Quantity") or "").strip()
            if not product_id or not quantity:
                continue
            lines.append(
                OpeningStockLineCreate(
                    product_id=UUID(product_id),
                    storage_node_id=(
                        UUID(str(row.get("StorageNodeId")).strip())
                        if row.get("StorageNodeId")
                        else None
                    ),
                    quantity=Decimal(quantity),
                    minimum_level=(
                        Decimal(str(row["MinimumLevel"]))
                        if row.get("MinimumLevel")
                        else None
                    ),
                    maximum_level=(
                        Decimal(str(row["MaximumLevel"]))
                        if row.get("MaximumLevel")
                        else None
                    ),
                    reorder_level=(
                        Decimal(str(row["ReorderLevel"]))
                        if row.get("ReorderLevel")
                        else None
                    ),
                    safety_stock=(
                        Decimal(str(row["SafetyStock"]))
                        if row.get("SafetyStock")
                        else None
                    ),
                    remarks=(row.get("Remarks") or "").strip() or None,
                )
            )
        return self.import_opening_stock_json(
            OpeningStockImportRequest(
                reference_number=reference_number,
                posting_date=posting_date,
                branch_id=branch_id,
                warehouse_id=warehouse_id,
                remarks=remarks,
                auto_post=auto_post,
                lines=lines,
            ),
            firm_scope=firm_scope,
            actor_id=actor_id,
        )

    def import_opening_stock_xlsx(
        self,
        workbook_bytes: bytes,
        *,
        reference_number: str,
        posting_date: date,
        branch_id: UUID,
        warehouse_id: UUID,
        remarks: str | None,
        auto_post: bool,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> OpeningStockBatch:
        """Build opening-stock lines from an XLSX upload."""
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError as error:
            raise ValidationError(
                "XLSX import dependency is unavailable. Install openpyxl."
            ) from error
        workbook = load_workbook(filename=BytesIO(workbook_bytes), read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("The XLSX import file is empty.")
        header = [str(value or "").strip() for value in rows[0]]
        index = {name: position for position, name in enumerate(header)}
        lines: list[OpeningStockLineCreate] = []
        for values in rows[1:]:
            product_id = str(values[index.get("ProductId", -1)] or "").strip()
            quantity = str(values[index.get("Quantity", -1)] or "").strip()
            if not product_id or not quantity:
                continue
            lines.append(
                OpeningStockLineCreate(
                    product_id=UUID(product_id),
                    storage_node_id=(
                        UUID(str(values[index["StorageNodeId"]]).strip())
                        if index.get("StorageNodeId", -1) >= 0
                        and values[index["StorageNodeId"]] is not None
                        else None
                    ),
                    quantity=Decimal(quantity),
                    minimum_level=(
                        Decimal(str(values[index["MinimumLevel"]]))
                        if index.get("MinimumLevel", -1) >= 0
                        and values[index["MinimumLevel"]] is not None
                        else None
                    ),
                    maximum_level=(
                        Decimal(str(values[index["MaximumLevel"]]))
                        if index.get("MaximumLevel", -1) >= 0
                        and values[index["MaximumLevel"]] is not None
                        else None
                    ),
                    reorder_level=(
                        Decimal(str(values[index["ReorderLevel"]]))
                        if index.get("ReorderLevel", -1) >= 0
                        and values[index["ReorderLevel"]] is not None
                        else None
                    ),
                    safety_stock=(
                        Decimal(str(values[index["SafetyStock"]]))
                        if index.get("SafetyStock", -1) >= 0
                        and values[index["SafetyStock"]] is not None
                        else None
                    ),
                    remarks=(
                        str(values[index["Remarks"]]).strip()
                        if index.get("Remarks", -1) >= 0
                        and values[index["Remarks"]] is not None
                        else None
                    ),
                )
            )
        return self.import_opening_stock_json(
            OpeningStockImportRequest(
                reference_number=reference_number,
                posting_date=posting_date,
                branch_id=branch_id,
                warehouse_id=warehouse_id,
                remarks=remarks,
                auto_post=auto_post,
                lines=lines,
            ),
            firm_scope=firm_scope,
            actor_id=actor_id,
        )

    def create_adjustment(
        self,
        data: InventoryAdjustmentCreate,
        *,
        firm_scope: UUID,
        actor_id: UUID,
    ) -> InventoryTransaction:
        """Post a stock adjustment movement."""
        (
            base_quantity,
            entered_quantity,
            entered_uom_id,
            conversion_version,
        ) = self._resolve_base_quantity(
            firm_scope=firm_scope,
            product_id=data.product_id,
            quantity=(
                data.entered_quantity
                if data.entered_quantity is not None
                else data.quantity
            ),
            entered_uom_id=data.entered_uom_id,
            conversion_version=None,
            on_date=data.transaction_date,
        )
        delta_sign = Decimal("1") if data.quantity >= 0 else Decimal("-1")
        delta = abs(base_quantity) * delta_sign
        inventory = self._ensure_inventory_projection(
            firm_id=firm_scope,
            branch_id=data.branch_id,
            warehouse_id=data.warehouse_id,
            storage_node_id=data.storage_node_id,
            product_id=data.product_id,
            actor_id=actor_id,
        )
        transaction = self._stage_movement(
            inventory,
            actor_id=actor_id,
            movement=_Movement(
                transaction_type=InventoryTransactionType.ADJUSTMENT.value,
                reference_number=data.reference_number.strip().upper(),
                reference_type=data.reference_type.strip().upper(),
                transaction_date=data.transaction_date,
                quantity=abs(base_quantity),
                current_delta=delta,
                entered_quantity=abs(entered_quantity),
                entered_uom_id=entered_uom_id,
                conversion_version=conversion_version,
                remarks=data.remarks,
            ),
        )
        # An adjustment is the movement with no paperwork behind it, so nothing
        # on screen would ever hint that the ledger had stopped agreeing with
        # the stock it controls. The value is taken from the stock ledger row
        # the movement just wrote, signed by the direction stock went.
        #
        # The flush is required, not defensive: request sessions are built with
        # `autoflush=False`, so the row staged above is invisible to a query
        # until it is written. Without it the value read as nothing and the
        # adjustment posted no journal at all -- which unit tests could not
        # catch, because their session factory autoflushes by default.
        self._session.flush()
        entry = self._session.scalar(
            select(StockLedgerEntry).where(
                StockLedgerEntry.transaction_id == transaction.id
            )
        )
        movement_value = Decimal(str(entry.total_cost or ZERO)) if entry else ZERO
        DocumentPostingService(self._session).post_stock_adjustment(
            firm_id=firm_scope,
            transaction_id=transaction.id,
            reference_number=data.reference_number.strip().upper(),
            transaction_date=data.transaction_date,
            value_delta=movement_value if delta >= ZERO else -movement_value,
            actor_id=actor_id,
            remarks=data.remarks,
        )
        self._commit()
        self._session.refresh(transaction)
        return transaction

    def record_goods_receipt(
        self,
        *,
        firm_scope: UUID,
        actor_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        reference_number: str,
        transaction_date: date,
        total_quantity: Decimal,
        blocked_quantity: Decimal = Decimal("0"),
        damaged_quantity: Decimal = Decimal("0"),
        entered_quantity: Decimal | None = None,
        entered_uom_id: UUID | None = None,
        conversion_version: int | None = None,
        remarks: str | None = None,
        unit_cost: Decimal | None = None,
        batch_id: UUID | None = None,
    ) -> InventoryTransaction:
        """Post the stock a goods receipt brought in.

        ``batch_id`` puts the goods in that batch's row rather than the
        product's single row, which is what makes two deliveries of one
        medicine countable apart.
        """
        (
            base_quantity,
            entered_quantity,
            entered_uom_id,
            conversion_version,
        ) = self._resolve_base_quantity(
            firm_scope=firm_scope,
            product_id=product_id,
            quantity=(
                entered_quantity if entered_quantity is not None else total_quantity
            ),
            entered_uom_id=entered_uom_id,
            conversion_version=conversion_version,
            on_date=transaction_date,
        )
        conversion_factor = (
            base_quantity / entered_quantity
            if entered_quantity not in {None, ZERO}
            else Decimal("1")
        )
        blocked_base = Decimal(str(blocked_quantity)) * conversion_factor
        damaged_base = Decimal(str(damaged_quantity)) * conversion_factor
        inventory = self._ensure_inventory_projection(
            firm_id=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_node_id=storage_node_id,
            product_id=product_id,
            actor_id=actor_id,
            batch_id=batch_id,
        )
        transaction = self._stage_movement(
            inventory,
            actor_id=actor_id,
            movement=_Movement(
                transaction_type=InventoryTransactionType.GOODS_RECEIPT.value,
                batch_id=batch_id,
                unit_cost=unit_cost,
                reference_number=reference_number.strip().upper(),
                reference_type="GOODS_RECEIPT",
                transaction_date=transaction_date,
                quantity=base_quantity,
                current_delta=base_quantity,
                blocked_delta=blocked_base + damaged_base,
                damaged_delta=damaged_base,
                entered_quantity=entered_quantity,
                entered_uom_id=entered_uom_id,
                conversion_version=conversion_version,
                remarks=remarks,
            ),
        )
        self._session.flush()
        return transaction

    def valuation_for(self, *, firm_scope: UUID, product_id: UUID) -> ProductValuation:
        """Return a product's running valuation, creating it on first use.

        Args:
            firm_scope: The owning firm.
            product_id: The product being valued.

        Returns:
            The valuation state row.

        """
        row = self._session.scalar(
            select(ProductValuation).where(
                ProductValuation.firm_id == firm_scope,
                ProductValuation.product_id == product_id,
                ProductValuation.is_deleted.is_(False),
            )
        )
        if row is None:
            row = ProductValuation(firm_id=firm_scope, product_id=product_id)
            self._session.add(row)
            self._session.flush()
        return row

    def _apply_valuation(
        self, inventory: InventoryRecord, movement: _Movement, actor_id: UUID
    ) -> tuple[Decimal | None, Decimal | None, Decimal]:
        """Roll the moving weighted average forward for one movement.

        Receipts move the average toward the price paid; issues leave it alone
        and consume at it, which is what makes the value released equal the cost
        of goods sold. A receipt with no stated cost is valued at the current
        average rather than at zero, so an unpriced movement cannot silently
        destroy the average.

        Args:
            inventory: The projection the movement applies to.
            movement: The staged movement.
            actor_id: The user responsible.

        Returns:
            The unit cost applied, the total cost of the movement, and the
            average cost after it.

        """
        valuation = self.valuation_for(
            firm_scope=inventory.firm_id, product_id=inventory.product_id
        )
        delta = movement.current_delta
        average = Decimal(str(valuation.average_cost))
        on_hand = Decimal(str(valuation.quantity_on_hand))

        if delta == ZERO:
            # Reservations and status moves shift no goods and no value.
            return None, None, average

        if delta > ZERO:
            unit_cost = average if movement.unit_cost is None else movement.unit_cost
            new_quantity = on_hand + delta
            new_value = Decimal(str(valuation.total_value)) + (delta * unit_cost)
            average = (new_value / new_quantity) if new_quantity != ZERO else ZERO
        else:
            unit_cost = average
            new_quantity = on_hand + delta
            new_value = Decimal(str(valuation.total_value)) + (delta * average)
            if new_quantity <= ZERO:
                # Stock fully issued: hold no residual value on nothing.
                new_quantity = new_quantity if new_quantity < ZERO else ZERO
                new_value = ZERO if new_quantity == ZERO else new_value

        valuation.quantity_on_hand = new_quantity
        valuation.total_value = quantize_money(new_value)
        valuation.average_cost = average
        valuation.updated_by = actor_id
        return unit_cost, quantize_money(abs(delta) * unit_cost), average

    def reverse_transaction(
        self,
        transaction_id: UUID,
        *,
        firm_scope: UUID,
        actor_id: UUID,
        reason: str | None = None,
    ) -> InventoryTransaction:
        """Post the exact inverse of an existing movement.

        Cancelling a document that already moved stock must put the stock back,
        otherwise a cancelled goods receipt or purchase return leaves phantom
        quantities behind. The reversal is itself an immutable movement linked to
        the original, so the ledger keeps both halves and cannot be replayed.

        Args:
            transaction_id: The movement to reverse.
            firm_scope: The firm that owns the movement.
            actor_id: The user performing the reversal.
            reason: Optional narration stored on the reversing movement.

        Returns:
            The reversing inventory transaction.

        Raises:
            ResourceNotFoundError: If the movement or its projection is missing.
            ValidationError: If the movement was already reversed.

        """
        original = self._session.scalar(
            select(InventoryTransaction).where(
                InventoryTransaction.id == transaction_id,
                InventoryTransaction.firm_id == firm_scope,
                InventoryTransaction.is_deleted.is_(False),
            )
        )
        if original is None:
            raise ResourceNotFoundError("Inventory transaction not found.")
        already_reversed = self._session.scalar(
            select(InventoryTransaction.id).where(
                InventoryTransaction.reversal_of_transaction_id == original.id,
                InventoryTransaction.is_deleted.is_(False),
            )
        )
        if already_reversed is not None:
            raise ValidationError("This inventory movement was already reversed.")
        inventory = self._session.get(InventoryRecord, original.inventory_id)
        if inventory is None:
            raise ResourceNotFoundError("Inventory record not found.")
        transaction = self._stage_movement(
            inventory,
            actor_id=actor_id,
            movement=_Movement(
                transaction_type=f"{original.transaction_type}{REVERSAL_SUFFIX}"[:40],
                reference_number=original.reference_number,
                reference_type=original.reference_type,
                transaction_date=original.transaction_date,
                quantity=-original.quantity,
                current_delta=-original.current_quantity_delta,
                reserved_delta=-original.reserved_quantity_delta,
                blocked_delta=-original.blocked_quantity_delta,
                damaged_delta=-original.damaged_quantity_delta,
                quarantine_delta=-original.quarantine_quantity_delta,
                in_transit_delta=-original.in_transit_quantity_delta,
                entered_quantity=(
                    -original.entered_quantity
                    if original.entered_quantity is not None
                    else None
                ),
                entered_uom_id=original.entered_uom_id,
                conversion_version=original.conversion_version,
                remarks=reason,
            ),
        )
        transaction.reversal_of_transaction_id = original.id
        self._session.flush()
        return transaction

    def record_purchase_return(
        self,
        *,
        firm_scope: UUID,
        actor_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        reference_number: str,
        transaction_date: date,
        return_quantity: Decimal,
        sellable_quantity: Decimal,
        damaged_quantity: Decimal = Decimal("0"),
        scrap_quantity: Decimal = Decimal("0"),
        quarantine_quantity: Decimal = Decimal("0"),
        entered_quantity: Decimal | None = None,
        entered_uom_id: UUID | None = None,
        conversion_version: int | None = None,
        remarks: str | None = None,
        batch_id: UUID | None = None,
    ) -> InventoryTransaction:
        """Post the stock a purchase return sent back, from one batch.

        ``batch_id`` takes the goods out of that batch's row rather than the
        product's untracked one. Without it a batch could be received, sold
        from, and then returned to the supplier off stock that was never in it.
        """
        (
            base_quantity,
            entered_quantity,
            entered_uom_id,
            conversion_version,
        ) = self._resolve_base_quantity(
            firm_scope=firm_scope,
            product_id=product_id,
            quantity=(
                entered_quantity if entered_quantity is not None else return_quantity
            ),
            entered_uom_id=entered_uom_id,
            conversion_version=conversion_version,
            on_date=transaction_date,
        )
        conversion_factor = (
            base_quantity / entered_quantity
            if entered_quantity not in {None, ZERO}
            else Decimal("1")
        )
        sellable_base = Decimal(str(sellable_quantity)) * conversion_factor
        damaged_base = Decimal(str(damaged_quantity)) * conversion_factor
        scrap_base = Decimal(str(scrap_quantity)) * conversion_factor
        quarantine_base = Decimal(str(quarantine_quantity)) * conversion_factor
        inventory = self._ensure_inventory_projection(
            firm_id=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_node_id=storage_node_id,
            product_id=product_id,
            actor_id=actor_id,
            batch_id=batch_id,
        )
        transaction = self._stage_movement(
            inventory,
            actor_id=actor_id,
            movement=_Movement(
                transaction_type=InventoryTransactionType.RETURN.value,
                batch_id=batch_id,
                reference_number=reference_number.strip().upper(),
                reference_type="PURCHASE_RETURN",
                transaction_date=transaction_date,
                quantity=base_quantity,
                current_delta=-base_quantity,
                blocked_delta=ZERO,
                damaged_delta=ZERO,
                quarantine_delta=ZERO,
                entered_quantity=entered_quantity,
                entered_uom_id=entered_uom_id,
                conversion_version=conversion_version,
                remarks=remarks
                or (
                    f"purchase_return buckets sellable={sellable_base} "
                    f"damaged={damaged_base} scrap={scrap_base} "
                    f"quarantine={quarantine_base}"
                ),
            ),
        )
        if sellable_base > base_quantity:
            raise ValidationError(
                "Sellable return quantity cannot exceed return quantity."
            )
        self._session.flush()
        return transaction

    def record_sales_order_reservation(
        self,
        *,
        firm_scope: UUID,
        actor_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        reference_number: str,
        transaction_date: date,
        reserve_quantity: Decimal,
        entered_quantity: Decimal | None = None,
        entered_uom_id: UUID | None = None,
        conversion_version: int | None = None,
        remarks: str | None = None,
        batch_id: UUID | None = None,
    ) -> InventoryTransaction:
        """Reserve stock for a sales order line, against one batch.

        ``batch_id`` holds that batch's stock rather than the product's
        untracked row. Callers split the line with
        ``allocate_for_reservation`` and call this once per batch; the part of
        a reservation no batch can cover is passed with no batch, because there
        is no batch behind it.
        """
        (
            base_quantity,
            entered_quantity,
            entered_uom_id,
            conversion_version,
        ) = self._resolve_base_quantity(
            firm_scope=firm_scope,
            product_id=product_id,
            quantity=(
                entered_quantity if entered_quantity is not None else reserve_quantity
            ),
            entered_uom_id=entered_uom_id,
            conversion_version=conversion_version,
            on_date=transaction_date,
        )
        inventory = self._ensure_inventory_projection(
            firm_id=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_node_id=storage_node_id,
            product_id=product_id,
            actor_id=actor_id,
            batch_id=batch_id,
        )
        transaction = self._stage_movement(
            inventory,
            actor_id=actor_id,
            movement=_Movement(
                transaction_type=InventoryTransactionType.RESERVE.value,
                batch_id=batch_id,
                reference_number=reference_number.strip().upper(),
                reference_type="SALES_ORDER",
                transaction_date=transaction_date,
                quantity=base_quantity,
                reserved_delta=base_quantity,
                entered_quantity=entered_quantity,
                entered_uom_id=entered_uom_id,
                conversion_version=conversion_version,
                remarks=remarks or "sales_order reservation",
            ),
        )
        self._session.flush()
        return transaction

    def release_sales_order_reservation(
        self,
        *,
        firm_scope: UUID,
        actor_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        reference_number: str,
        transaction_date: date,
        release_quantity: Decimal,
        entered_quantity: Decimal | None = None,
        entered_uom_id: UUID | None = None,
        conversion_version: int | None = None,
        remarks: str | None = None,
        batch_id: UUID | None = None,
    ) -> InventoryTransaction:
        """Release a sales order's reservation."""
        (
            base_quantity,
            entered_quantity,
            entered_uom_id,
            conversion_version,
        ) = self._resolve_base_quantity(
            firm_scope=firm_scope,
            product_id=product_id,
            quantity=(
                entered_quantity if entered_quantity is not None else release_quantity
            ),
            entered_uom_id=entered_uom_id,
            conversion_version=conversion_version,
            on_date=transaction_date,
        )
        inventory = self._ensure_inventory_projection(
            firm_id=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_node_id=storage_node_id,
            product_id=product_id,
            actor_id=actor_id,
            batch_id=batch_id,
        )
        transaction = self._stage_movement(
            inventory,
            actor_id=actor_id,
            movement=_Movement(
                transaction_type=InventoryTransactionType.UNRESERVE.value,
                batch_id=batch_id,
                reference_number=reference_number.strip().upper(),
                reference_type="SALES_ORDER",
                transaction_date=transaction_date,
                quantity=base_quantity,
                reserved_delta=-base_quantity,
                entered_quantity=entered_quantity,
                entered_uom_id=entered_uom_id,
                conversion_version=conversion_version,
                remarks=remarks or "sales_order reservation release",
            ),
        )
        self._session.flush()
        return transaction

    def allocate_for_reservation(
        self,
        *,
        firm_scope: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        quantity: Decimal,
    ) -> list[tuple[UUID | None, Decimal]]:
        """Choose which batches a sales order holds, earliest expiry first.

        Committing stock at approval is what stops two salespeople promising
        the same box, and until now it committed the *product*: the movement
        went to the untracked row whatever the goods were held in. A firm whose
        stock is all in batches then had reservations on a row with nothing in
        it, driving its available negative while the batch rows sat untouched
        and apparently free.

        Reserving in the same order the goods will ship in keeps the two
        halves of the sales flow talking about the same stock -- dispatch
        releases a batch's reservation and immediately draws from it, because
        both rank by expiry.

        **Short stock does not fail here.** An order may be taken for more than
        is on the shelf; that is a back order, and the reports count on it. The
        batches cover what they can and the remainder is returned as a single
        pair with no batch, which is the truth: there is no batch behind it.

        Returns:
            The batches to hold, in the order to hold them, and any uncovered
            remainder last under ``None``.

        """
        rows = self._expiry_ranked_rows(
            firm_scope=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            product_id=product_id,
            column=InventoryRecord.available_quantity,
        )
        outstanding = Decimal(str(quantity))
        allocation: list[tuple[UUID | None, Decimal]] = []
        for row in rows:
            if outstanding <= ZERO:
                break
            take = min(outstanding, Decimal(str(row.available_quantity)))
            if take <= ZERO:
                continue
            allocation.append((row.batch_id, take))
            outstanding -= take
        if outstanding > ZERO:
            allocation.append((None, outstanding))
        return allocation

    def allocate_for_release(
        self,
        *,
        firm_scope: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        quantity: Decimal,
    ) -> list[tuple[UUID | None, Decimal]]:
        """Choose which reservations to let go, earliest expiry first.

        The mirror of ``allocate_for_reservation``, and it has to walk the rows
        that actually hold a reservation rather than the ones holding stock:
        releasing a batch that was never held would drive its reserved
        quantity negative.

        Earliest expiry first again, so a dispatch that releases and then
        allocates frees exactly the batch it is about to draw from. Anything
        left over comes off the untracked row, which is where a reservation
        goes that no batch could cover.

        Returns:
            The reservations to release, in the order to release them.

        """
        rows = self._expiry_ranked_rows(
            firm_scope=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            product_id=product_id,
            column=InventoryRecord.reserved_quantity,
        )
        outstanding = Decimal(str(quantity))
        allocation: list[tuple[UUID | None, Decimal]] = []
        for row in rows:
            if outstanding <= ZERO:
                break
            take = min(outstanding, Decimal(str(row.reserved_quantity)))
            if take <= ZERO:
                continue
            allocation.append((row.batch_id, take))
            outstanding -= take
        if outstanding > ZERO:
            allocation.append((None, outstanding))
        return allocation

    def _expiry_ranked_rows(
        self,
        *,
        firm_scope: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        product_id: UUID,
        column: InstrumentedAttribute[Decimal],
    ) -> list[InventoryRecord]:
        """Return this product's stock rows, the batch nearest expiry first.

        Expiry is ranked explicitly rather than left to the backend's NULL
        ordering -- PostgreSQL sorts NULLs first in ASC and SQLite last, so a
        batch with no expiry date would be picked first on one and last on the
        other. A batch without an expiry is not urgent, so it goes last, and
        ties break on the batch id to keep two runs of the same allocation
        identical.
        """
        return list(
            self._session.scalars(
                select(InventoryRecord)
                .outerjoin(BatchRecord, BatchRecord.id == InventoryRecord.batch_id)
                .where(
                    InventoryRecord.firm_id == firm_scope,
                    InventoryRecord.branch_id == branch_id,
                    InventoryRecord.warehouse_id == warehouse_id,
                    InventoryRecord.product_id == product_id,
                    InventoryRecord.is_deleted.is_(False),
                    column > ZERO,
                )
                .order_by(
                    case((BatchRecord.expiry_date.is_(None), 1), else_=0).asc(),
                    BatchRecord.expiry_date.asc(),
                    InventoryRecord.batch_id.asc(),
                )
            ).all()
        )

    def allocate_for_dispatch(
        self,
        *,
        firm_scope: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        quantity: Decimal,
    ) -> list[tuple[UUID | None, Decimal]]:
        """Choose which batches a dispatch consumes, earliest expiry first.

        A product held in one bay can now be several stock rows, one per batch,
        so a single document line may have to come out of more than one of
        them: sixty strips from the batch expiring in March and forty from
        June. This returns the split as (batch_id, quantity) pairs, and the
        caller stages one movement per pair.

        First expiry, first out. Expiry is ranked explicitly rather than left
        to the backend's NULL ordering -- PostgreSQL sorts NULLs first in ASC
        and SQLite last, so a batch with no expiry date would be picked first
        on one and last on the other. A batch without an expiry is not urgent,
        so it goes last, and ties break on the batch id to keep two runs of the
        same dispatch identical.

        Untracked stock -- the row whose ``batch_id`` is NULL -- is returned as
        a single pair, so a product nobody tracks behaves exactly as it did.

        Returns:
            The batches to draw from, in the order to draw from them. Raises if
            the available stock across all of them is short.

        """
        rows = self._expiry_ranked_rows(
            firm_scope=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            product_id=product_id,
            column=InventoryRecord.available_quantity,
        )
        # `require_batch_on_issue` is the product saying its goods cannot leave
        # unidentified. Untracked stock -- the row whose batch is NULL -- is
        # exactly what that forbids, so it is dropped from the candidates and
        # the dispatch is short rather than silently shipping stock nobody can
        # trace.
        product = self._session.get(Product, product_id)
        if product is not None and product.require_batch_on_issue:
            rows = [row for row in rows if row.batch_id is not None]
        outstanding = Decimal(str(quantity))
        allocation: list[tuple[UUID | None, Decimal]] = []
        for row in rows:
            if outstanding <= ZERO:
                break
            take = min(outstanding, Decimal(str(row.available_quantity)))
            if take <= ZERO:
                continue
            allocation.append((row.batch_id, take))
            outstanding -= take
        if outstanding > ZERO:
            raise ValidationError(
                "Insufficient available stock to dispatch: short by "
                f"{outstanding}."
                + (
                    " This product may only be issued from a batch."
                    if product is not None and product.require_batch_on_issue
                    else ""
                )
            )
        return allocation

    def record_delivery_note_dispatch(
        self,
        *,
        firm_scope: UUID,
        actor_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        reference_number: str,
        transaction_date: date,
        dispatch_quantity: Decimal,
        entered_quantity: Decimal | None = None,
        entered_uom_id: UUID | None = None,
        conversion_version: int | None = None,
        remarks: str | None = None,
        batch_id: UUID | None = None,
    ) -> InventoryTransaction:
        """Post the stock a delivery note dispatched, from one batch.

        Callers holding a line that spans batches call
        ``allocate_for_dispatch`` first and then this once per allocated batch.
        """
        (
            base_quantity,
            entered_quantity,
            entered_uom_id,
            conversion_version,
        ) = self._resolve_base_quantity(
            firm_scope=firm_scope,
            product_id=product_id,
            quantity=(
                entered_quantity if entered_quantity is not None else dispatch_quantity
            ),
            entered_uom_id=entered_uom_id,
            conversion_version=conversion_version,
            on_date=transaction_date,
        )
        inventory = self._ensure_inventory_projection(
            firm_id=firm_scope,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_node_id=storage_node_id,
            product_id=product_id,
            actor_id=actor_id,
            batch_id=batch_id,
        )
        transaction = self._stage_movement(
            inventory,
            actor_id=actor_id,
            movement=_Movement(
                transaction_type=InventoryTransactionType.DISPATCH.value,
                batch_id=batch_id,
                reference_number=reference_number.strip().upper(),
                reference_type="DELIVERY_NOTE",
                transaction_date=transaction_date,
                quantity=base_quantity,
                current_delta=-base_quantity,
                entered_quantity=entered_quantity,
                entered_uom_id=entered_uom_id,
                conversion_version=conversion_version,
                remarks=remarks or "delivery_note dispatch",
            ),
        )
        self._session.flush()
        return transaction

    def export_inventory_csv(self, *, firm_scope: UUID, search: str | None) -> str:
        """Render the filtered projections as CSV."""
        rows, _ = self.list_inventory(
            firm_scope=firm_scope,
            filters=InventoryListFilters(include_deleted=False),
            page=1,
            page_size=5000,
            search=search,
            sort_by="product_code",
            descending=False,
        )
        output = [
            "ProductCode,ProductName,BranchCode,WarehouseCode,StorageNodeCode,Current,Available,Reserved,Blocked,Damaged,Quarantine,InTransit,ReorderLevel,Status"
        ]
        for item in rows:
            output.append(
                ",".join(
                    [
                        self._csv(item.product_id, item, "product_code"),
                        self._csv(item.product_id, item, "product_name"),
                        self._csv(item.branch_id, item, "branch_code"),
                        self._csv(item.warehouse_id, item, "warehouse_code"),
                        self._csv(item.storage_node_id, item, "storage_node_code"),
                        str(item.current_quantity),
                        str(item.available_quantity),
                        str(item.reserved_quantity),
                        str(item.blocked_quantity),
                        str(item.damaged_quantity),
                        str(item.quarantine_quantity),
                        str(item.in_transit_quantity),
                        str(item.reorder_level or ""),
                        item.status,
                    ]
                )
            )
        return "\n".join(output)

    def export_inventory_xlsx(self, *, firm_scope: UUID, search: str | None) -> bytes:
        """Render the filtered projections as XLSX."""
        try:
            from openpyxl import Workbook
        except ImportError as error:
            raise ValidationError(
                "XLSX export dependency is unavailable. Install openpyxl."
            ) from error
        rows, _ = self.list_inventory(
            firm_scope=firm_scope,
            filters=InventoryListFilters(include_deleted=False),
            page=1,
            page_size=5000,
            search=search,
            sort_by="product_code",
            descending=False,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventory"
        sheet.append(
            [
                "ProductCode",
                "ProductName",
                "BranchCode",
                "WarehouseCode",
                "StorageNodeCode",
                "Current",
                "Available",
                "Reserved",
                "Blocked",
                "Damaged",
                "Quarantine",
                "InTransit",
                "ReorderLevel",
                "Status",
            ]
        )
        for item in rows:
            sheet.append(
                [
                    self._lookup_product_code(item.product_id),
                    self._lookup_product_name(item.product_id),
                    self._lookup_branch_code(item.branch_id),
                    self._lookup_warehouse_code(item.warehouse_id),
                    self._lookup_storage_code(item.storage_node_id),
                    item.current_quantity,
                    item.available_quantity,
                    item.reserved_quantity,
                    item.blocked_quantity,
                    item.damaged_quantity,
                    item.quarantine_quantity,
                    item.in_transit_quantity,
                    item.reorder_level or 0,
                    item.status,
                ]
            )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def export_ledger_csv(self, *, firm_scope: UUID, search: str | None) -> str:
        """Render the filtered ledger rows as CSV."""
        rows, _ = self.list_ledger(
            firm_scope=firm_scope,
            filters=StockLedgerListFilters(),
            page=1,
            page_size=5000,
            search=search,
            sort_by="transaction_date",
            descending=False,
        )
        output = [
            "TransactionDate,TransactionType,ReferenceNumber,ProductCode,WarehouseCode,Quantity,CurrentDelta,ReservedDelta,BlockedDelta,DamagedDelta,QuarantineDelta,InTransitDelta,NewCurrent,NewAvailable"
        ]
        for item in rows:
            output.append(
                ",".join(
                    [
                        item.transaction_date.isoformat(),
                        item.transaction_type,
                        item.reference_number,
                        self._lookup_product_code(item.product_id),
                        self._lookup_warehouse_code(item.warehouse_id),
                        str(item.quantity),
                        str(item.current_quantity_delta),
                        str(item.reserved_quantity_delta),
                        str(item.blocked_quantity_delta),
                        str(item.damaged_quantity_delta),
                        str(item.quarantine_quantity_delta),
                        str(item.in_transit_quantity_delta),
                        str(item.new_current_quantity),
                        str(item.new_available_quantity),
                    ]
                )
            )
        return "\n".join(output)

    def export_ledger_xlsx(self, *, firm_scope: UUID, search: str | None) -> bytes:
        """Render the filtered ledger rows as XLSX."""
        try:
            from openpyxl import Workbook
        except ImportError as error:
            raise ValidationError(
                "XLSX export dependency is unavailable. Install openpyxl."
            ) from error
        rows, _ = self.list_ledger(
            firm_scope=firm_scope,
            filters=StockLedgerListFilters(),
            page=1,
            page_size=5000,
            search=search,
            sort_by="transaction_date",
            descending=False,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "StockLedger"
        sheet.append(
            [
                "TransactionDate",
                "TransactionType",
                "ReferenceNumber",
                "ProductCode",
                "WarehouseCode",
                "Quantity",
                "CurrentDelta",
                "ReservedDelta",
                "BlockedDelta",
                "DamagedDelta",
                "QuarantineDelta",
                "InTransitDelta",
                "NewCurrent",
                "NewAvailable",
            ]
        )
        for item in rows:
            sheet.append(
                [
                    item.transaction_date.isoformat(),
                    item.transaction_type,
                    item.reference_number,
                    self._lookup_product_code(item.product_id),
                    self._lookup_warehouse_code(item.warehouse_id),
                    item.quantity,
                    item.current_quantity_delta,
                    item.reserved_quantity_delta,
                    item.blocked_quantity_delta,
                    item.damaged_quantity_delta,
                    item.quarantine_quantity_delta,
                    item.in_transit_quantity_delta,
                    item.new_current_quantity,
                    item.new_available_quantity,
                ]
            )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def inventory_response(self, row: InventoryRecord) -> InventoryResponse:
        """Expose one stock projection."""
        batch_number, batch_expiry = self._lookup_batch(row.batch_id)
        return InventoryResponse.model_validate(
            {
                "batch_id": row.batch_id,
                "batch_number": batch_number,
                "batch_expiry_date": batch_expiry,
                "id": row.id,
                "firm_id": row.firm_id,
                "branch_id": row.branch_id,
                "branch_code": self._lookup_branch_code(row.branch_id),
                "branch_name": self._lookup_branch_name(row.branch_id),
                "warehouse_id": row.warehouse_id,
                "warehouse_code": self._lookup_warehouse_code(row.warehouse_id),
                "warehouse_name": self._lookup_warehouse_name(row.warehouse_id),
                "storage_node_id": row.storage_node_id,
                "storage_node_code": self._lookup_storage_code(row.storage_node_id),
                "storage_node_name": self._lookup_storage_name(row.storage_node_id),
                "product_id": row.product_id,
                "product_code": self._lookup_product_code(row.product_id),
                "product_name": self._lookup_product_name(row.product_id),
                "business_profile_id": row.business_profile_id,
                "business_profile_code": self._lookup_profile_code(
                    row.business_profile_id
                ),
                "current_quantity": row.current_quantity,
                "reserved_quantity": row.reserved_quantity,
                "available_quantity": row.available_quantity,
                "blocked_quantity": row.blocked_quantity,
                "damaged_quantity": row.damaged_quantity,
                "quarantine_quantity": row.quarantine_quantity,
                "in_transit_quantity": row.in_transit_quantity,
                "display_quantity": row.display_quantity,
                "display_uom_id": row.display_uom_id,
                "minimum_level": row.minimum_level,
                "maximum_level": row.maximum_level,
                "reorder_level": row.reorder_level,
                "safety_stock": row.safety_stock,
                "last_transaction_at": row.last_transaction_at,
                "status": row.status,
                "is_deleted": row.is_deleted,
                "created_at": row.created_at,
                "updated_at": row.updated_at,
            }
        )

    def transaction_response(
        self, row: InventoryTransaction
    ) -> InventoryTransactionResponse:
        """Expose one inventory movement."""
        payload = self._movement_payload(row)
        payload["entered_quantity"] = row.entered_quantity
        payload["entered_uom_id"] = row.entered_uom_id
        payload["conversion_version"] = row.conversion_version
        return InventoryTransactionResponse.model_validate(payload)

    def _movement_payload(
        self, row: InventoryTransaction | StockLedgerEntry
    ) -> dict[str, object]:
        """Build the fields a transaction and its ledger row have in common.

        The ledger row is not an ``InventoryTransaction``: it records the
        as-entered quantity under ``original_quantity``/``original_uom_id`` and
        has no ``conversion_version``. Reusing the transaction builder for it
        raised AttributeError, so ``GET /inventory/ledger`` failed for every
        firm that had ever moved stock.
        """
        return {
            "id": row.id,
            "inventory_id": row.inventory_id,
            "firm_id": row.firm_id,
            "branch_id": row.branch_id,
            "branch_code": self._lookup_branch_code(row.branch_id),
            "branch_name": self._lookup_branch_name(row.branch_id),
            "warehouse_id": row.warehouse_id,
            "warehouse_code": self._lookup_warehouse_code(row.warehouse_id),
            "warehouse_name": self._lookup_warehouse_name(row.warehouse_id),
            "storage_node_id": row.storage_node_id,
            "storage_node_code": self._lookup_storage_code(row.storage_node_id),
            "storage_node_name": self._lookup_storage_name(row.storage_node_id),
            "product_id": row.product_id,
            "product_code": self._lookup_product_code(row.product_id),
            "product_name": self._lookup_product_name(row.product_id),
            "batch_id": row.batch_id,
            "batch_number": self._lookup_batch(row.batch_id)[0],
            "business_profile_id": row.business_profile_id,
            "transaction_type": row.transaction_type,
            "reference_number": row.reference_number,
            "reference_type": row.reference_type,
            "transaction_date": row.transaction_date,
            "quantity": row.quantity,
            "current_quantity_delta": row.current_quantity_delta,
            "reserved_quantity_delta": row.reserved_quantity_delta,
            "blocked_quantity_delta": row.blocked_quantity_delta,
            "damaged_quantity_delta": row.damaged_quantity_delta,
            "quarantine_quantity_delta": row.quarantine_quantity_delta,
            "in_transit_quantity_delta": row.in_transit_quantity_delta,
            "previous_current_quantity": row.previous_current_quantity,
            "new_current_quantity": row.new_current_quantity,
            "previous_reserved_quantity": row.previous_reserved_quantity,
            "new_reserved_quantity": row.new_reserved_quantity,
            "previous_available_quantity": row.previous_available_quantity,
            "new_available_quantity": row.new_available_quantity,
            "previous_blocked_quantity": row.previous_blocked_quantity,
            "new_blocked_quantity": row.new_blocked_quantity,
            "previous_damaged_quantity": row.previous_damaged_quantity,
            "new_damaged_quantity": row.new_damaged_quantity,
            "previous_quarantine_quantity": row.previous_quarantine_quantity,
            "new_quarantine_quantity": row.new_quarantine_quantity,
            "previous_in_transit_quantity": row.previous_in_transit_quantity,
            "new_in_transit_quantity": row.new_in_transit_quantity,
            "remarks": row.remarks,
            "created_at": row.created_at,
        }

    def ledger_response(self, row: StockLedgerEntry) -> StockLedgerResponse:
        """Expose one immutable ledger row."""
        payload = self._movement_payload(row)
        payload["transaction_id"] = row.transaction_id
        payload["entered_quantity"] = row.original_quantity
        payload["entered_uom_id"] = row.original_uom_id
        return StockLedgerResponse.model_validate(payload)

    def opening_stock_batch_response(
        self, row: OpeningStockBatch
    ) -> OpeningStockBatchResponse:
        """Expose one opening-stock batch."""
        return OpeningStockBatchResponse.model_validate(
            {
                "id": row.id,
                "firm_id": row.firm_id,
                "branch_id": row.branch_id,
                "branch_code": self._lookup_branch_code(row.branch_id),
                "branch_name": self._lookup_branch_name(row.branch_id),
                "warehouse_id": row.warehouse_id,
                "warehouse_code": self._lookup_warehouse_code(row.warehouse_id),
                "warehouse_name": self._lookup_warehouse_name(row.warehouse_id),
                "reference_number": row.reference_number,
                "posting_date": row.posting_date,
                "source_format": row.source_format,
                "status": row.status,
                "remarks": row.remarks,
                "posted_at": row.posted_at,
                "lines": [
                    self._opening_stock_line_response(line)
                    for line in row.lines
                    if not line.is_deleted
                ],
                "created_at": row.created_at,
                "updated_at": row.updated_at,
            }
        )

    def _opening_stock_line_response(
        self, row: OpeningStockLine
    ) -> OpeningStockLineResponse:
        return OpeningStockLineResponse.model_validate(
            {
                "id": row.id,
                "line_number": row.line_number,
                "product_id": row.product_id,
                "product_code": self._lookup_product_code(row.product_id),
                "product_name": self._lookup_product_name(row.product_id),
                "storage_node_id": row.storage_node_id,
                "storage_node_code": self._lookup_storage_code(row.storage_node_id),
                "storage_node_name": self._lookup_storage_name(row.storage_node_id),
                "business_profile_id": row.business_profile_id,
                "quantity": row.quantity,
                "entered_quantity": row.entered_quantity,
                "entered_uom_id": row.entered_uom_id,
                "conversion_version": row.conversion_version,
                "batch_number": row.batch_number,
                "batch_id": row.batch_id,
                "expiry_date": row.expiry_date,
                "minimum_level": row.minimum_level,
                "maximum_level": row.maximum_level,
                "reorder_level": row.reorder_level,
                "safety_stock": row.safety_stock,
                "remarks": row.remarks,
                "transaction_id": row.transaction_id,
            }
        )

    def _apply_inventory_filters(
        self,
        statement: Select[Any],
        count: Select[Any],
        filters: InventoryListFilters,
    ) -> tuple[Select[Any], Select[Any]]:
        if not filters.include_deleted:
            statement = statement.where(InventoryRecord.is_deleted.is_(False))
            count = count.where(InventoryRecord.is_deleted.is_(False))
        if filters.status is not None:
            statement = statement.where(InventoryRecord.status == filters.status.value)
            count = count.where(InventoryRecord.status == filters.status.value)
        for field, value in {
            InventoryRecord.branch_id: filters.branch_id,
            InventoryRecord.warehouse_id: filters.warehouse_id,
            InventoryRecord.storage_node_id: filters.storage_node_id,
            InventoryRecord.product_id: filters.product_id,
            InventoryRecord.business_profile_id: filters.business_profile_id,
        }.items():
            if value is not None:
                statement = statement.where(field == value)
                count = count.where(field == value)
        if filters.low_stock_only:
            condition = InventoryRecord.current_quantity <= func.coalesce(
                InventoryRecord.reorder_level, InventoryRecord.minimum_level, ZERO
            )
            statement = statement.where(condition)
            count = count.where(condition)
        if filters.out_of_stock_only:
            statement = statement.where(InventoryRecord.current_quantity <= 0)
            count = count.where(InventoryRecord.current_quantity <= 0)
        if filters.negative_only:
            condition = or_(
                InventoryRecord.current_quantity < 0,
                InventoryRecord.available_quantity < 0,
            )
            statement = statement.where(condition)
            count = count.where(condition)
        return statement, count

    def _apply_transaction_filters(
        self,
        statement: Select[Any],
        count: Select[Any],
        filters: InventoryTransactionListFilters,
    ) -> tuple[Select[Any], Select[Any]]:
        if filters.transaction_type is not None:
            statement = statement.where(
                InventoryTransaction.transaction_type == filters.transaction_type
            )
            count = count.where(
                InventoryTransaction.transaction_type == filters.transaction_type
            )
        for field, value in {
            InventoryTransaction.branch_id: filters.branch_id,
            InventoryTransaction.warehouse_id: filters.warehouse_id,
            InventoryTransaction.storage_node_id: filters.storage_node_id,
            InventoryTransaction.product_id: filters.product_id,
        }.items():
            if value is not None:
                statement = statement.where(field == value)
                count = count.where(field == value)
        if filters.reference_number:
            statement = statement.where(
                InventoryTransaction.reference_number.ilike(
                    f"%{filters.reference_number.strip()}%"
                )
            )
            count = count.where(
                InventoryTransaction.reference_number.ilike(
                    f"%{filters.reference_number.strip()}%"
                )
            )
        if filters.reference_type:
            statement = statement.where(
                InventoryTransaction.reference_type
                == filters.reference_type.strip().upper()
            )
            count = count.where(
                InventoryTransaction.reference_type
                == filters.reference_type.strip().upper()
            )
        if filters.transaction_from is not None:
            statement = statement.where(
                InventoryTransaction.transaction_date >= filters.transaction_from
            )
            count = count.where(
                InventoryTransaction.transaction_date >= filters.transaction_from
            )
        if filters.transaction_to is not None:
            statement = statement.where(
                InventoryTransaction.transaction_date <= filters.transaction_to
            )
            count = count.where(
                InventoryTransaction.transaction_date <= filters.transaction_to
            )
        return statement, count

    def _apply_ledger_filters(
        self,
        statement: Select[Any],
        count: Select[Any],
        filters: StockLedgerListFilters,
    ) -> tuple[Select[Any], Select[Any]]:
        if filters.transaction_type is not None:
            statement = statement.where(
                StockLedgerEntry.transaction_type == filters.transaction_type
            )
            count = count.where(
                StockLedgerEntry.transaction_type == filters.transaction_type
            )
        for field, value in {
            StockLedgerEntry.branch_id: filters.branch_id,
            StockLedgerEntry.warehouse_id: filters.warehouse_id,
            StockLedgerEntry.storage_node_id: filters.storage_node_id,
            StockLedgerEntry.product_id: filters.product_id,
        }.items():
            if value is not None:
                statement = statement.where(field == value)
                count = count.where(field == value)
        if filters.reference_number:
            statement = statement.where(
                StockLedgerEntry.reference_number.ilike(
                    f"%{filters.reference_number.strip()}%"
                )
            )
            count = count.where(
                StockLedgerEntry.reference_number.ilike(
                    f"%{filters.reference_number.strip()}%"
                )
            )
        if filters.reference_type:
            statement = statement.where(
                StockLedgerEntry.reference_type
                == filters.reference_type.strip().upper()
            )
            count = count.where(
                StockLedgerEntry.reference_type
                == filters.reference_type.strip().upper()
            )
        if filters.transaction_from is not None:
            statement = statement.where(
                StockLedgerEntry.transaction_date >= filters.transaction_from
            )
            count = count.where(
                StockLedgerEntry.transaction_date >= filters.transaction_from
            )
        if filters.transaction_to is not None:
            statement = statement.where(
                StockLedgerEntry.transaction_date <= filters.transaction_to
            )
            count = count.where(
                StockLedgerEntry.transaction_date <= filters.transaction_to
            )
        return statement, count

    def _apply_opening_stock_filters(
        self,
        statement: Select[Any],
        count: Select[Any],
        filters: OpeningStockBatchListFilters,
    ) -> tuple[Select[Any], Select[Any]]:
        if not filters.include_deleted:
            statement = statement.where(OpeningStockBatch.is_deleted.is_(False))
            count = count.where(OpeningStockBatch.is_deleted.is_(False))
        if filters.status is not None:
            statement = statement.where(
                OpeningStockBatch.status == filters.status.value
            )
            count = count.where(OpeningStockBatch.status == filters.status.value)
        if filters.branch_id is not None:
            statement = statement.where(
                OpeningStockBatch.branch_id == filters.branch_id
            )
            count = count.where(OpeningStockBatch.branch_id == filters.branch_id)
        if filters.warehouse_id is not None:
            statement = statement.where(
                OpeningStockBatch.warehouse_id == filters.warehouse_id
            )
            count = count.where(OpeningStockBatch.warehouse_id == filters.warehouse_id)
        if filters.posting_from is not None:
            statement = statement.where(
                OpeningStockBatch.posting_date >= filters.posting_from
            )
            count = count.where(OpeningStockBatch.posting_date >= filters.posting_from)
        if filters.posting_to is not None:
            statement = statement.where(
                OpeningStockBatch.posting_date <= filters.posting_to
            )
            count = count.where(OpeningStockBatch.posting_date <= filters.posting_to)
        return statement, count

    def _ensure_inventory_projection(
        self,
        *,
        firm_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
        actor_id: UUID,
        batch_id: UUID | None = None,
    ) -> InventoryRecord:
        _, _, storage_node, _, profile = self._validate_references(
            firm_id=firm_id,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_node_id=storage_node_id,
            product_id=product_id,
        )
        locator = self._storage_locator(storage_node.id if storage_node else None)
        row = self._find_inventory_row(
            firm_id=firm_id,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_locator=locator,
            product_id=product_id,
            batch_id=batch_id,
        )
        if row is not None:
            return row
        row = InventoryRecord(
            firm_id=firm_id,
            branch_id=branch_id,
            warehouse_id=warehouse_id,
            storage_node_id=storage_node.id if storage_node else None,
            storage_locator=locator,
            product_id=product_id,
            batch_id=batch_id,
            business_profile_id=profile.id if profile is not None else None,
            current_quantity=ZERO,
            reserved_quantity=ZERO,
            available_quantity=ZERO,
            blocked_quantity=ZERO,
            damaged_quantity=ZERO,
            quarantine_quantity=ZERO,
            in_transit_quantity=ZERO,
            display_quantity=ZERO,
            display_uom_id=self._default_display_uom_id(product_id),
            status="ACTIVE",
            created_by=actor_id,
            updated_by=actor_id,
        )
        self._session.add(row)
        self._session.flush()
        return row

    def _stage_movement(
        self, inventory: InventoryRecord, *, actor_id: UUID, movement: _Movement
    ) -> InventoryTransaction:
        previous_current = inventory.current_quantity
        previous_reserved = inventory.reserved_quantity
        previous_blocked = inventory.blocked_quantity
        previous_damaged = inventory.damaged_quantity
        previous_quarantine = inventory.quarantine_quantity
        previous_in_transit = inventory.in_transit_quantity
        previous_available = inventory.available_quantity

        new_current = previous_current + movement.current_delta
        new_reserved = previous_reserved + movement.reserved_delta
        new_blocked = previous_blocked + movement.blocked_delta
        new_damaged = previous_damaged + movement.damaged_delta
        new_quarantine = previous_quarantine + movement.quarantine_delta
        new_in_transit = previous_in_transit + movement.in_transit_delta
        new_available = self._available_quantity(
            current_quantity=new_current,
            reserved_quantity=new_reserved,
            blocked_quantity=new_blocked,
        )

        self._validate_non_negative_bucket("Reserved", new_reserved)
        self._validate_non_negative_bucket("Blocked", new_blocked)
        self._validate_non_negative_bucket("Damaged", new_damaged)
        self._validate_non_negative_bucket("Quarantine", new_quarantine)
        self._validate_non_negative_bucket("In transit", new_in_transit)

        inventory.current_quantity = new_current
        inventory.reserved_quantity = new_reserved
        inventory.available_quantity = new_available
        inventory.blocked_quantity = new_blocked
        inventory.damaged_quantity = new_damaged
        inventory.quarantine_quantity = new_quarantine
        inventory.in_transit_quantity = new_in_transit
        inventory.display_quantity = new_current
        if movement.entered_uom_id is not None:
            inventory.display_uom_id = movement.entered_uom_id
        elif inventory.display_uom_id is None:
            inventory.display_uom_id = self._default_display_uom_id(
                inventory.product_id
            )
        inventory.last_transaction_at = movement.transaction_date
        inventory.updated_by = actor_id

        unit_cost, total_cost, average_after = self._apply_valuation(
            inventory, movement, actor_id
        )
        transaction = InventoryTransaction(
            inventory_id=inventory.id,
            firm_id=inventory.firm_id,
            branch_id=inventory.branch_id,
            warehouse_id=inventory.warehouse_id,
            storage_node_id=inventory.storage_node_id,
            product_id=inventory.product_id,
            batch_id=movement.batch_id or inventory.batch_id,
            business_profile_id=inventory.business_profile_id,
            transaction_type=movement.transaction_type,
            reference_number=movement.reference_number,
            reference_type=movement.reference_type,
            transaction_date=movement.transaction_date,
            quantity=movement.quantity,
            current_quantity_delta=movement.current_delta,
            reserved_quantity_delta=movement.reserved_delta,
            blocked_quantity_delta=movement.blocked_delta,
            damaged_quantity_delta=movement.damaged_delta,
            quarantine_quantity_delta=movement.quarantine_delta,
            in_transit_quantity_delta=movement.in_transit_delta,
            previous_current_quantity=previous_current,
            new_current_quantity=new_current,
            previous_reserved_quantity=previous_reserved,
            new_reserved_quantity=new_reserved,
            previous_available_quantity=previous_available,
            new_available_quantity=new_available,
            previous_blocked_quantity=previous_blocked,
            new_blocked_quantity=new_blocked,
            previous_damaged_quantity=previous_damaged,
            new_damaged_quantity=new_damaged,
            previous_quarantine_quantity=previous_quarantine,
            new_quarantine_quantity=new_quarantine,
            previous_in_transit_quantity=previous_in_transit,
            new_in_transit_quantity=new_in_transit,
            remarks=movement.remarks,
            entered_quantity=movement.entered_quantity or movement.quantity,
            entered_uom_id=movement.entered_uom_id,
            conversion_version=movement.conversion_version,
            created_by=actor_id,
            updated_by=actor_id,
        )
        self._session.add(transaction)
        self._session.flush()
        self._session.add(
            StockLedgerEntry(
                transaction_id=transaction.id,
                inventory_id=inventory.id,
                firm_id=inventory.firm_id,
                branch_id=inventory.branch_id,
                warehouse_id=inventory.warehouse_id,
                storage_node_id=inventory.storage_node_id,
                product_id=inventory.product_id,
                batch_id=transaction.batch_id,
                business_profile_id=inventory.business_profile_id,
                transaction_type=transaction.transaction_type,
                reference_number=transaction.reference_number,
                reference_type=transaction.reference_type,
                transaction_date=transaction.transaction_date,
                quantity=transaction.quantity,
                current_quantity_delta=transaction.current_quantity_delta,
                reserved_quantity_delta=transaction.reserved_quantity_delta,
                blocked_quantity_delta=transaction.blocked_quantity_delta,
                damaged_quantity_delta=transaction.damaged_quantity_delta,
                quarantine_quantity_delta=transaction.quarantine_quantity_delta,
                in_transit_quantity_delta=transaction.in_transit_quantity_delta,
                previous_current_quantity=transaction.previous_current_quantity,
                new_current_quantity=transaction.new_current_quantity,
                previous_reserved_quantity=transaction.previous_reserved_quantity,
                new_reserved_quantity=transaction.new_reserved_quantity,
                previous_available_quantity=transaction.previous_available_quantity,
                new_available_quantity=transaction.new_available_quantity,
                previous_blocked_quantity=transaction.previous_blocked_quantity,
                new_blocked_quantity=transaction.new_blocked_quantity,
                previous_damaged_quantity=transaction.previous_damaged_quantity,
                new_damaged_quantity=transaction.new_damaged_quantity,
                previous_quarantine_quantity=transaction.previous_quarantine_quantity,
                new_quarantine_quantity=transaction.new_quarantine_quantity,
                previous_in_transit_quantity=transaction.previous_in_transit_quantity,
                new_in_transit_quantity=transaction.new_in_transit_quantity,
                remarks=transaction.remarks,
                unit_cost=unit_cost,
                total_cost=total_cost,
                average_cost_after=average_after,
                original_quantity=transaction.entered_quantity or transaction.quantity,
                original_uom_id=transaction.entered_uom_id,
                base_quantity=transaction.quantity,
                created_by=actor_id,
                updated_by=actor_id,
            )
        )
        record_audit(
            self._session,
            action="inventory.transaction.created",
            entity_type="inventory_transaction",
            entity_id=transaction.id,
            actor_id=actor_id,
            firm_id=inventory.firm_id,
            after_data={
                "inventory_id": str(inventory.id),
                "transaction_type": transaction.transaction_type,
                "reference_number": transaction.reference_number,
                "new_current_quantity": str(new_current),
                "new_available_quantity": str(new_available),
            },
        )
        return transaction

    def _build_opening_stock_lines(
        self,
        *,
        firm_id: UUID,
        warehouse_id: UUID,
        lines: Iterable[OpeningStockLineCreate],
        actor_id: UUID,
    ) -> list[OpeningStockLine]:
        items: list[OpeningStockLine] = []
        seen: set[tuple[UUID, str, str]] = set()
        profile = self._resolved_profile(firm_id)
        for index, line in enumerate(lines, start=1):
            product = self._session.scalar(
                select(Product).where(
                    Product.id == line.product_id,
                    Product.firm_id == firm_id,
                    Product.is_deleted.is_(False),
                )
            )
            if product is None:
                raise ValidationError(
                    "Opening stock line references an unknown product."
                )
            storage_node = None
            if line.storage_node_id is not None:
                storage_node = self._session.scalar(
                    select(WarehouseStorageNode).where(
                        WarehouseStorageNode.id == line.storage_node_id,
                        WarehouseStorageNode.warehouse_id == warehouse_id,
                        WarehouseStorageNode.is_deleted.is_(False),
                    )
                )
                if storage_node is None:
                    raise ValidationError(
                        "Opening stock line storage node does not belong to "
                        "the selected warehouse."
                    )
            locator = self._storage_locator(storage_node.id if storage_node else None)
            # The batch is part of the key, so one count of one shelf can
            # record two deliveries of a product expiring months apart.
            unique_key = (line.product_id, locator, (line.batch_number or "").strip())
            if unique_key in seen:
                raise ValidationError(
                    "Duplicate opening stock lines for the same product, "
                    "storage location and batch are not allowed."
                )
            seen.add(unique_key)
            items.append(
                OpeningStockLine(
                    line_number=index,
                    product_id=product.id,
                    storage_node_id=storage_node.id if storage_node else None,
                    storage_locator=locator,
                    business_profile_id=profile.id if profile is not None else None,
                    quantity=line.quantity,
                    unit_cost=line.unit_cost,
                    entered_quantity=line.entered_quantity,
                    entered_uom_id=line.entered_uom_id,
                    conversion_version=line.conversion_version,
                    batch_number=(line.batch_number or "").strip() or None,
                    expiry_date=line.expiry_date,
                    minimum_level=line.minimum_level,
                    maximum_level=line.maximum_level,
                    reorder_level=line.reorder_level,
                    safety_stock=line.safety_stock,
                    remarks=line.remarks,
                    created_by=actor_id,
                    updated_by=actor_id,
                )
            )
        return items

    def _find_inventory_row(
        self,
        *,
        firm_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_locator: str,
        product_id: UUID,
        batch_id: UUID | None,
    ) -> InventoryRecord | None:
        return self._session.scalar(
            select(InventoryRecord).where(
                InventoryRecord.firm_id == firm_id,
                InventoryRecord.branch_id == branch_id,
                InventoryRecord.warehouse_id == warehouse_id,
                InventoryRecord.storage_locator == storage_locator,
                InventoryRecord.product_id == product_id,
                # `== None` renders IS NULL, which is what selects the single
                # row an untracked product keeps. Comparing a nullable column
                # with `==` to a None variable is the intent here, not a
                # mistake -- batch-tracked stock and untracked stock are
                # different rows and must not collapse into one.
                InventoryRecord.batch_id == batch_id,
                InventoryRecord.is_deleted.is_(False),
            )
        )

    def _validate_references(
        self,
        *,
        firm_id: UUID,
        branch_id: UUID,
        warehouse_id: UUID,
        storage_node_id: UUID | None,
        product_id: UUID,
    ) -> tuple[
        Branch, Warehouse, WarehouseStorageNode | None, Product, BusinessProfile | None
    ]:
        branch = self._session.scalar(
            select(Branch).where(
                Branch.id == branch_id,
                Branch.firm_id == firm_id,
                Branch.is_deleted.is_(False),
            )
        )
        if branch is None:
            raise ValidationError("Branch does not belong to the active firm.")
        warehouse = self._session.scalar(
            select(Warehouse).where(
                Warehouse.id == warehouse_id,
                Warehouse.branch_id == branch.id,
                Warehouse.firm_id == firm_id,
                Warehouse.is_deleted.is_(False),
            )
        )
        if warehouse is None:
            raise ValidationError("Warehouse does not belong to the selected branch.")
        storage_node = None
        if storage_node_id is not None:
            storage_node = self._session.scalar(
                select(WarehouseStorageNode).where(
                    WarehouseStorageNode.id == storage_node_id,
                    WarehouseStorageNode.warehouse_id == warehouse.id,
                    WarehouseStorageNode.is_deleted.is_(False),
                )
            )
            if storage_node is None:
                raise ValidationError(
                    "Storage node does not belong to the selected warehouse."
                )
        product = self._session.scalar(
            select(Product).where(
                Product.id == product_id,
                Product.firm_id == firm_id,
                Product.is_deleted.is_(False),
            )
        )
        if product is None:
            raise ValidationError("Product does not belong to the active firm.")
        return branch, warehouse, storage_node, product, self._resolved_profile(firm_id)

    def _validate_branch_warehouse_scope(
        self, *, firm_id: UUID, branch_id: UUID, warehouse_id: UUID
    ) -> None:
        branch = self._session.scalar(
            select(Branch.id).where(
                Branch.id == branch_id,
                Branch.firm_id == firm_id,
                Branch.is_deleted.is_(False),
            )
        )
        if branch is None:
            raise ValidationError("Branch does not belong to the active firm.")
        warehouse = self._session.scalar(
            select(Warehouse.id).where(
                Warehouse.id == warehouse_id,
                Warehouse.firm_id == firm_id,
                Warehouse.branch_id == branch_id,
                Warehouse.is_deleted.is_(False),
            )
        )
        if warehouse is None:
            raise ValidationError("Warehouse does not belong to the selected branch.")

    def _assert_unique_opening_reference(
        self, firm_id: UUID, reference_number: str, excluding_id: UUID | None = None
    ) -> None:
        statement = select(OpeningStockBatch.id).where(
            OpeningStockBatch.firm_id == firm_id,
            OpeningStockBatch.reference_number == reference_number.strip().upper(),
            OpeningStockBatch.is_deleted.is_(False),
        )
        if excluding_id is not None:
            statement = statement.where(OpeningStockBatch.id != excluding_id)
        if self._session.scalar(statement) is not None:
            raise ConflictError("Opening stock reference number already exists.")

    def _resolved_profile(self, firm_id: UUID) -> BusinessProfile | None:
        assignment = self._session.scalar(
            select(FirmBusinessProfile).where(
                FirmBusinessProfile.firm_id == firm_id,
                FirmBusinessProfile.is_deleted.is_(False),
                FirmBusinessProfile.is_active.is_(True),
            )
        )
        if assignment is not None:
            profile = self._session.scalar(
                select(BusinessProfile).where(
                    BusinessProfile.id == assignment.business_profile_id,
                    BusinessProfile.is_deleted.is_(False),
                    BusinessProfile.status == "ACTIVE",
                )
            )
            if profile is not None:
                return profile
        return self._session.scalar(
            select(BusinessProfile).where(
                BusinessProfile.is_deleted.is_(False),
                BusinessProfile.status == "ACTIVE",
                BusinessProfile.is_default.is_(True),
            )
        )

    def _storage_locator(self, storage_node_id: UUID | None) -> str:
        return str(storage_node_id) if storage_node_id is not None else "ROOT"

    def _default_display_uom_id(self, product_id: UUID) -> UUID | None:
        product = self._session.scalar(
            select(Product).where(
                Product.id == product_id, Product.is_deleted.is_(False)
            )
        )
        if product is None:
            return None
        return product.inventory_uom_id or product.base_uom_id

    def _resolve_base_quantity(
        self,
        *,
        firm_scope: UUID,
        product_id: UUID,
        quantity: Decimal,
        entered_uom_id: UUID | None,
        conversion_version: int | None,
        on_date: date,
    ) -> tuple[Decimal, Decimal, UUID | None, int | None]:
        entered = Decimal(str(quantity))
        if entered_uom_id is None:
            return entered, entered, None, conversion_version
        product = self._session.scalar(
            select(Product).where(
                Product.id == product_id,
                Product.firm_id == firm_scope,
                Product.is_deleted.is_(False),
            )
        )
        if product is None:
            raise ValidationError("Transaction product is unavailable for conversion.")
        target_uom_id = product.base_uom_id or product.inventory_uom_id
        if target_uom_id is None or target_uom_id == entered_uom_id:
            return entered, entered, entered_uom_id, conversion_version
        statement = select(ConversionRule).where(
            ConversionRule.firm_id == firm_scope,
            ConversionRule.is_deleted.is_(False),
            ConversionRule.status == "ACTIVE",
            ConversionRule.from_uom_id == entered_uom_id,
            ConversionRule.to_uom_id == target_uom_id,
            ConversionRule.effective_from <= on_date,
            or_(
                ConversionRule.effective_to.is_(None),
                ConversionRule.effective_to >= on_date,
            ),
            or_(
                ConversionRule.product_id == product_id,
                ConversionRule.product_id.is_(None),
            ),
        )
        if conversion_version is not None:
            statement = statement.where(ConversionRule.version == conversion_version)
        rule = self._session.scalars(
            statement.order_by(
                ConversionRule.product_id.desc(), ConversionRule.version.desc()
            )
        ).first()
        if rule is None:
            raise ValidationError(
                "No active conversion rule is configured for the selected UOM."
            )
        base_quantity = entered * rule.conversion_factor
        return base_quantity, entered, entered_uom_id, rule.version

    def _available_quantity(
        self,
        *,
        current_quantity: Decimal,
        reserved_quantity: Decimal,
        blocked_quantity: Decimal,
    ) -> Decimal:
        return current_quantity - reserved_quantity - blocked_quantity

    def _validate_non_negative_bucket(self, label: str, value: Decimal) -> None:
        if value < 0:
            raise ValidationError(f"{label} quantity cannot become negative.")

    def _apply_thresholds(
        self,
        inventory: InventoryRecord,
        *,
        minimum_level: Decimal | None,
        maximum_level: Decimal | None,
        reorder_level: Decimal | None,
        safety_stock: Decimal | None,
        actor_id: UUID,
    ) -> None:
        if minimum_level is not None:
            inventory.minimum_level = minimum_level
        if maximum_level is not None:
            inventory.maximum_level = maximum_level
        if reorder_level is not None:
            inventory.reorder_level = reorder_level
        if safety_stock is not None:
            inventory.safety_stock = safety_stock
        inventory.updated_by = actor_id

    def _lookup_branch_code(self, branch_id: UUID) -> str:
        return str(
            self._session.scalar(select(Branch.code).where(Branch.id == branch_id))
            or ""
        )

    def _lookup_branch_name(self, branch_id: UUID) -> str:
        return str(
            self._session.scalar(select(Branch.name).where(Branch.id == branch_id))
            or ""
        )

    def _lookup_warehouse_code(self, warehouse_id: UUID) -> str:
        return str(
            self._session.scalar(
                select(Warehouse.code).where(Warehouse.id == warehouse_id)
            )
            or ""
        )

    def _lookup_warehouse_name(self, warehouse_id: UUID) -> str:
        return str(
            self._session.scalar(
                select(Warehouse.name).where(Warehouse.id == warehouse_id)
            )
            or ""
        )

    def _lookup_storage_code(self, storage_node_id: UUID | None) -> str | None:
        if storage_node_id is None:
            return None
        value = self._session.scalar(
            select(WarehouseStorageNode.code).where(
                WarehouseStorageNode.id == storage_node_id
            )
        )
        return str(value) if value is not None else None

    def _lookup_storage_name(self, storage_node_id: UUID | None) -> str | None:
        if storage_node_id is None:
            return None
        value = self._session.scalar(
            select(WarehouseStorageNode.name).where(
                WarehouseStorageNode.id == storage_node_id
            )
        )
        return str(value) if value is not None else None

    def _lookup_product_code(self, product_id: UUID) -> str:
        return str(
            self._session.scalar(select(Product.code).where(Product.id == product_id))
            or ""
        )

    def _lookup_product_name(self, product_id: UUID) -> str:
        return str(
            self._session.scalar(select(Product.name).where(Product.id == product_id))
            or ""
        )

    def _lookup_batch(self, batch_id: UUID | None) -> tuple[str | None, date | None]:
        """Return a batch's number and expiry, in one query rather than two."""
        if batch_id is None:
            return None, None
        row = self._session.execute(
            select(BatchRecord.batch_number, BatchRecord.expiry_date).where(
                BatchRecord.id == batch_id
            )
        ).first()
        if row is None:
            return None, None
        return row[0], row[1]

    def _lookup_profile_code(self, profile_id: UUID | None) -> str | None:
        if profile_id is None:
            return None
        value = self._session.scalar(
            select(BusinessProfile.code).where(BusinessProfile.id == profile_id)
        )
        return str(value) if value is not None else None

    def _csv(self, _id: UUID | None, row: InventoryRecord, attribute: str) -> str:
        response = self.inventory_response(row)
        value = getattr(response, attribute, None)
        return str(value or "")

    def _commit(self) -> None:
        try:
            self._session.commit()
        except IntegrityError as error:
            self._session.rollback()
            raise ConflictError(
                "The operation violates inventory uniqueness constraints."
            ) from error
