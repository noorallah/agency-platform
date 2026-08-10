"""Purchase backend validation, lifecycle, and API-scope tests."""

import asyncio
from datetime import date
from decimal import Decimal
from io import BytesIO
from uuid import UUID, uuid4

import pytest
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.datastructures import UploadFile

from app.batch_serial.models import batch_serial as _batch_models  # noqa: F401
from app.branches.models import Branch, Warehouse, WarehouseStorageNode
from app.business.models import BusinessFeature, BusinessProfile, ProfileFeature
from app.common.audit.models import AuditLog
from app.common.scope import (
    ResolvedFirmScope,
    optional_firm_scope,
    required_firm_scope,
)
from app.core.database.base import Base
from app.core.enums import TokenType
from app.core.exceptions import (
    AuthorizationError,
    ConflictError,
    ResourceNotFoundError,
    ValidationError,
)
from app.core.security.authorization import Principal, require_permission
from app.core.security.jwt import TokenClaims
from app.document_framework.models import (
    DocumentLifecycleEvent,
    DocumentNumberingRule,
    DocumentTypeDefinition,
)
from app.firms.models import Firm
from app.goods_receipt.models import GoodsReceipt
from app.identity.models import UserFirm
from app.inventory.models import inventory as _inventory_models  # noqa: F401
from app.products.models import Product
from app.purchase.api.router import (
    ActionReasonRequest,
    cancel_purchase_order,
    close_purchase_order,
    create_purchase_order,
    delete_purchase_order,
    export_purchase_orders,
    get_purchase_order,
    import_purchase_orders,
    list_purchase_orders,
    purchase_order_history,
    purchase_summary,
    restore_purchase_order,
    update_purchase_order,
)
from app.purchase.models import PurchaseOrderLine
from app.purchase.schemas import (
    PurchaseOrderCreate,
    PurchaseOrderImportRequest,
    PurchaseOrderStatus,
    PurchaseOrderUpdate,
)
from app.purchase.services import PurchaseService
from app.sales.models import GeoCountry
from app.search.services import SearchService
from app.tax.models import tax_framework as _tax_models  # noqa: F401
from app.tax.schemas import TaxComponentWrite, TaxProfileWrite, TaxSystemWrite
from app.tax.services import TaxFrameworkService
from app.uom.models import uom as _uom_models  # noqa: F401
from app.uom.schemas import ConversionRuleCreate, UomCreate
from app.uom.services import UomService
from app.vendors.models import Vendor


def _firm_scope(
    principal: Principal, session: Session, firm_id: UUID | None
) -> ResolvedFirmScope:
    """Resolve firm scope exactly as a request does, through the shared helper.

    Routers no longer carry a private resolver; membership is validated once in
    ``app.common.scope`` against the platform store.
    """
    return required_firm_scope(
        optional_firm_scope(principal=principal, db=session, x_firm_id=firm_id)
    )


def _session_factory() -> sessionmaker[Session]:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, expire_on_commit=False)


def _principal(
    user_id: UUID,
    permissions: set[str],
    *,
    firm_id: UUID | None = None,
    roles: set[str] | None = None,
) -> Principal:
    claim_roles = roles or set()
    return Principal(
        subject=user_id,
        roles=frozenset(claim_roles),
        permissions=frozenset(permissions),
        claims=TokenClaims(
            sub=str(user_id),
            type=TokenType.ACCESS,
            iat=1,
            exp=4_102_444_800,
            roles=sorted(claim_roles),
            permissions=sorted(permissions),
        ),
        firm_id=firm_id,
    )


def _firm(session: Session, code: str) -> Firm:
    row = Firm(
        name=f"{code} Firm",
        code=code,
        country="IN",
        currency_code="INR",
        financial_year_start=date(2026, 4, 1),
    )
    session.add(row)
    session.commit()
    return row


def _business_profile(session: Session, actor_id: UUID) -> BusinessProfile:
    """Seed the default profile, enabling the features these tests exercise.

    Purchase orders here carry attachments, and ATTACHMENTS is enforced, so a
    profile that does not enable it would refuse every one of them. Feature
    gating makes a profile's assignments load-bearing: a fixture has to grant
    what its documents use, the same as a real firm.
    """
    row = BusinessProfile(
        code="GENERIC",
        name="Generic",
        industry_type="GENERIC",
        status="ACTIVE",
        is_default=True,
        default_settings={},
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(row)
    session.flush()
    feature = BusinessFeature(
        code="ATTACHMENTS",
        name="Attachments",
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(feature)
    session.flush()
    session.add(
        ProfileFeature(
            business_profile_id=row.id,
            feature_id=feature.id,
            is_enabled=True,
            created_by=actor_id,
            updated_by=actor_id,
        )
    )
    session.commit()
    return row


def _country(session: Session, actor_id: UUID) -> GeoCountry:
    row = GeoCountry(
        code="IN",
        name="India",
        iso2="IN",
        iso3="IND",
        phone_code="+91",
        is_active=True,
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(row)
    session.commit()
    return row


def _branch(
    session: Session,
    *,
    firm_id: UUID,
    actor_id: UUID,
    code: str = "BR-001",
    status: str = "ACTIVE",
    business_profile_id: UUID | None = None,
) -> Branch:
    row = Branch(
        firm_id=firm_id,
        code=code,
        name=f"Branch {code}",
        display_name=f"Branch {code}",
        business_profile_id=business_profile_id,
        currency_code="INR",
        working_hours={"start": "09:00", "end": "18:00"},
        status=status,
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(row)
    session.commit()
    return row


def _warehouse(
    session: Session,
    *,
    firm_id: UUID,
    branch_id: UUID,
    actor_id: UUID,
    code: str = "WH-001",
    status: str = "ACTIVE",
    business_profile_id: UUID | None = None,
) -> Warehouse:
    row = Warehouse(
        firm_id=firm_id,
        branch_id=branch_id,
        code=code,
        name=f"Warehouse {code}",
        display_name=f"Warehouse {code}",
        business_profile_id=business_profile_id,
        status=status,
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(row)
    session.commit()
    return row


def _storage_node(
    session: Session,
    *,
    warehouse_id: UUID,
    actor_id: UUID,
    code: str = "BIN-01",
    is_active: bool = True,
) -> WarehouseStorageNode:
    row = WarehouseStorageNode(
        warehouse_id=warehouse_id,
        node_type="BIN",
        code=code,
        name=f"Bin {code}",
        path=f"/{code}",
        is_active=is_active,
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(row)
    session.commit()
    return row


def _vendor(
    session: Session,
    *,
    firm_id: UUID,
    actor_id: UUID,
    code: str = "VEN-001",
    status: str = "ACTIVE",
    business_profile_id: UUID | None = None,
    is_deleted: bool = False,
) -> Vendor:
    row = Vendor(
        firm_id=firm_id,
        code=code,
        name=f"Vendor {code}",
        display_name=f"Vendor {code}",
        business_profile_id=business_profile_id,
        status=status,
        is_deleted=is_deleted,
        deleted_by=actor_id if is_deleted else None,
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(row)
    session.commit()
    return row


def _product(
    session: Session,
    *,
    firm_id: UUID,
    actor_id: UUID,
    code: str = "SKU-001",
    status: str = "ACTIVE",
    tax_profile_group_code: str | None = None,
    base_uom_id: UUID | None = None,
    inventory_uom_id: UUID | None = None,
    purchase_uom_id: UUID | None = None,
    is_deleted: bool = False,
) -> Product:
    row = Product(
        firm_id=firm_id,
        code=code,
        name=f"Product {code}",
        product_type="STOCK_ITEM",
        tax_profile_group_code=tax_profile_group_code,
        base_uom_id=base_uom_id,
        inventory_uom_id=inventory_uom_id,
        purchase_uom_id=purchase_uom_id,
        status=status,
        is_deleted=is_deleted,
        deleted_by=actor_id if is_deleted else None,
        created_by=actor_id,
        updated_by=actor_id,
    )
    session.add(row)
    session.commit()
    return row


# Products reference tax profiles by the version-stable group code rather than the
# profile UUID, so a product keeps its mapping when a profile is superseded.
_TAX_PROFILE_GROUP_CODE = "VAT_5"


def _tax_profile(session: Session, *, firm_id: UUID, actor_id: UUID) -> UUID:
    framework = TaxFrameworkService(session)
    country = _country(session, actor_id)
    system = framework.create_system(
        TaxSystemWrite(
            country_id=country.id,
            code="VAT",
            name="Value Added Tax",
            display_name="VAT",
            status="ACTIVE",
        ),
        firm_id=firm_id,
        actor_id=actor_id,
    )
    component = framework.create_component(
        TaxComponentWrite(
            tax_system_id=system.id,
            code="VAT5",
            name="VAT 5",
            label="VAT 5%",
            percentage="5",
            status="ACTIVE",
        ),
        firm_id=firm_id,
        actor_id=actor_id,
    )
    profile = framework.create_profile(
        TaxProfileWrite(
            tax_system_id=system.id,
            code="VAT_5",
            name="VAT 5%",
            label="VAT 5%",
            status="ACTIVE",
            components=[
                {
                    "tax_component_id": component.id,
                    "percentage": "5",
                    "calculation_order": 1,
                    "included_in_price": False,
                    "recoverable": False,
                }
            ],
        ),
        firm_id=firm_id,
        actor_id=actor_id,
    )
    return profile.id


def _uom_pair(session: Session, *, firm_id: UUID, actor_id: UUID) -> tuple[UUID, UUID]:
    service = UomService(session)
    piece = service.create_uom(
        UomCreate(code="piece", name="Piece", symbol="pc"),
        actor_id=actor_id,
    )
    box = service.create_uom(UomCreate(code="box", name="Box"), actor_id=actor_id)
    service.create_conversion_rule(
        ConversionRuleCreate(
            from_uom_id=box.id,
            to_uom_id=piece.id,
            conversion_factor=Decimal("10"),
            effective_from=date(2026, 1, 1),
            version=1,
        ),
        firm_scope=firm_id,
        actor_id=actor_id,
    )
    return piece.id, box.id


def _purchase_data(
    *,
    branch_id: UUID,
    warehouse_id: UUID,
    vendor_id: UUID,
    product_id: UUID,
    purchase_uom_id: UUID | None = None,
    inventory_uom_id: UUID | None = None,
    tax_profile_id: UUID | None = None,
    storage_node_id: UUID | None = None,
    po_number: str = "PO-VAL-001",
    status: str = "DRAFT",
    ordered_quantity: str = "2",
    free_quantity: str = "1",
    unit_price: str = "10",
    discount_percent: str = "10",
    header_discount_amount: str = "1",
    additional_charges: str = "2",
    round_off: str = "0.5",
) -> PurchaseOrderCreate:
    return PurchaseOrderCreate.model_validate(
        {
            "po_number": po_number,
            "branch_id": branch_id,
            "warehouse_id": warehouse_id,
            "vendor_id": vendor_id,
            "purchase_date": "2026-08-02",
            "expected_delivery_date": "2026-08-10",
            "vendor_contact": "Accounts",
            "vendor_address": "Industrial Area",
            "department": "Procurement",
            "reference_number": f"REF-{po_number}",
            "external_reference": f"EXT-{po_number}",
            "remarks": "Validation order",
            "status": status,
            "header_discount_amount": header_discount_amount,
            "additional_charges": additional_charges,
            "round_off": round_off,
            "lines": [
                {
                    "product_id": product_id,
                    "purchase_uom_id": purchase_uom_id,
                    "inventory_uom_id": inventory_uom_id,
                    "ordered_quantity": ordered_quantity,
                    "free_quantity": free_quantity,
                    "unit_price": unit_price,
                    "discount_percent": discount_percent,
                    "tax_profile_id": tax_profile_id,
                    "warehouse_id": warehouse_id,
                    "storage_node_id": storage_node_id,
                    "remarks": "Primary line",
                }
            ],
            "delivery_schedules": [
                {
                    "line_number": 1,
                    "delivery_date": "2026-08-08",
                    "quantity": ordered_quantity,
                    "remarks": "Partial delivery",
                }
            ],
            "attachments": [
                {
                    "file_name": "quote.pdf",
                    "file_path": "/purchase/quote.pdf",
                    "mime_type": "application/pdf",
                }
            ],
            "notes": [{"note_type": "INTERNAL", "note": "Validate workflow-free PO."}],
        }
    )


async def _read_stream(response: object) -> bytes:
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk.encode("utf-8") if isinstance(chunk, str) else chunk)
    return b"".join(chunks)


def test_purchase_service_calculations_lifecycle_audit_and_history() -> None:
    session = _session_factory()()
    actor_id = uuid4()
    firm = _firm(session, "PO-SVC")
    profile = _business_profile(session, actor_id)
    branch = _branch(
        session,
        firm_id=firm.id,
        actor_id=actor_id,
        business_profile_id=profile.id,
    )
    warehouse = _warehouse(
        session,
        firm_id=firm.id,
        branch_id=branch.id,
        actor_id=actor_id,
        business_profile_id=profile.id,
    )
    storage = _storage_node(session, warehouse_id=warehouse.id, actor_id=actor_id)
    vendor = _vendor(
        session,
        firm_id=firm.id,
        actor_id=actor_id,
        business_profile_id=profile.id,
    )
    inventory_uom_id, purchase_uom_id = _uom_pair(
        session, firm_id=firm.id, actor_id=actor_id
    )
    tax_profile_id = _tax_profile(session, firm_id=firm.id, actor_id=actor_id)
    product = _product(
        session,
        firm_id=firm.id,
        actor_id=actor_id,
        tax_profile_group_code=_TAX_PROFILE_GROUP_CODE,
        base_uom_id=inventory_uom_id,
        inventory_uom_id=inventory_uom_id,
        purchase_uom_id=purchase_uom_id,
    )
    service = PurchaseService(session)

    created = service.create_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            purchase_uom_id=purchase_uom_id,
            inventory_uom_id=inventory_uom_id,
            tax_profile_id=tax_profile_id,
            storage_node_id=storage.id,
        ),
        firm_id=firm.id,
        actor_id=actor_id,
    )
    response = service.order_response(created)
    assert response.po_number == "PO-VAL-001"
    # subtotal is the taxable base: gross 20.00 less the 2.00 line discount.
    # It previously reported gross before discount, which no other document did.
    assert response.subtotal == Decimal("18.0000")
    assert response.line_discount_total == Decimal("2.0000")
    assert response.tax_total == Decimal("0.9000")
    # grand_total is unchanged by the redefinition; only the name's meaning was.
    assert response.grand_total == Decimal("20.4000")
    assert response.lines[0].base_quantity == Decimal("30.0000")
    assert response.lines[0].conversion_factor == Decimal("10")
    assert response.lines[0].conversion_version == 1
    assert len(response.delivery_schedules) == 1
    assert len(response.attachments) == 1
    assert len(response.notes) == 1
    assert session.scalar(
        select(DocumentTypeDefinition).where(
            DocumentTypeDefinition.firm_id == firm.id,
            DocumentTypeDefinition.code == "PURCHASE_ORDER",
        )
    )
    assert session.scalar(
        select(DocumentNumberingRule).where(
            DocumentNumberingRule.firm_id == firm.id,
            DocumentNumberingRule.code == "PURCHASE_ORDER_DEFAULT",
        )
    )
    events = session.scalars(
        select(DocumentLifecycleEvent).where(
            DocumentLifecycleEvent.firm_id == firm.id,
            DocumentLifecycleEvent.source_document_id == created.id,
        )
    ).all()
    assert [event.action for event in events] == ["CREATED"]

    updated = service.update_order(
        created.id,
        PurchaseOrderUpdate.model_validate(
            {
                **{
                    key: value
                    for key, value in _purchase_data(
                        branch_id=branch.id,
                        warehouse_id=warehouse.id,
                        vendor_id=vendor.id,
                        product_id=product.id,
                        purchase_uom_id=purchase_uom_id,
                        inventory_uom_id=inventory_uom_id,
                        tax_profile_id=tax_profile_id,
                        storage_node_id=storage.id,
                        status="SUBMITTED",
                        ordered_quantity="4",
                        free_quantity="0",
                        unit_price="12",
                        discount_percent="0",
                        header_discount_amount="2",
                        additional_charges="1",
                        round_off="0.5",
                    )
                    .model_dump(mode="json")
                    .items()
                    if key != "po_number"
                }
            },
        ),
        firm_scope=firm.id,
        actor_id=actor_id,
    )
    updated_response = service.order_response(updated)
    assert updated_response.status == PurchaseOrderStatus.SUBMITTED
    assert updated_response.subtotal == Decimal("48.0000")
    assert updated_response.tax_total == Decimal("2.4000")
    assert updated_response.grand_total == Decimal("49.9000")

    history = service.order_history(order_id=created.id, firm_scope=firm.id)
    assert [entry.action for entry in history] == [
        "purchase.created",
        "purchase.updated",
    ]

    audit_actions = session.scalars(
        select(AuditLog.action).where(AuditLog.entity_id == created.id)
    ).all()
    assert set(audit_actions) >= {"purchase.created", "purchase.updated"}

    deleted_order = service.create_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            tax_profile_id=tax_profile_id,
            po_number="PO-DEL-001",
        ),
        firm_id=firm.id,
        actor_id=actor_id,
    )
    service.delete_order(deleted_order.id, firm_scope=firm.id, actor_id=actor_id)
    with pytest.raises(ResourceNotFoundError):
        service.get_order(deleted_order.id, firm_scope=firm.id)
    restored = service.restore_order(
        deleted_order.id, firm_scope=firm.id, actor_id=actor_id
    )
    assert restored.is_deleted is False

    cancelled = service.create_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            tax_profile_id=tax_profile_id,
            po_number="PO-CAN-001",
        ),
        firm_id=firm.id,
        actor_id=actor_id,
    )
    service.cancel_order(
        cancelled.id,
        firm_scope=firm.id,
        actor_id=actor_id,
        reason="Vendor unavailable",
    )
    with pytest.raises(ValidationError, match="Cancelled/closed purchase orders"):
        service.update_order(
            cancelled.id,
            PurchaseOrderUpdate.model_validate(
                {
                    key: value
                    for key, value in _purchase_data(
                        branch_id=branch.id,
                        warehouse_id=warehouse.id,
                        vendor_id=vendor.id,
                        product_id=product.id,
                        tax_profile_id=tax_profile_id,
                        po_number="IGNORED",
                    )
                    .model_dump(mode="json")
                    .items()
                    if key != "po_number"
                }
            ),
            firm_scope=firm.id,
            actor_id=actor_id,
        )
    with pytest.raises(ValidationError, match="cannot be closed"):
        service.close_order(
            cancelled.id, firm_scope=firm.id, actor_id=actor_id, reason="Invalid"
        )

    closed = service.create_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            tax_profile_id=tax_profile_id,
            po_number="PO-CLOSE-001",
            status="APPROVED",
        ),
        firm_id=firm.id,
        actor_id=actor_id,
    )
    closed = service.close_order(
        closed.id, firm_scope=firm.id, actor_id=actor_id, reason="Completed"
    )
    assert closed.status == PurchaseOrderStatus.CLOSED.value
    assert closed.close_reason == "Completed"


def test_purchase_service_validations_multi_firm_search_and_import_duplicates() -> None:
    session = _session_factory()()
    actor_id = uuid4()
    first_firm = _firm(session, "PO-A")
    second_firm = _firm(session, "PO-B")
    profile = _business_profile(session, actor_id)
    branch = _branch(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        business_profile_id=profile.id,
    )
    warehouse = _warehouse(
        session,
        firm_id=first_firm.id,
        branch_id=branch.id,
        actor_id=actor_id,
        business_profile_id=profile.id,
    )
    storage = _storage_node(session, warehouse_id=warehouse.id, actor_id=actor_id)
    vendor = _vendor(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        business_profile_id=profile.id,
    )
    inventory_uom_id, purchase_uom_id = _uom_pair(
        session, firm_id=first_firm.id, actor_id=actor_id
    )
    tax_profile_id = _tax_profile(session, firm_id=first_firm.id, actor_id=actor_id)
    product = _product(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        tax_profile_group_code=_TAX_PROFILE_GROUP_CODE,
        base_uom_id=inventory_uom_id,
        inventory_uom_id=inventory_uom_id,
        purchase_uom_id=purchase_uom_id,
    )
    service = PurchaseService(session)

    created = service.create_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            purchase_uom_id=purchase_uom_id,
            inventory_uom_id=inventory_uom_id,
            tax_profile_id=tax_profile_id,
            storage_node_id=storage.id,
            po_number="PO-SEARCH-001",
        ),
        firm_id=first_firm.id,
        actor_id=actor_id,
    )
    with pytest.raises(ResourceNotFoundError):
        service.get_order(created.id, firm_scope=second_firm.id)
    with pytest.raises(ConflictError, match="already exists"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-SEARCH-001",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    inactive_branch = _branch(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        code="BR-INACTIVE",
        status="INACTIVE",
    )
    with pytest.raises(ValidationError, match="Inactive branches"):
        service.create_order(
            _purchase_data(
                branch_id=inactive_branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-ERR-001",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    inactive_warehouse = _warehouse(
        session,
        firm_id=first_firm.id,
        branch_id=branch.id,
        actor_id=actor_id,
        code="WH-INACTIVE",
        status="INACTIVE",
    )
    with pytest.raises(ValidationError, match="Inactive warehouses"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=inactive_warehouse.id,
                vendor_id=vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-ERR-002",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    inactive_vendor = _vendor(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        code="VEN-INACTIVE",
        status="BLOCKED",
    )
    with pytest.raises(ValidationError, match="Inactive or blocked vendors"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=inactive_vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-ERR-003",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    deleted_vendor = _vendor(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        code="VEN-DELETED",
        is_deleted=True,
    )
    with pytest.raises(ValidationError, match="vendor is not available"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=deleted_vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-ERR-004",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    inactive_product = _product(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        code="SKU-INACTIVE",
        status="INACTIVE",
    )
    with pytest.raises(ValidationError, match="Inactive/blocked products"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=inactive_product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-ERR-005",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    deleted_product = _product(
        session,
        firm_id=first_firm.id,
        actor_id=actor_id,
        code="SKU-DELETED",
        is_deleted=True,
    )
    with pytest.raises(ValidationError, match="product is not available"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=deleted_product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-ERR-006",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    inactive_storage = _storage_node(
        session,
        warehouse_id=warehouse.id,
        actor_id=actor_id,
        code="BIN-INACTIVE",
        is_active=False,
    )
    with pytest.raises(ValidationError, match="Inactive storage areas"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                storage_node_id=inactive_storage.id,
                po_number="PO-ERR-007",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    with pytest.raises(ValidationError, match="Selected tax profile"):
        service.create_order(
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=product.id,
                tax_profile_id=uuid4(),
                po_number="PO-ERR-008",
            ),
            firm_id=first_firm.id,
            actor_id=actor_id,
        )

    result = SearchService(session).search(
        query="PO-SEARCH",
        principal=_principal(uuid4(), {"PURCHASE_VIEW"}, firm_id=first_firm.id),
        category="masters",
        page=1,
        page_size=20,
    )
    assert result.total == 1
    assert result.results[0].entity_type == "purchase_orders"

    duplicate_import = PurchaseOrderImportRequest(
        records=[
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-IMP-001",
            ),
            _purchase_data(
                branch_id=branch.id,
                warehouse_id=warehouse.id,
                vendor_id=vendor.id,
                product_id=product.id,
                tax_profile_id=tax_profile_id,
                po_number="PO-IMP-001",
            ),
        ]
    )
    with pytest.raises(ConflictError, match="Duplicate purchase order numbers found"):
        service.import_orders(
            duplicate_import, firm_scope=first_firm.id, actor_id=actor_id
        )


def test_purchase_api_routes_import_export_summary_history_and_permissions() -> None:
    pytest.importorskip("openpyxl")
    factory = _session_factory()
    setup = factory()
    actor_id = uuid4()
    firm = _firm(setup, "PO-API")
    _business_profile(setup, actor_id)
    branch = _branch(setup, firm_id=firm.id, actor_id=actor_id)
    warehouse = _warehouse(
        setup, firm_id=firm.id, branch_id=branch.id, actor_id=actor_id
    )
    vendor = _vendor(setup, firm_id=firm.id, actor_id=actor_id)
    inventory_uom_id, purchase_uom_id = _uom_pair(
        setup, firm_id=firm.id, actor_id=actor_id
    )
    tax_profile_id = _tax_profile(setup, firm_id=firm.id, actor_id=actor_id)
    product = _product(
        setup,
        firm_id=firm.id,
        actor_id=actor_id,
        tax_profile_group_code=_TAX_PROFILE_GROUP_CODE,
        base_uom_id=inventory_uom_id,
        inventory_uom_id=inventory_uom_id,
        purchase_uom_id=purchase_uom_id,
    )
    setup.add(UserFirm(user_id=actor_id, firm_id=firm.id, is_active=True))
    setup.commit()
    setup.close()

    permissions = {
        "PURCHASE_VIEW",
        "PURCHASE_CREATE",
        "PURCHASE_UPDATE",
        "PURCHASE_DELETE",
        "PURCHASE_RESTORE",
        "PURCHASE_IMPORT",
        "PURCHASE_EXPORT",
        "PURCHASE_APPROVE",
        "PURCHASE_CANCEL",
    }
    principal = _principal(actor_id, permissions, firm_id=firm.id)
    session = factory()
    scope = _firm_scope(principal, session, firm.id)

    created = create_purchase_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            purchase_uom_id=purchase_uom_id,
            inventory_uom_id=inventory_uom_id,
            tax_profile_id=tax_profile_id,
            po_number="PO-API-001",
        ),
        scope,
        session,
    )
    assert created.data.po_number == "PO-API-001"

    listed = list_purchase_orders(
        scope=scope,
        page=1,
        page_size=20,
        search="PO-API",
        sort_by="created_at",
        sort_direction="desc",
        vendor_id=None,
        status_value=None,
        branch_id=None,
        warehouse_id=None,
        buyer_id=None,
        purchase_type=None,
        created_from=None,
        created_to=None,
        include_deleted=False,
        db=session,
    )
    assert listed.pagination.total_records == 1

    fetched = get_purchase_order(created.data.id, scope, False, session)
    assert fetched.data.id == created.data.id

    updated = update_purchase_order(
        created.data.id,
        PurchaseOrderUpdate.model_validate(
            {
                key: value
                for key, value in _purchase_data(
                    branch_id=branch.id,
                    warehouse_id=warehouse.id,
                    vendor_id=vendor.id,
                    product_id=product.id,
                    purchase_uom_id=purchase_uom_id,
                    inventory_uom_id=inventory_uom_id,
                    tax_profile_id=tax_profile_id,
                    po_number="IGNORED",
                    status="APPROVED",
                    ordered_quantity="5",
                    free_quantity="0",
                    unit_price="11",
                    discount_percent="5",
                )
                .model_dump(mode="json")
                .items()
                if key != "po_number"
            }
        ),
        scope,
        session,
    )
    assert updated.data.status == PurchaseOrderStatus.APPROVED

    summary = purchase_summary(scope, session)
    assert summary.data.total == 1
    assert summary.data.open == 1

    history = purchase_order_history(created.data.id, scope, session)
    assert len(history.data) >= 2

    csv_import = asyncio.run(
        import_purchase_orders(
            scope=scope,
            db=session,
            format="csv",
            payload=None,
            file=UploadFile(
                filename="purchase.csv",
                file=BytesIO(
                    (
                        "PoNumber,BranchId,WarehouseId,VendorId,ProductId,PurchaseDate,"
                        "OrderedQty,UnitPrice,PurchaseUomId,InventoryUomId,TaxProfileId,Remarks\n"
                        f"PO-API-CSV-001,{branch.id},{warehouse.id},{vendor.id},{product.id},2026-08-02,"
                        f"3,9,{purchase_uom_id},{inventory_uom_id},{tax_profile_id},CSV import\n"
                    ).encode()
                ),
            ),
        )
    )
    assert csv_import.data[0].po_number == "PO-API-CSV-001"

    from openpyxl import Workbook, load_workbook

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(
        [
            "PoNumber",
            "BranchId",
            "WarehouseId",
            "VendorId",
            "ProductId",
            "PurchaseDate",
            "OrderedQty",
            "UnitPrice",
            "PurchaseUomId",
            "InventoryUomId",
            "TaxProfileId",
            "Remarks",
        ]
    )
    sheet.append(
        [
            "PO-API-XLSX-001",
            str(branch.id),
            str(warehouse.id),
            str(vendor.id),
            str(product.id),
            "2026-08-02",
            "4",
            "8",
            str(purchase_uom_id),
            str(inventory_uom_id),
            str(tax_profile_id),
            "XLSX import",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    xlsx_import = asyncio.run(
        import_purchase_orders(
            scope=scope,
            db=session,
            format="xlsx",
            payload=None,
            file=UploadFile(
                filename="purchase.xlsx",
                file=BytesIO(buffer.getvalue()),
            ),
        )
    )
    assert xlsx_import.data[0].po_number == "PO-API-XLSX-001"

    json_import = asyncio.run(
        import_purchase_orders(
            scope=scope,
            db=session,
            format="json",
            payload=PurchaseOrderImportRequest(
                records=[
                    _purchase_data(
                        branch_id=branch.id,
                        warehouse_id=warehouse.id,
                        vendor_id=vendor.id,
                        product_id=product.id,
                        purchase_uom_id=purchase_uom_id,
                        inventory_uom_id=inventory_uom_id,
                        tax_profile_id=tax_profile_id,
                        po_number="PO-API-JSON-001",
                    )
                ]
            ).model_dump_json(),
            file=None,
        )
    )
    assert json_import.data[0].po_number == "PO-API-JSON-001"

    csv_export = export_purchase_orders(scope, "csv", "PO-API", session)
    csv_bytes = asyncio.run(_read_stream(csv_export))
    csv_text = csv_bytes.decode("utf-8")
    assert (
        "PO Number,Date,Vendor ID,Branch ID,Warehouse ID,Status,Subtotal,Tax Total,Grand Total"
        in csv_text
    )
    assert "PO-API-001" in csv_text

    xlsx_export = export_purchase_orders(scope, "xlsx", "PO-API", session)
    workbook_bytes = asyncio.run(_read_stream(xlsx_export))
    exported = load_workbook(filename=BytesIO(workbook_bytes), read_only=True)
    exported_sheet = exported.active
    assert exported_sheet is not None
    rows = list(exported_sheet.iter_rows(values_only=True))
    assert rows[0][0] == "PO Number"
    assert any(row[0] == "PO-API-001" for row in rows[1:])

    cancelled = create_purchase_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            tax_profile_id=tax_profile_id,
            po_number="PO-API-CANCEL-001",
        ),
        scope,
        session,
    )
    cancelled_response = cancel_purchase_order(
        cancelled.data.id,
        ActionReasonRequest(reason="User cancelled"),
        scope,
        session,
    )
    assert cancelled_response.data.status == PurchaseOrderStatus.CANCELLED

    closed = create_purchase_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            tax_profile_id=tax_profile_id,
            po_number="PO-API-CLOSE-001",
            status="APPROVED",
        ),
        scope,
        session,
    )
    closed_response = close_purchase_order(
        closed.data.id,
        ActionReasonRequest(reason="Closed after validation"),
        scope,
        session,
    )
    assert closed_response.data.status == PurchaseOrderStatus.CLOSED

    deleted = create_purchase_order(
        _purchase_data(
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            vendor_id=vendor.id,
            product_id=product.id,
            tax_profile_id=tax_profile_id,
            po_number="PO-API-DEL-001",
        ),
        scope,
        session,
    )
    delete_purchase_order(deleted.data.id, scope, session)
    restored = restore_purchase_order(deleted.data.id, scope, session)
    assert restored.data.is_deleted is False

    with pytest.raises(AuthorizationError):
        require_permission("PURCHASE_DELETE")(
            _principal(actor_id, {"PURCHASE_VIEW"}, firm_id=firm.id)
        )


def test_generated_purchase_order_number_carries_firm_branch_and_year() -> None:
    """Auto-generated numbers keep their composed shape.

    Every purchase test supplied an explicit po_number, so the generated format
    was asserted nowhere. Extracting the shared document base silently dropped
    the company code from it, and nothing failed.
    """
    session = _session_factory()()
    actor_id = uuid4()
    firm = _firm(session, "NUM")
    branch = _branch(session, firm_id=firm.id, actor_id=actor_id)
    warehouse = _warehouse(
        session, firm_id=firm.id, branch_id=branch.id, actor_id=actor_id
    )
    vendor = _vendor(session, firm_id=firm.id, actor_id=actor_id)
    product = _product(session, firm_id=firm.id, actor_id=actor_id)

    row = PurchaseService(session).create_order(
        PurchaseOrderCreate(
            vendor_id=vendor.id,
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            purchase_date=date(2026, 8, 4),
            status=PurchaseOrderStatus.DRAFT,
            lines=[
                {
                    "product_id": str(product.id),
                    "ordered_quantity": "5",
                    "unit_price": "10",
                }
            ],
        ),
        firm_id=firm.id,
        actor_id=actor_id,
    )

    # prefix - company - branch - financial year - sequence
    assert row.po_number == f"PO-{firm.code}-{branch.code}-2026-2027-000001"


def test_editing_a_purchase_order_keeps_its_line_identities() -> None:
    """Purchase order line ids survive an edit.

    Goods receipts and purchase invoices record which order line they came from
    in source_document_line_id, a bare UUID with no foreign key. Re-inserting
    lines on every save left those references pointing at rows that no longer
    existed.
    """
    session = _session_factory()()
    actor_id = uuid4()
    firm = _firm(session, "EDIT")
    branch = _branch(session, firm_id=firm.id, actor_id=actor_id)
    warehouse = _warehouse(
        session, firm_id=firm.id, branch_id=branch.id, actor_id=actor_id
    )
    vendor = _vendor(session, firm_id=firm.id, actor_id=actor_id)
    product = _product(session, firm_id=firm.id, actor_id=actor_id)
    service = PurchaseService(session)

    def _payload(quantity: str) -> PurchaseOrderCreate:
        return PurchaseOrderCreate(
            vendor_id=vendor.id,
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            purchase_date=date(2026, 8, 4),
            status=PurchaseOrderStatus.DRAFT,
            lines=[
                {
                    "product_id": str(product.id),
                    "ordered_quantity": quantity,
                    "unit_price": "10",
                }
            ],
        )

    order = service.create_order(_payload("5"), firm_id=firm.id, actor_id=actor_id)
    before = session.scalar(
        select(PurchaseOrderLine.id).where(
            PurchaseOrderLine.purchase_order_id == order.id
        )
    )
    assert before is not None

    service.update_order(
        order.id,
        PurchaseOrderUpdate(**_payload("8").model_dump(exclude={"po_number"})),
        firm_scope=firm.id,
        actor_id=actor_id,
    )
    rows = session.scalars(
        select(PurchaseOrderLine).where(PurchaseOrderLine.purchase_order_id == order.id)
    ).all()
    assert len(rows) == 1
    assert rows[0].id == before, "the line must keep its identity across an edit"
    assert rows[0].ordered_quantity == Decimal("8.0000")


def test_an_order_with_receipts_against_it_cannot_be_deleted() -> None:
    """Cancelling a received order was refused; deleting one was not.

    A goods receipt records the purchase_order_id it came from, so removing the
    order leaves the receipt pointing at a document no listing shows -- and
    delete is the more destructive of the two operations.
    """
    session = _session_factory()()
    actor_id = uuid4()
    firm = _firm(session, "PODEL")
    branch = _branch(session, firm_id=firm.id, actor_id=actor_id)
    warehouse = _warehouse(
        session, firm_id=firm.id, branch_id=branch.id, actor_id=actor_id
    )
    vendor = _vendor(session, firm_id=firm.id, actor_id=actor_id)
    product = _product(session, firm_id=firm.id, actor_id=actor_id)
    service = PurchaseService(session)

    order = service.create_order(
        _purchase_data(
            vendor_id=vendor.id,
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            product_id=product.id,
        ),
        firm_id=firm.id,
        actor_id=actor_id,
    )

    session.add(
        GoodsReceipt(
            firm_id=firm.id,
            purchase_order_id=order.id,
            purchase_order_number=order.po_number,
            vendor_id=vendor.id,
            branch_id=branch.id,
            warehouse_id=warehouse.id,
            grn_number="GRN-PODEL-1",
            receipt_date=date(2026, 8, 5),
            status="DRAFT",
            created_by=actor_id,
            updated_by=actor_id,
        )
    )
    session.commit()

    with pytest.raises(ValidationError, match="Goods have been received"):
        service.delete_order(order.id, firm_scope=firm.id, actor_id=actor_id)
